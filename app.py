from __future__ import annotations

import base64
import cgi
import csv
import hashlib
import html
import io
import json
import os
import re
import secrets
import shutil
import sqlite3
import time
import urllib.parse
import urllib.error
import urllib.request
import zipfile
from datetime import datetime, timedelta
from http import cookies
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path

from reportlab.lib import colors
from reportlab.lib.pagesizes import A0, A1, A2, A3, A4, landscape, portrait
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
from pypdf import PageObject, PdfReader, PdfWriter, Transformation
from pypdf.generic import ContentStream
from PIL import Image


ROOT = Path(__file__).resolve().parent
DATA = ROOT / "data"
UPLOADS = DATA / "uploads"
OUTPUTS = DATA / "outputs"
BATCHES = DATA / "batches"
WORKSPACE_FILES = DATA / "workspace_files"
PROFILE_UPLOADS = DATA / "profile_analyses"
DB = DATA / "ske.db"
HOST = os.environ.get("HOST", "0.0.0.0")
PORT = int(os.environ.get("PORT", "10000"))

PVIT_BASE_URL = os.environ.get("MYPVIT_BASE_URL", "https://api.mypvit.pro").rstrip("/")
PVIT_MODE = os.environ.get("MYPVIT_MODE", "TEST")
PVIT_MERCHANT_SLUG = os.environ.get("MYPVIT_MERCHANT_SLUG", "MR_1778797178")
PVIT_SECRET = os.environ.get("MYPVIT_SECRET", "")
PVIT_ACCOUNT_CODE = os.environ.get("MYPVIT_ACCOUNT_CODE", "")
PVIT_REST_PATH = os.environ.get("MYPVIT_REST_PATH", "/v2/INFXDEQPC2CUCPUB/rest")
PVIT_LINK_PATH = os.environ.get("MYPVIT_LINK_PATH", "/LHHUPSB32IKL50UU/link")
PVIT_BALANCE_PATH = os.environ.get("MYPVIT_BALANCE_PATH", "/OY5WKD4P55VGSAES/balance")
PVIT_STATUS_PATH = os.environ.get("MYPVIT_STATUS_PATH", "/FT8EGPPYBMKNTVDE/status")
PVIT_KYC_PATH = os.environ.get("MYPVIT_KYC_PATH", "/v2/CMK2GBAWUCNPNUNU/kyc")
PVIT_FEES_PATH = os.environ.get("MYPVIT_FEES_PATH", "/v2/30U1DUMY8R3VLPCH/get-fees")
PVIT_QR_PATH = os.environ.get("MYPVIT_QR_PATH", "/v2/PJBEJ9PFZWZP464U/generate-qr-code")
PVIT_OPERATORS_PATH = os.environ.get("MYPVIT_OPERATORS_PATH", "/v2/JSOBJUIWFTUORS2S/get-operators")
PVIT_COUNTRIES_PATH = os.environ.get("MYPVIT_COUNTRIES_PATH", "/v2/QWD6RUY5AYML9EY/get-countries")
PVIT_HEALTH_PATH = os.environ.get("MYPVIT_HEALTH_PATH", "/Q9BA00XLPSEAA6JV/services/health")
PUBLIC_SITE_URL = os.environ.get("BTP_PUBLIC_URL", f"http://{HOST}:{PORT}").rstrip("/")
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "").strip()
OPENAI_MODEL = os.environ.get("OPENAI_MODEL", "gpt-4o-mini")
PVIT_SECRET_RECEIVER_TOKEN = os.environ.get("MYPVIT_SECRET_RECEIVER_TOKEN", "pvit-btp-b9a1c7a2759644a7b005").strip()
PVIT_SECRET_LEGACY_TOKEN = "btp-smart-tools-test"
PVIT_ACCEPTED_RECEIVER_TOKENS = {
    PVIT_SECRET_RECEIVER_TOKEN,
    "pvit-btp-b9a1c7a2759644a7b005",
    "pvit-b9a1c7a2759644a7b005",
    PVIT_SECRET_LEGACY_TOKEN,
}

PAYMENT_OFFERS = {
    "pdf": {"label": "Pack ponctuel", "amount": 1200, "display": "1 200 FCFA", "credits": 1, "subscription": "none", "days": 0, "description": "1 credit pour generer 1 PDF avec cartouche."},
    "discovery": {"label": "Pack decouverte", "amount": 5000, "display": "5 000 FCFA", "credits": 10, "subscription": "none", "days": 0, "description": "10 credits : 1 analyse de cartouche + 5 generations PDF."},
    "standard": {"label": "Pack standard", "amount": 10000, "display": "10 000 FCFA", "credits": 20, "subscription": "none", "days": 0, "description": "20 credits : 2 analyses de cartouches + 10 generations PDF."},
    "monthly": {"label": "Abonnement mensuel", "amount": 12000, "display": "12 000 FCFA", "credits": 30, "subscription": "monthly", "days": 31, "description": "30 credits/mois : cartouches, analyses IA et generations PDF."},
    "annual": {"label": "Abonnement annuel", "amount": 108000, "display": "108 000 FCFA", "credits": 450, "subscription": "annual", "days": 365, "description": "450 credits/an pour bureaux et entreprises BTP."},
    "enterprise": {"label": "Pack entreprise", "amount": 250000, "display": "250 000 FCFA", "credits": 1500, "subscription": "enterprise", "days": 365, "description": "1500 credits/an pour entreprises, bureaux d'etudes et grands volumes de cartouches."},
}

CREDIT_COST_PDF = 1
CREDIT_COST_TEMPLATE_ANALYSIS = 5
CREDIT_COST_TEMPLATE_RESCAN = 3

ADMIN_EMAIL = os.environ.get("BTP_ADMIN_EMAIL", "sessouedem15@gmail.com").strip()
ADMIN_PASSWORD = os.environ.get("BTP_ADMIN_PASSWORD", "").strip()

SESSIONS: dict[str, int] = {}
PAGE_FORMATS = {
    "A4 Portrait": portrait(A4),
    "A4 Paysage": landscape(A4),
    "A3 Portrait": portrait(A3),
    "A3 Paysage": landscape(A3),
    "A2 Portrait": portrait(A2),
    "A2 Paysage": landscape(A2),
    "A1 Portrait": portrait(A1),
    "A1 Paysage": landscape(A1),
    "A0 Portrait": portrait(A0),
    "A0 Paysage": landscape(A0),
}
THEMES = {
    "Bleu BTP": "#08213A",
    "Noir Pro": "#111827",
    "Rouge Entreprise": "#991B1B",
    "Vert Chantier": "#166534",
    "Gris Technique": "#374151",
}
PLAN_TYPE_OPTIONS = [
    "Plan de situation",
    "Plan de masse",
    "Plan de coffrage",
    "Plan de ferraillage",
    "Profil en travers",
    "Profil en long",
    "Profil en travers type",
    "Plan VRD",
    "Plan topographique",
    "Plan d'execution",
    "Plan beton",
    "Plan de route",
    "Plan de dallot",
    "Plan d'implantation",
    "Plan de fondations",
]
PROFILE_ANALYSIS_TYPES = [
    "Profil en long",
    "Profil en travers",
    "Profil en travers type",
    "Projet routier",
    "Plan de route",
    "Plan topographique",
    "Plan beton",
    "Plan ferraillage",
    "Export Covadis",
    "Fichier Excel / CSV",
    "Image de profil",
    "DWG / DXF",
]
SCALE_OPTIONS = [
    "1/20",
    "1/25",
    "1/50",
    "1/75",
    "1/100",
    "1/200",
    "1/250",
    "1/500",
    "1/1000",
    "1/2000",
    "1/5000",
]
REVISION_DESCRIPTION_OPTIONS = [
    "Premiere emission",
    "Emission pour controle",
    "Emission pour validation",
    "Emission pour execution",
    "Modification suivant observations",
    "Mise a jour quantites",
    "Mise a jour cartouche",
    "Bon pour execution",
]
REVISION_INDEX_OPTIONS = [
    "REV 00",
    "REV 01",
    "REV 02",
    "REV 03",
    "A",
    "B",
    "C",
    "D",
    "Indice 0",
    "Indice 1",
]
PLATFORM_TEMPLATES = {
    "platform_standard": {
        "name": "BTP Standard",
        "category": "Standard",
        "color": "#08213A",
        "description": "Modele clair, propre et polyvalent pour plans BTP courants.",
    },
    "platform_premium": {
        "name": "BTP Premium Bleu",
        "category": "Premium",
        "color": "#8A6A20",
        "description": "Cartouche prestige avec accents or, bandeaux forts et rendu haut de gamme.",
    },
    "platform_modern": {
        "name": "BTP Moderne Gris",
        "category": "Moderne",
        "color": "#334155",
        "description": "Style minimaliste, gris technique, plus leger et contemporain.",
    },
    "platform_topo": {
        "name": "BTP Topographie",
        "category": "Topographie",
        "color": "#0F766E",
        "description": "Identite verte topo avec repere lateral et lecture chantier.",
    },
    "platform_engineering": {
        "name": "BTP Ingenierie",
        "category": "Ingenierie",
        "color": "#1D4ED8",
        "description": "Style bleu ingenierie avec grille technique et cartouche structure.",
    },
}


def now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def datalist_html() -> str:
    plan_options = "".join(f'<option value="{html.escape(value)}">' for value in PLAN_TYPE_OPTIONS)
    scale_options = "".join(f'<option value="{html.escape(value)}">' for value in SCALE_OPTIONS)
    revision_options = "".join(f'<option value="{html.escape(value)}">' for value in REVISION_DESCRIPTION_OPTIONS)
    revision_index_options = "".join(f'<option value="{html.escape(value)}">' for value in REVISION_INDEX_OPTIONS)
    return f"""
    <datalist id="plan_type_options">{plan_options}</datalist>
    <datalist id="scale_options">{scale_options}</datalist>
    <datalist id="revision_index_options">{revision_index_options}</datalist>
    <datalist id="revision_description_options">{revision_options}</datalist>
    """


def cartouche_focus_panel() -> str:
    return f"""
    <div class="cartouche-focus">
      <div>
        <span class="pill">Service actif maintenant</span>
        <h2>Cartouches PDF professionnelles</h2>
        <p>Le site est concentre sur une chose principale : importer un plan PDF ou une image, appliquer une cartouche propre, ajouter logo, format, echelle, legende, tableaux, puis telecharger le PDF final.</p>
      </div>
      <div class="focus-cost">
        <b>Credits</b>
        <span>1 PDF = {CREDIT_COST_PDF} credit</span>
        <span>1 cartouche importee = {CREDIT_COST_TEMPLATE_ANALYSIS} credits une seule fois</span>
        <span>Le modele est analyse, stocke dans le compte et reutilisable ensuite.</span>
      </div>
    </div>
    """


def cartouche_steps_html() -> str:
    steps = [
        ("1", "Charger", "PDF, image ou logo de l'entreprise."),
        ("2", "Completer", "Projet, echelle, format, revision et modele."),
        ("3", "Verifier", "Apercu, controle qualite, puis PDF final."),
    ]
    return "<div class='steps'>" + "".join(
        f"<div class='step'><b>{num}</b><strong>{html.escape(title)}</strong><span>{html.escape(text)}</span></div>"
        for num, title, text in steps
    ) + "</div>"


def cartouche_quality_html(source_loaded: bool = False) -> str:
    source_state = "pret a verifier apres chargement" if not source_loaded else "fichier charge"
    checks = [
        ("Format", "Detection automatique A4/A3/A2/A1/A0 si le PDF le permet."),
        ("Echelle", "Lecture automatique quand l'echelle est presente dans le nom ou le texte du PDF."),
        ("Cartouche", "Le modele personnalise est analyse une seule fois, puis garde dans le compte."),
        ("Logo", "PNG/JPG accepte, ajuste automatiquement dans la cartouche."),
        ("Legende", "Lignes, couleurs et textes modifiables avant generation."),
        ("Credit", f"Verification avant debit : {CREDIT_COST_PDF} credit par PDF final."),
    ]
    rows = "".join(
        f"<li><b>{html.escape(title)}</b><span>{html.escape(text)}</span></li>"
        for title, text in checks
    )
    return f"""
    <div class="quality-card">
      <div class="row" style="justify-content:space-between">
        <h3>Controle qualite cartouche</h3>
        <span class="pill">{html.escape(source_state)}</span>
      </div>
      <ul>{rows}</ul>
      <p class="muted">Pour DWG/DXF, exporte le dessin en PDF depuis AutoCAD avant import afin de conserver le vrai rendu du plan.</p>
    </div>
    """


def download_url(name: str | Path) -> str:
    return "/download/" + urllib.parse.quote(Path(str(name)).name)


def guess_plan_metadata(source_path: Path | None) -> dict:
    if not source_path:
        return {}
    text = source_path.stem.replace("_", " ").replace("-", " ")
    if source_path.suffix.lower() == ".pdf":
        try:
            reader = PdfReader(str(source_path))
            extracted = (reader.pages[0].extract_text() or "")[:3000] if reader.pages else ""
            text = f"{text} {extracted}"
        except Exception:
            pass
    low = text.lower()
    plan_type = ""
    keyword_map = [
        (("ferraillage", "armature", "acier"), "Plan de ferraillage"),
        (("coffrage",), "Plan de coffrage"),
        (("fondation", "semelle"), "Plan de fondations"),
        (("profil en travers", "travers"), "Profil en travers"),
        (("profil en long", "longitudinal"), "Profil en long"),
        (("vrd", "reseau", "rÃ©seau", "assainissement"), "Plan VRD"),
        (("topo", "topographique", "implantation"), "Plan topographique"),
        (("route", "chaussÃ©e", "chaussee"), "Plan de route"),
        (("dallot", "dalot"), "Plan de dallot"),
        (("masse",), "Plan de masse"),
        (("situation",), "Plan de situation"),
        (("bÃ©ton", "beton", "radier", "voile"), "Plan beton"),
        (("execution", "exÃ©cution"), "Plan d'execution"),
    ]
    for keywords, value in keyword_map:
        if any(keyword in low for keyword in keywords):
            plan_type = value
            break
    scale = ""
    scale_denoms = "20|25|50|75|100|125|150|200|250|500|750|1000|1500|2000|2500|5000|10000"
    scale_patterns = [
        rf"\b1\s*[/:\-\s]\s*({scale_denoms})\b",
        rf"\b(?:ech|echelle|Ã©chelle|scale)\s*[:=\-]?\s*1\s*[/:\-\s]\s*({scale_denoms})\b",
        rf"\b(?:ech|echelle|Ã©chelle)\s*[:=\-]?\s*({scale_denoms})\b",
    ]
    match = None
    for pattern in scale_patterns:
        match = re.search(pattern, low)
        if match:
            break
    if match:
        scale = f"1/{match.group(1)}"
    return {"plan_type": plan_type, "scale": scale}


def detect_image_style_color(path: Path, fallback: str = "#08213A") -> str:
    try:
        with Image.open(path) as img:
            img = img.convert("RGB")
            img.thumbnail((160, 160))
            colors_count = img.getcolors(maxcolors=160 * 160) or []
        weighted: dict[tuple[int, int, int], int] = {}
        for count, (r, g, b) in colors_count:
            # Ignore white paper/background and very light scan noise.
            if r > 225 and g > 225 and b > 225:
                continue
            # Ignore near-black text lines so the dominant brand/accent color wins.
            if r < 35 and g < 35 and b < 35:
                continue
            key = (round(r / 16) * 16, round(g / 16) * 16, round(b / 16) * 16)
            weighted[key] = weighted.get(key, 0) + count
        if not weighted:
            return fallback
        r, g, b = max(weighted.items(), key=lambda item: item[1])[0]
        return f"#{max(0,min(255,r)):02X}{max(0,min(255,g)):02X}{max(0,min(255,b)):02X}"
    except Exception:
        return fallback


def ensure_dirs():
    for path in (DATA, UPLOADS, OUTPUTS, BATCHES, WORKSPACE_FILES, PROFILE_UPLOADS):
        path.mkdir(parents=True, exist_ok=True)


def db() -> sqlite3.Connection:
    con = sqlite3.connect(DB)
    con.row_factory = sqlite3.Row
    return con


def hash_password(password: str, salt: str | None = None) -> str:
    salt = salt or secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 120000).hex()
    return f"{salt}:{digest}"


def verify_password(password: str, stored: str) -> bool:
    try:
        salt, digest = stored.split(":", 1)
        return hash_password(password, salt).split(":", 1)[1] == digest
    except Exception:
        return False


def init_db():
    ensure_dirs()
    with db() as con:
        con.executescript(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                email TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                name TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'user',
                credits INTEGER NOT NULL DEFAULT 0,
                subscription TEXT NOT NULL DEFAULT 'none',
                subscription_until TEXT,
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS generations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                project TEXT NOT NULL,
                company TEXT NOT NULL,
                plan_type TEXT NOT NULL,
                scale TEXT NOT NULL,
                format_plan TEXT NOT NULL DEFAULT 'A4 Paysage',
                plan_number TEXT,
                source_file TEXT,
                output_file TEXT NOT NULL,
                status TEXT NOT NULL,
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS batches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                project TEXT NOT NULL,
                total_files INTEGER NOT NULL,
                output_zip TEXT NOT NULL,
                status TEXT NOT NULL,
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS payments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                offer TEXT NOT NULL,
                amount TEXT NOT NULL,
                method TEXT NOT NULL,
                phone TEXT,
                provider TEXT NOT NULL DEFAULT 'mypvit',
                provider_mode TEXT NOT NULL DEFAULT 'TEST',
                transaction_ref TEXT,
                payment_url TEXT,
                provider_response TEXT,
                status TEXT NOT NULL,
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS service_modules (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                category TEXT NOT NULL,
                status TEXT NOT NULL,
                description TEXT NOT NULL,
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS cartouche_templates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                name TEXT NOT NULL,
                source_type TEXT NOT NULL,
                source_file TEXT NOT NULL,
                theme_color TEXT NOT NULL DEFAULT '#08213A',
                ai_status TEXT NOT NULL DEFAULT 'preparation_locale',
                analysis_json TEXT,
                analysis_summary TEXT,
                paid_status TEXT NOT NULL DEFAULT 'test_gratuit',
                final_template_file TEXT,
                last_used_at TEXT,
                status TEXT NOT NULL DEFAULT 'active',
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS company_spaces (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                owner_user_id INTEGER NOT NULL,
                name TEXT NOT NULL,
                logo_file TEXT,
                storage_status TEXT NOT NULL DEFAULT 'local_render_test',
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS workspace_projects (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                company_id INTEGER,
                name TEXT NOT NULL,
                client TEXT,
                location TEXT,
                status TEXT NOT NULL DEFAULT 'actif',
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS workspace_documents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                project_id INTEGER,
                title TEXT NOT NULL,
                category TEXT NOT NULL,
                original_name TEXT NOT NULL,
                stored_file TEXT NOT NULL,
                mime_type TEXT,
                size_kb REAL,
                version_label TEXT NOT NULL DEFAULT 'V1',
                confidentiality TEXT NOT NULL DEFAULT 'prive',
                status TEXT NOT NULL DEFAULT 'stocke',
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS profile_analyses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                project TEXT,
                document_type TEXT NOT NULL,
                source_file TEXT NOT NULL,
                source_name TEXT NOT NULL,
                analysis_json TEXT,
                analysis_html TEXT,
                status TEXT NOT NULL,
                created_at TEXT NOT NULL
            );
            """
        )
        for alter in [
            "ALTER TABLE generations ADD COLUMN format_plan TEXT NOT NULL DEFAULT 'A4 Paysage'",
            "ALTER TABLE generations ADD COLUMN plan_number TEXT",
            "ALTER TABLE cartouche_templates ADD COLUMN ai_status TEXT NOT NULL DEFAULT 'preparation_locale'",
            "ALTER TABLE cartouche_templates ADD COLUMN analysis_json TEXT",
            "ALTER TABLE cartouche_templates ADD COLUMN analysis_summary TEXT",
            "ALTER TABLE cartouche_templates ADD COLUMN paid_status TEXT NOT NULL DEFAULT 'test_gratuit'",
            "ALTER TABLE cartouche_templates ADD COLUMN final_template_file TEXT",
            "ALTER TABLE cartouche_templates ADD COLUMN last_used_at TEXT",
            "ALTER TABLE company_spaces ADD COLUMN logo_file TEXT",
            "ALTER TABLE company_spaces ADD COLUMN storage_status TEXT NOT NULL DEFAULT 'local_render_test'",
            "ALTER TABLE workspace_documents ADD COLUMN version_label TEXT NOT NULL DEFAULT 'V1'",
            "ALTER TABLE workspace_documents ADD COLUMN confidentiality TEXT NOT NULL DEFAULT 'prive'",
            "ALTER TABLE payments ADD COLUMN phone TEXT",
            "ALTER TABLE payments ADD COLUMN provider TEXT NOT NULL DEFAULT 'mypvit'",
            "ALTER TABLE payments ADD COLUMN provider_mode TEXT NOT NULL DEFAULT 'TEST'",
            "ALTER TABLE payments ADD COLUMN transaction_ref TEXT",
            "ALTER TABLE payments ADD COLUMN payment_url TEXT",
            "ALTER TABLE payments ADD COLUMN provider_response TEXT",
            "ALTER TABLE payments ADD COLUMN offer_key TEXT",
            "ALTER TABLE payments ADD COLUMN credits INTEGER NOT NULL DEFAULT 0",
            "ALTER TABLE payments ADD COLUMN credited INTEGER NOT NULL DEFAULT 0",
        ]:
            try:
                con.execute(alter)
            except sqlite3.OperationalError:
                pass
        if ADMIN_EMAIL and ADMIN_PASSWORD:
            exists = con.execute("SELECT id FROM users WHERE email=?", (ADMIN_EMAIL,)).fetchone()
            if not exists:
                con.execute(
                    "INSERT INTO users(email,password_hash,name,role,credits,subscription,subscription_until,created_at) VALUES(?,?,?,?,?,?,?,?)",
                    (ADMIN_EMAIL, hash_password(ADMIN_PASSWORD), "Administrateur BTP Smart Tools", "admin", 9999, "admin_free", (datetime.now() + timedelta(days=365)).strftime("%Y-%m-%d"), now()),
                )
            else:
                con.execute(
                    "UPDATE users SET password_hash=?, role='admin', credits=9999, subscription='admin_free' WHERE email=?",
                    (hash_password(ADMIN_PASSWORD), ADMIN_EMAIL),
                )
        modules = [
            ("Cartouches automatiques", "Plans", "actif", "Generation PDF avec cartouches, cadres, legendes et informations projet."),
            ("Cartouches personnalisees", "Plans", "actif", "Import de modeles entreprise, preparation d'analyse locale et structure prete pour IA vision."),
            ("Assistant support", "Aide", "actif", "Accueil, guide d'utilisation, questions frequentes et support de premier niveau."),
            ("Analyse profils routiers", "IA BTP", "a_venir", "Analyse de profils en long, profils en travers, plans routiers et explications techniques."),
            ("Analyse plans techniques", "IA BTP", "a_venir", "Lecture intelligente des plans, explications et annotations techniques."),
            ("Controle topo chantier", "Topographie", "a_venir", "Fiches voiles, radiers, ecarts et rapports PDF."),
            ("Dimensionnement BTP", "Bureau etudes", "a_venir", "Outils futurs pour routes, batiments et murs de soutenement."),
        ]
        for name, category, status, description in modules:
            exists = con.execute("SELECT id FROM service_modules WHERE name=?", (name,)).fetchone()
            if not exists:
                con.execute(
                    "INSERT INTO service_modules(name,category,status,description,created_at) VALUES(?,?,?,?,?)",
                    (name, category, status, description, now()),
                )
            else:
                con.execute(
                    "UPDATE service_modules SET category=?, status=?, description=? WHERE name=?",
                    (category, status, description, name),
                )
        exists = con.execute("SELECT value FROM settings WHERE key='next_plan_number'").fetchone()
        if not exists:
                con.execute("INSERT INTO settings(key,value) VALUES('next_plan_number','1')")


def next_plan_number(con: sqlite3.Connection) -> str:
    row = con.execute("SELECT value FROM settings WHERE key='next_plan_number'").fetchone()
    value = int(row["value"] if row else "1")
    con.execute(
        "INSERT INTO settings(key,value) VALUES('next_plan_number',?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (str(value + 1),),
    )
    return f"BTP-{value:03d}"


def detect_pdf_format(path: Path | None) -> str:
    if not path or not path.exists() or path.suffix.lower() != ".pdf":
        return "A4 Paysage"
    try:
        reader = PdfReader(str(path))
        page = reader.pages[0]
        w = float(page.mediabox.width)
        h = float(page.mediabox.height)
        orientation = "Paysage" if w >= h else "Portrait"
        short, long = sorted((w, h))
        sizes = {
            "A4": sorted(A4),
            "A3": sorted(A3),
            "A2": sorted(A2),
            "A1": sorted(A1),
            "A0": sorted(A0),
        }
        best = min(sizes, key=lambda name: abs(sizes[name][0] - short) + abs(sizes[name][1] - long))
        return f"{best} {orientation}"
    except Exception:
        return "A4 Paysage"


def resolve_format(format_choice: str, source_path: Path | None) -> tuple[str, tuple[float, float]]:
    detected = detect_pdf_format(source_path)
    final_name = detected if format_choice == "Automatique" else format_choice
    return final_name, PAGE_FORMATS.get(final_name, landscape(A4))


def get_template(template_id: str | int | None) -> sqlite3.Row | None:
    if not template_id:
        return None
    try:
        tid = int(template_id)
    except Exception:
        return None
    with db() as con:
        return con.execute("SELECT * FROM cartouche_templates WHERE id=? AND status='active'", (tid,)).fetchone()


def get_or_create_company_space(con: sqlite3.Connection, user: sqlite3.Row) -> sqlite3.Row:
    space = con.execute("SELECT * FROM company_spaces WHERE owner_user_id=? ORDER BY id LIMIT 1", (user["id"],)).fetchone()
    if space:
        return space
    con.execute(
        "INSERT INTO company_spaces(owner_user_id,name,logo_file,storage_status,created_at) VALUES(?,?,?,?,?)",
        (user["id"], user["name"] or "Entreprise BTP", "", "local_render_test", now()),
    )
    return con.execute("SELECT * FROM company_spaces WHERE owner_user_id=? ORDER BY id DESC LIMIT 1", (user["id"],)).fetchone()


def classify_document(filename: str, fallback: str = "Autre") -> str:
    name = filename.lower()
    suffix = Path(filename).suffix.lower()
    if suffix == ".pdf":
        if any(word in name for word in ("dao", "appel", "offre", "ccap", "cctp", "dqe", "bpu")):
            return "DAO / Administratif"
        if any(word in name for word in ("rapport", "pv", "reunion", "journal")):
            return "Rapport chantier"
        return "Plan PDF"
    if suffix in (".dwg", ".dxf"):
        return "DWG / AutoCAD"
    if suffix in (".png", ".jpg", ".jpeg"):
        if any(word in name for word in ("cartouche", "modele", "template")):
            return "Cartouche"
        return "Image"
    if suffix in (".doc", ".docx", ".xls", ".xlsx", ".txt"):
        return "DAO / Administratif"
    if suffix == ".zip":
        return "Archive ZIP"
    return fallback


def image_data_url_for_ai(path: Path) -> str:
    with Image.open(path) as img:
        img = img.convert("RGB")
        img.thumbnail((1600, 1600))
        buffer = io.BytesIO()
        img.save(buffer, format="JPEG", quality=88, optimize=True)
    encoded = base64.b64encode(buffer.getvalue()).decode("ascii")
    return f"data:image/jpeg;base64,{encoded}"


def extract_output_text(response: dict) -> str:
    if isinstance(response.get("output_text"), str):
        return response["output_text"]
    parts: list[str] = []
    for item in response.get("output", []) or []:
        for content in item.get("content", []) or []:
            text = content.get("text") or content.get("output_text")
            if isinstance(text, str):
                parts.append(text)
    return "\n".join(parts).strip()


def openai_template_analysis(path: Path, local_analysis: dict) -> dict | None:
    if not OPENAI_API_KEY:
        return None

    prompt = """
Tu es un assistant BTP specialise dans les cartouches de plans techniques.
Analyse le modele de cartouche fourni et retourne uniquement un JSON valide.
Objectif : transformer ce modele en cartouche dynamique reutilisable.
Le JSON doit contenir :
- engine
- model
- confidence de 0 a 1
- visual_style : description courte du style
- orientation
- dominant_color
- secondary_colors
- detected_elements : liste des elements detectes
- editable_fields : champs dynamiques a prevoir
- layout_zones : liste d'objets avec name, role, position_estimee, contenu_attendu
- recommendations : conseils pour reconstruire proprement la cartouche
Ne mets aucun texte hors JSON.
"""
    content: list[dict] = [
        {"type": "input_text", "text": prompt},
        {"type": "input_text", "text": "Analyse locale disponible : " + json.dumps(local_analysis, ensure_ascii=False)},
    ]
    if path.suffix.lower() in (".png", ".jpg", ".jpeg"):
        content.append({"type": "input_image", "image_url": image_data_url_for_ai(path)})
    elif path.suffix.lower() == ".pdf":
        pdf_text = ""
        try:
            reader = PdfReader(str(path))
            pdf_text = "\n".join((page.extract_text() or "") for page in reader.pages[:2])[:6000]
        except Exception:
            pdf_text = ""
        content.append(
            {
                "type": "input_text",
                "text": "Le fichier importe est un PDF. Texte extrait des premieres pages :\n" + (pdf_text or "Aucun texte exploitable extrait."),
            }
        )

    payload = {
        "model": OPENAI_MODEL,
        "input": [{"role": "user", "content": content}],
        "temperature": 0.2,
        "max_output_tokens": 1800,
    }
    request = urllib.request.Request(
        "https://api.openai.com/v1/responses",
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {OPENAI_API_KEY}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=45) as response:
        raw = response.read().decode("utf-8", errors="replace")
    api_response = json.loads(raw)
    output_text = extract_output_text(api_response)
    match = re.search(r"\{.*\}", output_text, re.S)
    if not match:
        raise ValueError("L'IA n'a pas retourne de JSON exploitable.")
    analysis = json.loads(match.group(0))
    analysis["engine"] = analysis.get("engine") or "openai_vision"
    analysis["model"] = OPENAI_MODEL
    analysis["source_file"] = path.name
    return analysis


def build_template_analysis(path: Path, source_type: str, theme_color: str) -> dict:
    analysis = {
        "engine": "local_preparation",
        "ai_ready": True,
        "source_type": source_type,
        "file_name": path.name,
        "file_size_kb": round(path.stat().st_size / 1024, 1) if path.exists() else 0,
        "dominant_color": theme_color,
        "detected_elements": [
            "cadre principal probable",
            "zone logo a confirmer",
            "zones textes a rendre dynamiques",
            "tableaux et revisions a verifier",
        ],
        "editable_fields": [
            "logo",
            "entreprise",
            "projet",
            "localisation",
            "type_plan",
            "echelle",
            "date",
            "numero_plan",
            "revision",
            "legende",
            "tableau_elements",
            "signatures",
        ],
        "next_step": "Connecter une API vision/OCR pour detecter automatiquement les cadres, textes et zones exactes.",
    }
    if path.suffix.lower() in (".png", ".jpg", ".jpeg"):
        try:
            with Image.open(path) as img:
                analysis["image_width"] = img.width
                analysis["image_height"] = img.height
                analysis["orientation"] = "Paysage" if img.width >= img.height else "Portrait"
        except Exception:
            analysis["image_note"] = "Image enregistree, dimensions non detectees."
    elif path.suffix.lower() == ".pdf":
        try:
            reader = PdfReader(str(path))
            page = reader.pages[0]
            box = page.mediabox
            analysis["pdf_pages"] = len(reader.pages)
            analysis["pdf_width"] = round(float(box.width), 1)
            analysis["pdf_height"] = round(float(box.height), 1)
            analysis["orientation"] = "Paysage" if float(box.width) >= float(box.height) else "Portrait"
        except Exception:
            analysis["pdf_note"] = "PDF enregistre, dimensions non detectees."
    try:
        ai_analysis = openai_template_analysis(path, analysis)
        if ai_analysis:
            ai_analysis.setdefault("source_type", source_type)
            ai_analysis.setdefault("file_name", path.name)
            ai_analysis.setdefault("file_size_kb", analysis.get("file_size_kb", 0))
            ai_analysis.setdefault("dominant_color", analysis.get("dominant_color", theme_color))
            ai_analysis["local_fallback"] = analysis
            return ai_analysis
    except Exception as exc:
        analysis["engine"] = "local_preparation_ai_failed"
        analysis["ai_error"] = str(exc)
    return analysis


def summarize_template_analysis(analysis: dict) -> str:
    orientation = analysis.get("orientation", "a confirmer")
    size = analysis.get("file_size_kb", 0)
    color = analysis.get("dominant_color", "#08213A")
    if str(analysis.get("engine", "")).startswith("openai"):
        confidence = analysis.get("confidence", "a confirmer")
        return f"Analyse IA terminee : orientation {orientation}, couleur principale {color}, confiance {confidence}, fichier {size} Ko."
    if analysis.get("ai_error"):
        return f"Analyse locale terminee, IA non finalisee : {analysis.get('ai_error')}. Orientation {orientation}, couleur {color}."
    return f"Preparation locale terminee : orientation {orientation}, couleur principale {color}, fichier {size} Ko."


def extract_profile_source_text(path: Path) -> str:
    suffix = path.suffix.lower()
    parts: list[str] = [f"Nom du fichier : {path.name}", f"Format : {suffix or 'sans extension'}"]
    if suffix == ".pdf":
        try:
            reader = PdfReader(str(path))
            for index, page in enumerate(reader.pages[:4], start=1):
                text = page.extract_text() or ""
                if text.strip():
                    parts.append(f"--- Page {index} ---\n{text[:3500]}")
        except Exception as exc:
            parts.append(f"Lecture PDF limitee : {exc}")
    elif suffix in (".txt", ".csv"):
        try:
            raw = path.read_text(encoding="utf-8", errors="replace")
        except Exception:
            raw = path.read_text(errors="replace")
        if suffix == ".csv":
            sample = []
            try:
                with path.open("r", encoding="utf-8", errors="replace", newline="") as f:
                    dialect = csv.Sniffer().sniff(f.read(2048))
                    f.seek(0)
                    reader = csv.reader(f, dialect)
                    for row in list(reader)[:25]:
                        sample.append(" | ".join(row))
                raw = "\n".join(sample) or raw
            except Exception:
                pass
        parts.append(raw[:12000])
    elif suffix == ".xlsx":
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames[:3]:
                sheet = wb[sheet_name]
                rows = []
                for row in sheet.iter_rows(max_row=35, max_col=12, values_only=True):
                    values = ["" if value is None else str(value) for value in row]
                    if any(values):
                        rows.append(" | ".join(values))
                if rows:
                    parts.append(f"--- Feuille {sheet_name} ---\n" + "\n".join(rows))
        except Exception as exc:
            parts.append(f"Excel charge, extraction avancee non disponible : {exc}")
    elif suffix in (".png", ".jpg", ".jpeg"):
        try:
            with Image.open(path) as img:
                parts.append(f"Image lisible : {img.width} x {img.height} pixels. L'analyse visuelle IA peut detecter les textes et couches si l'API est disponible.")
        except Exception as exc:
            parts.append(f"Image non lisible localement : {exc}")
    elif suffix in (".dwg", ".dxf"):
        if suffix == ".dxf":
            try:
                raw = path.read_text(encoding="utf-8", errors="replace")
                parts.append("DXF lisible en texte technique. Extrait brut :\n" + raw[:9000])
            except Exception:
                parts.append("Fichier DXF charge. La lecture complete demande un parseur CAO avance.")
        else:
            parts.append("Fichier DWG AutoCAD/Covadis charge. Le DWG n'est pas lisible directement par un navigateur web sans moteur CAO. Le fichier est conserve pour conversion/analyse avancee.")
    elif suffix == ".zip":
        try:
            with zipfile.ZipFile(path, "r") as zf:
                names = zf.namelist()[:80]
            parts.append("Archive ZIP detectee. Contenu partiel :\n" + "\n".join(names))
        except Exception as exc:
            parts.append(f"Archive ZIP non lisible : {exc}")
    return "\n\n".join(parts)[:16000]


def profile_keywords(text: str) -> dict:
    low = text.lower()
    groups = {
        "axes": ["axe", "alignement", "pk", "chainage", "abscisse"],
        "chaussée": ["chaussee", "chaussée", "voie", "largeur", "plateforme"],
        "bordures_caniveaux": ["bordure", "caniveau", "cunette", "fil d'eau", "fild'eau"],
        "couches": ["pst", "couche de forme", "fondation", "base", "bitume", "grave", "0/25", "enrobe", "béton bitumineux", "beton bitumineux"],
        "niveaux": ["altitude", "z=", "tn", "projet", "cote", "niveau"],
        "pentes": ["pente", "%", "devers", "rampe"],
        "terrassement": ["deblai", "déblai", "remblai", "purge", "terrassement", "volume"],
        "ouvrages": ["dalot", "buse", "regard", "mur", "voile", "radier"],
    }
    return {name: any(keyword in low for keyword in keywords) for name, keywords in groups.items()}


def local_profile_analysis(path: Path, document_type: str, project: str, source_text: str, ai_error: str = "") -> dict:
    detected = profile_keywords(source_text)
    present = [name.replace("_", " ") for name, ok in detected.items() if ok]
    suffix = path.suffix.lower()
    notes = []
    if suffix in (".dwg", ".dxf"):
        notes.append("Le fichier CAO est stocke. Pour lire toutes les entites DWG avec precision, il faudra ajouter un moteur CAO ou exporter le plan en PDF/DXF exploitable.")
    if suffix in (".png", ".jpg", ".jpeg") and not OPENAI_API_KEY:
        notes.append("L'image est prete, mais la lecture visuelle intelligente demande une cle IA active et du quota.")
    if not present:
        present = ["structure technique a confirmer", "annotations a verifier", "zones du projet a lire sur le plan"]
    summary = (
        f"Analyse locale du document {document_type}. Le systeme a conserve le fichier et prepare une lecture technique. "
        f"Elements reperes : {', '.join(present)}."
    )
    return {
        "engine": "local_profile_analyzer",
        "status": "analyse_locale",
        "project": project,
        "document_type": document_type,
        "source_file": path.name,
        "summary": summary,
        "detected_elements": present,
        "technical_explanation": [
            "Verifier l'axe du projet et les PK/chainages si presents.",
            "Comparer les cotes TN et projet pour identifier les zones de remblai/deblai.",
            "Controler les couches de chaussee : PST, forme, fondation, base et revetement.",
            "Controler les bordures, caniveaux, ouvrages hydrauliques et raccordements.",
        ],
        "layers": ["PST", "Couche de forme", "Fondation", "Couche de base", "Revetement bitumineux"],
        "risks": [
            "Les altitudes et epaisseurs doivent etre confirmees sur le plan original.",
            "Une lecture IA/OCR avancee donnera une interpretation plus precise des textes et symboles.",
        ] + notes,
        "recommendations": [
            "Importer de preference un PDF lisible ou une image nette pour une meilleure analyse.",
            "Ajouter le fichier Covadis/Excel si les altitudes, profils ou quantites existent dans un tableau.",
            "Garder le resultat dans l'historique afin de comparer les versions du projet.",
        ],
        "ai_error": ai_error,
    }


def technical_file_capability(path: Path | None, source_name: str = "") -> dict:
    suffix = (path.suffix if path else Path(source_name).suffix).lower()
    direct = suffix in (".pdf", ".png", ".jpg", ".jpeg", ".txt", ".csv", ".xlsx")
    if suffix in (".pdf",):
        return {
            "format": "PDF",
            "status": "lisible_directement",
            "label": "PDF lisible directement",
            "message": "Le plan PDF peut etre affiche dans le navigateur, zoome et analyse.",
            "next_steps": ["Analyse OCR/IA possible", "Apercu exact du PDF", "Conservation du fichier original"],
        }
    if suffix in (".png", ".jpg", ".jpeg"):
        return {
            "format": "Image / scan",
            "status": "lisible_directement",
            "label": "Image lisible directement",
            "message": "Le plan ou profil image peut etre affiche et envoye a l'analyse visuelle IA.",
            "next_steps": ["Zoom visuel", "Detection des textes si quota IA disponible", "Annotations futures sur image"],
        }
    if suffix in (".xlsx", ".xls"):
        return {
            "format": "Excel",
            "status": "tableau_extrait",
            "label": "Excel lisible en tableau",
            "message": "Le systeme extrait les premieres lignes du classeur pour controler profils, cotes et coordonnees.",
            "next_steps": ["Lecture des colonnes", "Detection X/Y/Z si possible", "Analyse des chainages et altitudes"],
        }
    if suffix in (".txt", ".csv", ".pts", ".asc", ".xyz"):
        return {
            "format": "Topo texte",
            "status": "texte_lisible",
            "label": "Fichier texte lisible",
            "message": "Le contenu brut est affiche et peut etre nettoye/analyse ligne par ligne.",
            "next_steps": ["Controle format", "Detection points", "Conversion Covadis/station possible"],
        }
    if suffix == ".dxf":
        return {
            "format": "DXF",
            "status": "lecture_partielle",
            "label": "DXF partiellement exploitable",
            "message": "Le DXF peut etre inspecte comme fichier technique. Un rendu graphique exact demandera un moteur CAO.",
            "next_steps": ["Extraction future des calques", "Conversion future en SVG/PDF", "Analyse IA apres rendu image"],
        }
    if suffix == ".dwg":
        return {
            "format": "DWG",
            "status": "conversion_requise",
            "label": "DWG : conversion requise",
            "message": "Le DWG ne s'affiche pas directement dans un navigateur. Pour un rendu exact, il faudra connecter un moteur CAO comme Autodesk Platform Services, ODA/Teigha ou une conversion AutoCAD serveur.",
            "next_steps": ["Conserver le DWG original", "Convertir en PDF/image haute qualite", "Analyser le rendu converti", "Ajouter plus tard un mini viewer CAO"],
        }
    if suffix == ".zip":
        return {
            "format": "ZIP",
            "status": "archive_detectee",
            "label": "Archive ZIP detectee",
            "message": "Le systeme peut lister le contenu et preparer les fichiers internes. L'analyse automatique complete dependra des formats contenus dans le ZIP.",
            "next_steps": ["Lister fichiers internes", "Extraire PDF/images/Excel", "Preparer DWG/DXF pour conversion"],
        }
    return {
        "format": suffix or "Inconnu",
        "status": "format_a_confirmer",
        "label": "Format a confirmer",
        "message": "Le fichier est stocke mais ce format demande une regle de lecture supplementaire.",
        "next_steps": ["Conserver le fichier", "Identifier le format", "Ajouter un convertisseur adapte"],
    }


def render_compatibility_panel(capability: dict) -> str:
    status = html.escape(str(capability.get("status", "")))
    steps = capability.get("next_steps") or []
    step_html = "".join(f"<li>{html.escape(str(step))}</li>" for step in steps)
    return f"""
    <div class="card">
      <span class="pill">Format : {html.escape(str(capability.get('format','')))}</span>
      <span class="pill">Statut : {status}</span>
      <h3>{html.escape(str(capability.get('label','Compatibilite fichier')))}</h3>
      <p>{html.escape(str(capability.get('message','')))}</p>
      <ul>{step_html}</ul>
    </div>
    """


def openai_profile_analysis(path: Path, document_type: str, project: str, source_text: str) -> dict | None:
    if not OPENAI_API_KEY:
        return None
    prompt = """
Tu es un ingenieur routier/topographe experimente.
Analyse le document fourni comme un assistant technique qui explique un profil ou projet routier a un chef de projet.
Retourne uniquement un JSON valide avec :
- engine
- confidence de 0 a 1
- summary : resume simple
- detected_elements : liste d'elements visibles ou probables
- axes : explication de l'axe, PK, alignements
- road_sections : largeurs, chaussee, trottoirs, bordures, caniveaux si detectes
- layers : couches detectees avec role et epaisseur si visible
- altitudes_slopes : altitudes, pentes, devers, raccordements si detectes
- earthworks : remblai/deblai/purge/terrassement si detectes
- risks : zones critiques ou points a verifier
- recommendations : conseils pratiques chantier
- simple_explanation : explication claire pour quelqu'un qui ne comprend pas bien le plan
Si une information n'est pas visible, indique qu'elle est a confirmer au lieu d'inventer.
Ne mets aucun texte hors JSON.
"""
    content: list[dict] = [
        {"type": "input_text", "text": prompt},
        {"type": "input_text", "text": f"Projet : {project or 'Non renseigne'}\nType de document : {document_type}\nTexte/metadonnees extraits :\n{source_text[:14000]}"},
    ]
    if path.suffix.lower() in (".png", ".jpg", ".jpeg"):
        content.append({"type": "input_image", "image_url": image_data_url_for_ai(path)})
    payload = {
        "model": OPENAI_MODEL,
        "input": [{"role": "user", "content": content}],
        "temperature": 0.15,
        "max_output_tokens": 2200,
    }
    request = urllib.request.Request(
        "https://api.openai.com/v1/responses",
        data=json.dumps(payload).encode("utf-8"),
        headers={"Authorization": f"Bearer {OPENAI_API_KEY}", "Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=55) as response:
        raw = response.read().decode("utf-8", errors="replace")
    api_response = json.loads(raw)
    output_text = extract_output_text(api_response)
    match = re.search(r"\{.*\}", output_text, re.S)
    if not match:
        raise ValueError("L'IA n'a pas retourne de JSON exploitable.")
    analysis = json.loads(match.group(0))
    analysis["engine"] = analysis.get("engine") or "openai_profile_analyzer"
    analysis["model"] = OPENAI_MODEL
    analysis["project"] = project
    analysis["document_type"] = document_type
    analysis["source_file"] = path.name
    return analysis


def build_profile_analysis(path: Path, document_type: str, project: str) -> dict:
    source_text = extract_profile_source_text(path)
    try:
        ai_analysis = openai_profile_analysis(path, document_type, project, source_text)
        if ai_analysis:
            return ai_analysis
    except Exception as exc:
        return local_profile_analysis(path, document_type, project, source_text, str(exc))
    return local_profile_analysis(path, document_type, project, source_text)


def profile_list_html(title: str, value) -> str:
    if not value:
        return ""
    if isinstance(value, str):
        items = [value]
    elif isinstance(value, dict):
        items = [f"{k} : {v}" for k, v in value.items()]
    else:
        items = []
        for item in value:
            if isinstance(item, dict):
                label = item.get("name") or item.get("element") or item.get("role") or item.get("type") or "Element"
                detail = ", ".join(f"{k}: {v}" for k, v in item.items() if v and k not in ("name", "element", "role", "type"))
                items.append(f"{label} - {detail}" if detail else str(label))
            else:
                items.append(str(item))
    lis = "".join(f"<li>{html.escape(str(item))}</li>" for item in items if str(item).strip())
    if not lis:
        return ""
    return f"<div class='card'><h3>{html.escape(title)}</h3><ul>{lis}</ul></div>"


def render_profile_analysis_result(analysis: dict) -> str:
    summary = html.escape(str(analysis.get("summary") or analysis.get("simple_explanation") or "Analyse preparee."))
    confidence = analysis.get("confidence", "local")
    engine = analysis.get("engine", "analyse_locale")
    sections = [
        profile_list_html("Elements detectes", analysis.get("detected_elements")),
        profile_list_html("Axes / PK / alignements", analysis.get("axes")),
        profile_list_html("Chaussee et sections", analysis.get("road_sections")),
        profile_list_html("Couches de structure", analysis.get("layers")),
        profile_list_html("Altitudes, pentes et raccordements", analysis.get("altitudes_slopes")),
        profile_list_html("Terrassements remblai/deblai", analysis.get("earthworks")),
        profile_list_html("Zones critiques", analysis.get("risks")),
        profile_list_html("Recommandations chantier", analysis.get("recommendations")),
    ]
    if analysis.get("ai_error"):
        sections.append(f"<p class='alert'>IA non finalisee : {html.escape(str(analysis.get('ai_error')))}</p>")
    return f"""
    <div class="card">
      <span class="pill">Moteur : {html.escape(str(engine))}</span>
      <span class="pill">Confiance : {html.escape(str(confidence))}</span>
      <h2>Resume technique</h2>
      <p>{summary}</p>
    </div>
    <div class="grid" style="margin-top:16px">{''.join(section for section in sections if section)}</div>
    """


def profile_source_preview_html(source_path: Path | None, source_name: str) -> str:
    if not source_path or not source_path.exists():
        return "<p class='alert'>Le fichier source n'est plus disponible pour l'aperçu.</p>"
    suffix = source_path.suffix.lower()
    if suffix in (".txt", ".csv", ".pts", ".asc", ".xyz"):
        try:
            text = source_path.read_text(encoding="utf-8-sig", errors="replace")
        except Exception:
            text = source_path.read_text(errors="replace")
        return f"<pre class='profile-text-preview'>{html.escape(text[:24000])}</pre>"
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook

            wb = load_workbook(source_path, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            rows = []
            for row in ws.iter_rows(max_row=45, max_col=12, values_only=True):
                cells = "".join(f"<td>{html.escape('' if value is None else str(value))}</td>" for value in row)
                if cells.strip():
                    rows.append(f"<tr>{cells}</tr>")
            return f"<div class='profile-table-preview'><p class='muted'>Aperçu réel de la première feuille Excel : {html.escape(ws.title)}</p><table>{''.join(rows)}</table></div>"
        except Exception as exc:
            return f"<p class='alert'>Aperçu Excel impossible : {html.escape(str(exc))}</p>"
    if suffix in (".dwg", ".dxf"):
        return f"""
        <div class="profile-schematic">
          <h3>Fichier AutoCAD chargé</h3>
          <p>Le navigateur ne peut pas afficher directement un fichier DWG/DXF comme AutoCAD.</p>
          <p><b>Fichier reçu :</b> {html.escape(source_name)}</p>
          <p>Le fichier est bien stocké pour l'analyse. Pour afficher exactement le dessin, il faudra générer une image/PDF depuis AutoCAD ou ajouter plus tard un moteur de conversion DWG côté serveur.</p>
        </div>
        """
    if suffix == ".zip":
        try:
            with zipfile.ZipFile(source_path, "r") as zf:
                names = zf.namelist()[:120]
            rows = "".join(f"<tr><td>{html.escape(name)}</td><td>{html.escape(Path(name).suffix.lower() or 'dossier')}</td></tr>" for name in names)
            return f"<div class='profile-table-preview'><p class='muted'>Contenu reel de l'archive ZIP.</p><table><tr><th>Fichier</th><th>Type</th></tr>{rows}</table></div>"
        except Exception as exc:
            return f"<p class='alert'>Archive ZIP non lisible : {html.escape(str(exc))}</p>"
    return f"<p class='alert'>Aperçu direct non disponible pour ce format : {html.escape(source_path.suffix)}</p>"


def render_profile_visual(source_url: str, source_name: str, analysis: dict, source_path: Path | None = None) -> str:
    suffix = Path(source_name).suffix.lower()
    detected = analysis.get("detected_elements") or []
    if isinstance(detected, str):
        detected = [detected]
    chips = "".join(f"<span class='pill'>{html.escape(str(item))}</span>" for item in detected[:10])
    if not chips:
        chips = "<span class='pill'>Axe</span><span class='pill'>Cotes</span><span class='pill'>Couches</span>"
    if suffix in (".png", ".jpg", ".jpeg"):
        preview = f"<img src='{html.escape(source_url)}' alt='Apercu du document' style='max-width:100%;max-height:640px;object-fit:contain;background:white;border-radius:10px'>"
    elif suffix == ".pdf":
        preview = f"<iframe src='{html.escape(source_url)}' title='Apercu PDF' style='width:100%;height:640px;border:0;background:white;border-radius:10px'></iframe><p class='muted'><a href='{html.escape(source_url)}' target='_blank'>Ouvrir le fichier original dans un nouvel onglet</a></p>"
    else:
        preview = profile_source_preview_html(source_path, source_name)
    return f"""
    <div class="card profile-visual-card">
      <h2>Lecture visuelle du document</h2>
      <p class="muted">Le fichier charge reste visible pendant l'analyse. Les repères ci-dessous guident la lecture comme sur un contrôle de profil.</p>
      <div class="profile-tags">{chips}</div>
      <div class="profile-viewer">
        {preview}
        <div class="profile-overlay-note">
          <b>Repères à vérifier</b>
          <span>Axe / TN / Projet / PST / couches / pentes / talus</span>
        </div>
      </div>
    </div>
    """


def render_profile_workspace(analysis: dict, source_url: str, source_name: str, source_path: Path | None = None) -> str:
    compatibility = render_compatibility_panel(technical_file_capability(source_path, source_name))
    return f"""
    <div class="grid" style="grid-template-columns:1.25fr .75fr;margin-top:16px">
      {render_profile_visual(source_url, source_name, analysis, source_path)}
      <div>
        {compatibility}
        {render_profile_analysis_result(analysis)}
      </div>
    </div>
    <style>
    .profile-viewer{{position:relative;margin-top:14px;background:#0f172a;border:1px solid rgba(148,163,184,.24);border-radius:12px;padding:12px;min-height:460px;display:flex;align-items:center;justify-content:center;overflow:hidden}}
    .profile-overlay-note{{position:absolute;left:22px;top:22px;background:rgba(5,11,20,.86);border:1px solid rgba(34,211,238,.35);border-radius:10px;padding:10px 12px;max-width:280px;box-shadow:0 12px 30px rgba(0,0,0,.28)}}
    .profile-overlay-note b{{display:block;color:#8deeff;margin-bottom:4px}}
    .profile-overlay-note span{{font-size:12px;color:#dbeafe}}
    .profile-tags{{display:flex;gap:8px;flex-wrap:wrap;margin:10px 0}}
    .profile-schematic{{width:100%;background:#fff;border-radius:10px;padding:10px;color:#102033}}
    .profile-text-preview{{width:100%;max-height:640px;overflow:auto;background:#fff;color:#111827;border-radius:10px;padding:14px;font-size:13px;line-height:1.45;text-align:left;white-space:pre-wrap}}
    .profile-table-preview{{width:100%;max-height:640px;overflow:auto;background:#fff;color:#111827;border-radius:10px;padding:12px}}
    .profile-table-preview table{{background:white;color:#111827;min-width:680px}}
    .profile-table-preview td{{border:1px solid #d7dee8;color:#111827;font-size:12px;padding:7px}}
    @media(max-width:900px){{.grid[style*="1.25fr"]{{grid-template-columns:1fr!important}}.profile-viewer{{min-height:320px}}.profile-viewer iframe{{height:420px!important}}}}
    </style>
    """


def format_fcfa(value: int) -> str:
    return f"{value:,}".replace(",", " ") + " FCFA"


def pvit_config_status() -> tuple[bool, list[str]]:
    missing = []
    if not PVIT_MERCHANT_SLUG:
        missing.append("SLUG marchand")
    if not PVIT_SECRET:
        missing.append("X-Secret / cle secrete")
    if not PVIT_ACCOUNT_CODE:
        missing.append("Compte d'operation TEST")
    if not PVIT_LINK_PATH:
        missing.append("Endpoint LINK")
    return not missing, missing


def pvit_endpoint(path: str) -> str:
    return f"{PVIT_BASE_URL}/{path.lstrip('/')}"


def pvit_request(path: str, payload: dict | None = None, method: str = "POST") -> dict:
    headers = {"Content-Type": "application/json"}
    if PVIT_SECRET:
        headers["X-Secret"] = PVIT_SECRET
    data = json.dumps(payload or {}).encode("utf-8") if method.upper() != "GET" else None
    request = urllib.request.Request(pvit_endpoint(path), data=data, headers=headers, method=method.upper())
    with urllib.request.urlopen(request, timeout=20) as response:
        raw_text = response.read().decode("utf-8", errors="replace")
    try:
        return json.loads(raw_text)
    except Exception:
        return {"raw": raw_text}


def save_setting(key: str, value: str) -> None:
    with db() as con:
        con.execute("INSERT INTO settings(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value", (key, value))


def get_setting(key: str, default: str = "") -> str:
    with db() as con:
        row = con.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
    return row["value"] if row else default


def extract_secret_from_payload(data: dict, raw: str) -> str:
    candidates = [
        "secret",
        "x_secret",
        "x-secret",
        "X-Secret",
        "secret_key",
        "secretKey",
        "key",
        "api_secret",
    ]
    for key in candidates:
        value = data.get(key)
        if isinstance(value, list):
            value = value[0] if value else ""
        if value:
            return str(value).strip()
    match = re.search(r"(?i)(?:x[-_]?secret|secret(?:_key)?|secretKey)\s*[:=]\s*['\"]?([A-Za-z0-9._~:/+=-]{12,})", raw)
    return match.group(1).strip() if match else ""


def append_pvit_receiver_log(line: str) -> None:
    try:
        log_path = DATA / "pvit_secret_receiver.log"
        log_path.parent.mkdir(parents=True, exist_ok=True)
        with log_path.open("a", encoding="utf-8") as f:
            f.write(f"[{now()}] {line}\n")
    except Exception:
        pass


def create_pvit_payment(user: sqlite3.Row, offer_key: str, phone: str, payment_id: int) -> dict:
    offer = PAYMENT_OFFERS.get(offer_key, PAYMENT_OFFERS["pdf"])
    configured, missing = pvit_config_status()
    reference = f"BTP-{payment_id}-{int(time.time())}"
    if not configured:
        return {
            "status": "config_incomplete",
            "reference": reference,
            "payment_url": "",
            "message": "Configuration PVit TEST incomplete : " + ", ".join(missing),
            "raw": {"missing": missing, "mode": PVIT_MODE, "merchant_slug": PVIT_MERCHANT_SLUG},
        }

    payload = {
        "merchant_slug": PVIT_MERCHANT_SLUG,
        "reference": reference,
        "amount": offer["amount"],
        "currency": "XAF",
        "description": f"BTP Smart Tools - {offer['label']}",
        "customer_email": user["email"],
        "customer_phone": phone,
        "callback_url": f"{PUBLIC_SITE_URL}/payment/callback",
        "success_url": f"{PUBLIC_SITE_URL}/payment/success",
        "failed_url": f"{PUBLIC_SITE_URL}/payment/failed",
        "account_code": PVIT_ACCOUNT_CODE,
        "mode": PVIT_MODE,
    }
    data = json.dumps(payload).encode("utf-8")
    request = urllib.request.Request(
        pvit_endpoint(PVIT_LINK_PATH),
        data=data,
        headers={
            "Content-Type": "application/json",
            "X-Secret": PVIT_SECRET,
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=20) as response:
            raw_text = response.read().decode("utf-8", errors="replace")
        try:
            raw = json.loads(raw_text)
        except Exception:
            raw = {"raw": raw_text}
        payment_url = raw.get("payment_url") or raw.get("url") or raw.get("link") or ""
        return {"status": "sent_to_pvit", "reference": reference, "payment_url": payment_url, "message": "Demande de paiement envoyee a PVit TEST.", "raw": raw}
    except urllib.error.HTTPError as exc:
        raw_text = exc.read().decode("utf-8", errors="replace")
        return {"status": "pvit_error", "reference": reference, "payment_url": "", "message": f"Erreur PVit HTTP {exc.code}", "raw": {"error": raw_text}}
    except Exception as exc:
        return {"status": "pvit_error", "reference": reference, "payment_url": "", "message": str(exc), "raw": {"error": str(exc)}}


def grant_payment_credits(con: sqlite3.Connection, payment_id: int) -> bool:
    payment = con.execute("SELECT * FROM payments WHERE id=?", (payment_id,)).fetchone()
    if not payment or payment["credited"]:
        return False
    offer_key = payment["offer_key"] if "offer_key" in payment.keys() and payment["offer_key"] else ""
    offer = PAYMENT_OFFERS.get(offer_key)
    credits = int(payment["credits"] or 0)
    if not offer and not credits:
        for key, candidate in PAYMENT_OFFERS.items():
            if candidate["label"] == payment["offer"]:
                offer_key = key
                offer = candidate
                credits = int(candidate.get("credits", 0))
                break
    if not offer:
        offer = PAYMENT_OFFERS["pdf"]
    if not credits:
        credits = int(offer.get("credits", 0))
    subscription = offer.get("subscription", "none")
    days = int(offer.get("days", 0) or 0)
    until = ""
    if days:
        until = (datetime.now() + timedelta(days=days)).strftime("%Y-%m-%d")
        con.execute(
            "UPDATE users SET credits=credits+?, subscription=?, subscription_until=? WHERE id=?",
            (credits, subscription, until, payment["user_id"]),
        )
    else:
        con.execute("UPDATE users SET credits=credits+? WHERE id=?", (credits, payment["user_id"]))
    con.execute("UPDATE payments SET credited=1, credits=?, offer_key=? WHERE id=?", (credits, offer_key, payment_id))
    return True


def render_page(title: str, body: str, user: sqlite3.Row | None = None) -> bytes:
    logged = user is not None
    is_admin = logged and user["role"] == "admin"
    nav = [
        ("Accueil", "/"),
        ("Connexion", "/login") if not logged else ("Tableau de bord", "/dashboard"),
        ("Creer un compte", "/register") if not logged else ("Abonnement", "/payment"),
        ("Cartouche", "/generator"),
        ("Modeles cartouche", "/templates"),
        ("Batch PDF", "/batch"),
        ("Aide", "/assistant"),
    ]
    if is_admin:
        nav.append(("Admin", "/admin"))
    if logged:
        nav.append(("Deconnexion", "/logout"))
    links = "".join(f'<a href="{href}">{label}</a>' for label, href in nav)
    user_badge = f"<span class='badge'>{html.escape(user['name'])} - {html.escape(user['role'])}</span>" if logged else "<span class='badge'>Non connecte</span>"
    page = f"""<!doctype html>
<html lang="fr">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{html.escape(title)} - BTP Smart Tools</title>
<style>
:root{{--navy:#050b14;--navy2:#0b1220;--blue:#2f8cff;--cyan:#22d3ee;--green:#10b981;--purple:#7c3aed;--amber:#f59e0b;--red:#ef4444;--paper:#07111f;--line:#203044;--text:#eaf2ff;--muted:#94a3b8;--panel:#0b1220}}
*{{box-sizing:border-box}}body{{margin:0;background:radial-gradient(circle at 12% 0%,rgba(47,140,255,.20),transparent 30%),radial-gradient(circle at 88% 8%,rgba(124,58,237,.18),transparent 28%),linear-gradient(135deg,#050b14,#07111f 48%,#0a1020);font-family:Segoe UI,Arial,sans-serif;color:var(--text)}}body:before{{content:"";position:fixed;inset:0;pointer-events:none;background-image:linear-gradient(rgba(34,211,238,.035) 1px,transparent 1px),linear-gradient(90deg,rgba(34,211,238,.035) 1px,transparent 1px);background-size:42px 42px;mask-image:linear-gradient(to bottom,rgba(0,0,0,.85),transparent)}}header{{background:rgba(5,11,20,.82);color:white;position:sticky;top:0;z-index:20;border-bottom:1px solid rgba(34,211,238,.18);backdrop-filter:blur(16px)}}.top{{max-width:1220px;margin:auto;display:flex;align-items:center;justify-content:space-between;gap:14px;padding:14px 18px}}.brand{{display:flex;align-items:center;gap:10px;font-weight:900}}.brand b{{font-size:36px;color:var(--cyan);letter-spacing:0;text-shadow:0 0 26px rgba(34,211,238,.55)}}.brand span{{font-size:11px;line-height:1.1;color:#dbeafe}}nav{{display:flex;gap:8px;flex-wrap:wrap}}nav a{{color:#eaf2ff;text-decoration:none;background:rgba(255,255,255,.055);border:1px solid rgba(148,163,184,.18);padding:10px 12px;border-radius:8px;font-weight:800;transition:.18s}}nav a:hover{{background:rgba(47,140,255,.22);border-color:rgba(34,211,238,.55);transform:translateY(-1px)}}.wrap{{max-width:1220px;margin:auto;padding:22px;position:relative}}.grid{{display:grid;grid-template-columns:1fr 1fr;gap:20px}}.grid3{{display:grid;grid-template-columns:repeat(3,1fr);gap:16px}}.card{{background:linear-gradient(180deg,rgba(15,23,42,.92),rgba(8,15,28,.92));border:1px solid rgba(148,163,184,.20);border-radius:14px;padding:22px;box-shadow:0 20px 54px rgba(0,0,0,.28);color:#eaf2ff}}.hero{{position:relative;overflow:hidden;background:linear-gradient(135deg,rgba(5,11,20,.98),rgba(9,25,48,.96) 54%,rgba(19,18,48,.94));color:white;border:1px solid rgba(34,211,238,.22);border-radius:20px;padding:48px;display:grid;grid-template-columns:1fr 1fr;gap:32px;align-items:center;box-shadow:0 28px 80px rgba(0,0,0,.42)}}.hero:before{{content:"";position:absolute;inset:-40%;background:radial-gradient(circle,rgba(34,211,238,.22),transparent 28%);animation:floatGlow 9s ease-in-out infinite}}@keyframes floatGlow{{0%,100%{{transform:translate(-4%,2%)}}50%{{transform:translate(4%,-2%)}}}}.hero>*{{position:relative}}.hero h1{{font-size:78px;line-height:.9;margin:0;letter-spacing:0;text-shadow:0 0 36px rgba(47,140,255,.34)}}.hero h2{{font-size:27px;margin:10px 0 14px;color:#dbeafe}}.hero p{{color:#b7c7dd;font-size:17px;line-height:1.55}}label{{display:block;font-size:13px;font-weight:900;color:#c7ddf6;margin:10px 0 5px}}input,select,textarea{{width:100%;border:1px solid rgba(148,163,184,.28);border-radius:9px;padding:12px;font:inherit;background:#f8fafc;color:#102033;outline:none}}input:focus,select:focus,textarea:focus{{border-color:var(--cyan);box-shadow:0 0 0 3px rgba(34,211,238,.16)}}textarea{{min-height:95px;resize:vertical}}button,.btn{{border:0;border-radius:9px;padding:12px 17px;font-weight:900;color:white;background:linear-gradient(135deg,#2f8cff,#22d3ee);cursor:pointer;min-height:44px;text-decoration:none;display:inline-flex;align-items:center;justify-content:center;box-shadow:0 12px 30px rgba(47,140,255,.25);transition:.18s}}button:hover,.btn:hover{{transform:translateY(-1px);filter:brightness(1.07)}}.green{{background:linear-gradient(135deg,#10b981,#22d3ee)}}.purple{{background:linear-gradient(135deg,#7c3aed,#2f8cff)}}.amber{{background:linear-gradient(135deg,#f59e0b,#b45309)}}.red{{background:linear-gradient(135deg,#ef4444,#b91c1c)}}.dark{{background:linear-gradient(135deg,#0f172a,#111827)}}.row{{display:flex;gap:10px;align-items:center;flex-wrap:wrap}}.badge,.pill{{background:rgba(34,211,238,.12);color:#8deeff;border:1px solid rgba(34,211,238,.28);border-radius:999px;padding:7px 11px;font-weight:900;font-size:12px}}.muted{{color:var(--muted)}}table{{width:100%;border-collapse:collapse}}th,td{{border-bottom:1px solid rgba(148,163,184,.18);padding:11px;text-align:left;font-size:14px}}th{{background:rgba(15,23,42,.75);color:#dbeafe}}.stat{{background:linear-gradient(180deg,rgba(15,23,42,.95),rgba(8,15,28,.96));border:1px solid rgba(34,211,238,.18);border-radius:12px;padding:18px;box-shadow:0 16px 36px rgba(0,0,0,.24);color:#eaf2ff}}.stat b{{font-size:30px;color:var(--cyan);display:block}}.sheet{{background:white;border:2px solid #172033;min-height:450px;display:grid;grid-template-columns:1fr 170px;color:#111827;box-shadow:0 0 0 1px rgba(34,211,238,.22),0 28px 70px rgba(0,0,0,.42);transform:perspective(900px) rotateY(-4deg);border-radius:5px;overflow:hidden}}.draw{{padding:24px;border-right:2px solid #172033;background:linear-gradient(180deg,#ffffff,#fbfdff)}}.cartouche{{background:#fdfefe}}.cartouche div{{border-bottom:1px solid #172033;padding:9px;font-size:11px}}.cartouche .ske{{font-size:40px;color:var(--blue);font-weight:900;text-align:center;background:#eef6ff}}.legend-line{{display:grid;grid-template-columns:120px 110px 1fr;gap:8px;align-items:end;margin-bottom:8px}}.alert{{padding:13px;border-radius:10px;background:rgba(34,211,238,.10);color:#c8f5ff;font-weight:900;border:1px solid rgba(34,211,238,.24)}}.feature-row{{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin-top:18px}}.feature{{background:rgba(255,255,255,.06);border:1px solid rgba(34,211,238,.16);border-radius:12px;padding:16px}}.feature b{{display:block;color:white;margin-bottom:6px}}.feature span{{color:#b7c7dd;font-size:13px}}.roadmap-card{{position:relative;overflow:hidden}}.roadmap-card:after{{content:"";position:absolute;right:-35px;top:-35px;width:110px;height:110px;border-radius:50%;background:radial-gradient(circle,rgba(34,211,238,.18),transparent 65%)}}@media(max-width:900px){{.grid,.grid3,.hero,.feature-row{{grid-template-columns:1fr}}.hero h1{{font-size:48px}}.hero{{padding:28px}}.sheet{{grid-template-columns:1fr;transform:none}}.draw{{border-right:0;border-bottom:2px solid #111827}}}}
</style>
<style>
.cartouche-focus{{display:grid;grid-template-columns:1.4fr .9fr;gap:16px;align-items:center;background:linear-gradient(135deg,rgba(34,211,238,.13),rgba(16,185,129,.08));border:1px solid rgba(34,211,238,.28);border-radius:14px;padding:18px;margin-bottom:16px}}
.cartouche-focus h2{{margin:8px 0 6px}}.cartouche-focus p{{margin:0;color:#c7d7ee;line-height:1.5}}
.focus-cost{{display:grid;gap:7px;background:rgba(2,6,23,.36);border:1px solid rgba(148,163,184,.22);border-radius:12px;padding:14px}}
.focus-cost b{{color:#8deeff;font-size:18px}}.focus-cost span{{color:#dbeafe;font-weight:700}}
.steps{{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin:0 0 16px}}
.step{{display:grid;grid-template-columns:42px 1fr;gap:5px 10px;align-items:center;background:rgba(255,255,255,.06);border:1px solid rgba(34,211,238,.18);border-radius:12px;padding:13px}}
.step b{{grid-row:span 2;width:34px;height:34px;border-radius:50%;display:flex;align-items:center;justify-content:center;background:#22d3ee;color:#08213A;font-weight:900}}
.step strong{{color:#fff}}.step span{{color:#b7c7dd;font-size:12px}}
.quality-card{{background:rgba(2,6,23,.38);border:1px solid rgba(16,185,129,.28);border-radius:12px;padding:16px;margin-top:16px}}
.quality-card h3{{margin:0}}.quality-card ul{{list-style:none;padding:0;margin:12px 0;display:grid;gap:8px}}
.quality-card li{{display:grid;grid-template-columns:120px 1fr;gap:10px;padding:9px;border-bottom:1px solid rgba(148,163,184,.16)}}
.quality-card li b{{color:#8deeff}}.quality-card li span{{color:#dbeafe}}
@media(max-width:900px){{.cartouche-focus,.steps{{grid-template-columns:1fr}}.quality-card li{{grid-template-columns:1fr}}}}
</style>
<style>
.preview-wrap{{position:sticky;top:92px}}
.preview-stage{{background:#dfe7f0;border:1px solid #cbd5e1;border-radius:12px;padding:12px;box-shadow:inset 0 1px 0 #fff}}
.preview-sheet{{height:480px;background:white;border:2px solid #172033;display:grid;grid-template-columns:minmax(0,1fr) 118px;overflow:hidden}}
.preview-plan{{position:relative;background:#fbfdff;display:flex;align-items:center;justify-content:center;border-right:2px solid #172033;color:#64748b;font-weight:800;text-align:center}}
.preview-plan iframe{{position:absolute;inset:0;width:100%;height:100%;border:0;background:white}}
.preview-plan img{{position:absolute;inset:0;width:100%;height:100%;object-fit:contain;background:white}}
.preview-plan svg{{width:82%;max-height:82%;opacity:.95}}
.preview-empty{{padding:16px;max-width:260px;line-height:1.35;color:#64748b}}
.preview-file-name{{position:absolute;left:10px;bottom:10px;right:10px;background:rgba(15,23,42,.82);color:white;border-radius:8px;padding:7px 9px;font-size:11px;font-weight:800;text-align:left;z-index:3;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.preview-file-name a{{color:white;text-decoration:underline}}
.preview-cartouche{{font-size:8px;background:#fdfefe;display:flex;flex-direction:column;color:#111827}}
.preview-logo{{height:52px;display:flex;align-items:center;justify-content:center;font-size:22px;font-weight:900;border-bottom:1px solid #172033;color:#08213A}}
.preview-logo img{{max-width:90%;max-height:44px;object-fit:contain}}
.preview-sec{{border-bottom:1px solid #cbd5e1;padding:5px 6px}}
.preview-sec b{{display:block;font-size:8px;margin-top:2px;line-height:1.15}}
.preview-tag{{display:inline-block;background:#08213A;color:white;font-size:6px;font-weight:900;padding:2px 5px;margin-bottom:2px}}
.preview-table{{margin:6px;border:1px solid #cbd5e1;display:grid;grid-template-columns:1fr 1fr;font-size:6px}}
.preview-table span{{padding:3px;border-bottom:1px solid #e2e8f0}}
.preview-list{{margin:0 6px 6px 6px;font-size:6px;display:grid;gap:2px}}
.preview-list div{{display:flex;gap:5px;align-items:center;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.preview-swatch{{width:7px;height:7px;border-radius:2px;display:inline-block;flex:0 0 auto}}
.preview-footer{{margin-top:auto;background:#08213A;color:white;padding:7px;text-align:center;font-weight:900}}
.preview-sheet.platform_premium{{border-color:#8A6A20}}.preview-sheet.platform_premium .preview-cartouche{{background:#fbf7ec}}.preview-sheet.platform_premium .preview-logo{{background:#07192E;color:#f8fafc;border-bottom:4px solid #8A6A20}}.preview-sheet.platform_premium .preview-tag,.preview-sheet.platform_premium .preview-footer{{background:#07192E;color:#f7d77b}}
.preview-sheet.platform_modern{{border-color:#334155}}.preview-sheet.platform_modern .preview-cartouche{{background:#fafbfc;border-left:10px solid #e2e8f0}}.preview-sheet.platform_modern .preview-logo{{height:48px;color:#334155}}.preview-sheet.platform_modern .preview-tag,.preview-sheet.platform_modern .preview-footer{{background:#334155}}
.preview-sheet.platform_topo{{border-color:#0f766e}}.preview-sheet.platform_topo .preview-cartouche{{background:#f0fdfa;border-left:6px solid #0f766e}}.preview-sheet.platform_topo .preview-logo{{color:#0f766e;border-radius:0 0 18px 18px}}.preview-sheet.platform_topo .preview-tag,.preview-sheet.platform_topo .preview-footer{{background:#0f766e}}
.preview-sheet.platform_engineering{{border-color:#1d4ed8}}.preview-sheet.platform_engineering .preview-cartouche{{background-image:linear-gradient(#dbeafe 1px,transparent 1px),linear-gradient(90deg,#dbeafe 1px,transparent 1px);background-size:18px 18px}}.preview-sheet.platform_engineering .preview-logo{{color:#1d4ed8}}.preview-sheet.platform_engineering .preview-tag,.preview-sheet.platform_engineering .preview-footer{{background:#1d4ed8}}
.assistant-float{{position:fixed;right:12px;bottom:10px;z-index:60;display:flex;align-items:center;gap:6px;text-decoration:none;color:#102033;max-width:168px;opacity:.78}}
.assistant-float:hover{{opacity:1;transform:translateY(-1px)}}
.assistant-float .bubble{{background:rgba(255,255,255,.94);border:1px solid rgba(47,140,255,.20);border-radius:9px;padding:6px 8px;box-shadow:0 8px 18px rgba(15,23,42,.12)}}
.assistant-float .bubble b{{display:block;color:#08213A;font-size:10px;margin-bottom:1px;white-space:nowrap}}
.assistant-float .bubble span{{display:none}}
.assistant-float .icon{{width:32px;height:32px;border-radius:10px;background:linear-gradient(135deg,#2f8cff,#10b981);color:white;display:flex;align-items:center;justify-content:center;font-weight:900;font-size:15px;box-shadow:0 8px 18px rgba(47,140,255,.22)}}
@media(max-width:700px){{.assistant-float{{right:8px;bottom:8px;max-width:38px}}.assistant-float .bubble{{display:none}}.assistant-float .icon{{width:36px;height:36px;border-radius:11px}}}}
</style>
</head>
<body>
<header><div class="top"><div class="brand"><b>BTP</b><span>SMART TOOLS<br>PLANS & CARTOUCHES<br>by SKE System</span></div><nav>{links}</nav>{user_badge}</div></header>
<main class="wrap">{body}</main>
<a class="assistant-float" href="/assistant" title="Ouvrir l'assistant support">
  <div class="bubble"><b>Assistant Cartouche</b><span>Aide pour PDF, logo, modele, credits et telechargement.</span></div>
  <div class="icon">?</div>
</a>
</body>
</html>"""
    return page.encode("utf-8")


def preview_panel(title: str = "Apercu du modele") -> str:
    return f"""
    <div class="card preview-wrap">
      <h2>{html.escape(title)}</h2>
      <p class="muted">Change le modele ou charge un PDF : l'apercu se met a jour avant generation.</p>
      <div class="preview-stage">
        <div id="previewSheet" class="preview-sheet platform_standard">
          <div class="preview-plan" id="previewPlan">
            <svg viewBox="0 0 440 280">
              <rect x="25" y="25" width="360" height="220" fill="none" stroke="#111827" stroke-width="2"/>
              <g transform="translate(54 58)" stroke="#111827" fill="#111827">
                <line x1="0" y1="-25" x2="0" y2="25" stroke-width="2"/>
                <line x1="-25" y1="0" x2="25" y2="0" stroke-width="2"/>
                <path d="M0 -31 L-6 -17 L0 -22 L6 -17 Z"/>
                <text x="-4" y="-36" font-size="12" font-weight="700">N</text>
                <text x="-4" y="43" font-size="10">S</text>
                <text x="30" y="4" font-size="10">E</text>
                <text x="-42" y="4" font-size="10">O</text>
              </g>
              <path d="M70 220 L70 92 L145 92 L145 60 L310 60 L310 220 Z" fill="none" stroke="#111827" stroke-width="4"/>
              <path d="M98 220 L98 118 L166 118 L166 88 L285 88 L285 220" fill="none" stroke="#2f8cff" stroke-width="3"/>
              <line x1="45" y1="150" x2="370" y2="150" stroke="#94a3b8" stroke-dasharray="8 8"/>
              <text x="122" y="172" fill="#10b981" font-size="16">apercu plan</text>
            </svg>
          </div>
          <div class="preview-cartouche">
            <div class="preview-logo" id="previewLogo">SKE</div>
            <div class="preview-sec"><span class="preview-tag">PROJET</span><b id="previewProject">Jardin Botanique</b></div>
            <div class="preview-sec"><span class="preview-tag">ENTREPRISE</span><b id="previewCompany">SKE System</b></div>
            <div class="preview-sec"><span class="preview-tag">TYPE DE PLAN</span><b id="previewType">Plan technique</b></div>
            <div class="preview-sec"><span class="preview-tag">ECHELLE</span><b id="previewScale">1/100</b></div>
            <div class="preview-sec"><span class="preview-tag">FORMAT / REV.</span><b><span id="previewFormat">Automatique</span> - <span id="previewRevision">REV 00</span></b></div>
            <div class="preview-sec"><span class="preview-tag">RESPONSABLES</span><b><span id="previewManager">Chef projet</span> / <span id="previewOperator">Operateur</span></b></div>
            <div class="preview-list" id="previewLegends"><div><span class="preview-swatch" style="background:#2487ff"></span>Canal principal</div></div>
            <div class="preview-table" id="previewTable"><span>REPERE</span><span>QTE</span><span>Zone projet</span><span>1</span></div>
            <div class="preview-footer" id="previewModelName">BTP Standard</div>
          </div>
        </div>
      </div>
      <p class="alert">Apercu rapide. Le PDF final reste genere par le moteur professionnel.</p>
    </div>
    """


PREVIEW_SCRIPT = """
<script>
const previewNames = {
  platform_standard: 'BTP Standard',
  platform_premium: 'BTP Premium Bleu',
  platform_modern: 'BTP Moderne Gris',
  platform_topo: 'BTP Topographie',
  platform_engineering: 'BTP Ingenierie'
};
function updatePreview(){
  const form = document.querySelector('form');
  if(!form || !document.getElementById('previewSheet')) return;
  const sel = form.querySelector('[name="template_id"]');
  const key = sel ? (sel.value || 'platform_standard') : 'platform_standard';
  const cls = key.startsWith('platform_') ? key : 'platform_standard';
  previewSheet.className = 'preview-sheet ' + cls;
  previewModelName.textContent = previewNames[cls] || 'Modele personnalise';
  const get = n => (form.querySelector('[name="'+n+'"]') || {}).value || '';
  previewProject.textContent = get('project') || 'Jardin Botanique';
  previewCompany.textContent = get('company') || 'SKE System';
  previewType.textContent = get('plan_type') || 'Plan technique';
  previewScale.textContent = get('scale') || '1/100';
  previewFormat.textContent = get('format_plan') || 'Automatique';
  previewRevision.textContent = get('revision') || 'REV 00';
  previewManager.textContent = get('project_manager') || 'Chef projet';
  previewOperator.textContent = get('operator') || 'Operateur';
  const themeColor = get('theme_color');
  if(themeColor){
    document.querySelectorAll('.preview-tag,.preview-footer').forEach(x => x.style.background = themeColor);
  }
  const legendNodes = [...document.querySelectorAll('#legends .legend-line')].slice(0,3);
  if(legendNodes.length){
    previewLegends.innerHTML = legendNodes.map(x => {
      const color = x.children[1]?.value || '#2487ff';
      const text = x.children[2]?.value || 'Legende';
      return `<div><span class="preview-swatch" style="background:${color}"></span>${text}</div>`;
    }).join('');
  }
  const cols = [document.getElementById('col1')?.value || 'DESIGNATION', document.getElementById('col3')?.value || 'QUANTITE'];
  const row = document.querySelector('#elementsRows .legend-line');
  const r1 = row?.children[0]?.value || 'Zone projet';
  const r3 = row?.children[2]?.value || '1';
  previewTable.innerHTML = `<span>${cols[0]}</span><span>${cols[1]}</span><span>${r1}</span><span>${r3}</span>`;
}
function bindPreview(){
  const form = document.querySelector('form');
  if(!form) return;
  form.addEventListener('input', updatePreview);
  form.addEventListener('change', updatePreview);
  const showImportedPreview = (fileInput) => {
      const f = fileInput.files && fileInput.files[0];
      if(f){
        const name = f.name.toLowerCase();
        const planField = form.querySelector('[name="plan_type"]');
        const scaleField = form.querySelector('[name="scale"]');
        const planGuesses = [
          [/ferraillage|armature|acier/, 'Plan de ferraillage'],
          [/coffrage/, 'Plan de coffrage'],
          [/fondation|semelle/, 'Plan de fondations'],
          [/profil.*travers|travers/, 'Profil en travers'],
          [/profil.*long/, 'Profil en long'],
          [/vrd|reseau|rÃ©seau|assainissement/, 'Plan VRD'],
          [/topo|topographique|implantation/, 'Plan topographique'],
          [/route|chaussee|chaussÃ©e/, 'Plan de route'],
          [/dallot|dalot/, 'Plan de dallot'],
          [/masse/, 'Plan de masse'],
          [/situation/, 'Plan de situation'],
          [/beton|bÃ©ton|radier|voile/, 'Plan beton']
        ];
        if(planField && (!planField.value || ['Plan technique','Plan beton'].includes(planField.value))){
          const found = planGuesses.find(([rx]) => rx.test(name));
          if(found) planField.value = found[1];
        }
        const scaleMatch = name.match(/(?:ech|echelle|scale)?[\\s_\\-:]*(?:1)?[\\/_:-]?(20|25|50|75|100|125|150|200|250|500|750|1000|1500|2000|2500|5000|10000)\\b/);
        if(scaleField && scaleMatch && (!scaleField.value || scaleField.value === '1/100')){
          scaleField.value = '1/' + scaleMatch[1];
        }
      }
      if(!f){
        updatePreview();
        return;
      }
      const lowerName = f.name.toLowerCase();
      const isPdf = f.type === 'application/pdf' || lowerName.endsWith('.pdf');
      const isImage = (f.type && f.type.startsWith('image/')) || /\\.(png|jpe?g|webp|bmp|gif|tiff?)$/.test(lowerName);
      const safeName = f.name.replace(/[&<>"']/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));
      const url = URL.createObjectURL(f);
      if(isPdf){
        previewPlan.innerHTML = '<iframe src="'+url+'#toolbar=0&navpanes=0" title="Apercu du PDF importe"></iframe><div class="preview-file-name">PDF charge : '+safeName+' - <a href="'+url+'" target="_blank">ouvrir</a></div>';
      }else if(isImage){
        previewPlan.innerHTML = '<img src="'+url+'" alt="Apercu du document importe"><div class="preview-file-name">Image chargee : '+safeName+'</div>';
      }else{
        previewPlan.innerHTML = '<div class="preview-empty">Fichier charge : '+safeName+'<br>Le navigateur ne peut pas afficher ce format directement. Pour voir le vrai plan dans l apercu et dans le PDF final, utilise un PDF ou une image. Pour DWG/DXF, exporte d abord le plan en PDF depuis AutoCAD.</div>';
      }
      updatePreview();
  };
  const fileInputs = form.querySelectorAll('input[type="file"][name="file"], input[type="file"][name="files"]');
  fileInputs.forEach(fileInput => fileInput.addEventListener('change', () => showImportedPreview(fileInput)));
  const logoInput = form.querySelector('input[type="file"][name="logo"]');
  if(logoInput){
    logoInput.addEventListener('change', () => {
      const f = logoInput.files && logoInput.files[0];
      if(f && f.type.startsWith('image/')){
        previewLogo.innerHTML = '<img src="'+URL.createObjectURL(f)+'">';
      }
    });
  }
  updatePreview();
}
document.addEventListener('DOMContentLoaded', bindPreview);
</script>
"""


def _matrix_multiply(a: tuple[float, float, float, float, float, float], b: tuple[float, float, float, float, float, float]):
    return (
        a[0] * b[0] + a[2] * b[1],
        a[1] * b[0] + a[3] * b[1],
        a[0] * b[2] + a[2] * b[3],
        a[1] * b[2] + a[3] * b[3],
        a[0] * b[4] + a[2] * b[5] + a[4],
        a[1] * b[4] + a[3] * b[5] + a[5],
    )


def _matrix_apply(m: tuple[float, float, float, float, float, float], x: float, y: float):
    return (m[0] * x + m[2] * y + m[4], m[1] * x + m[3] * y + m[5])


def estimate_pdf_content_bbox(page) -> tuple[float, float, float, float] | None:
    try:
        stream = ContentStream(page.get_contents(), page.pdf)
        media = page.mediabox
        mx0, my0 = float(media.left), float(media.bottom)
        mx1, my1 = float(media.right), float(media.top)
        stack: list[tuple[float, float, float, float, float, float]] = []
        ctm = (1.0, 0.0, 0.0, 1.0, 0.0, 0.0)
        text_matrix = (1.0, 0.0, 0.0, 1.0, 0.0, 0.0)
        points: list[tuple[float, float]] = []

        def add_point(x, y):
            px, py = _matrix_apply(ctm, float(x), float(y))
            if mx0 - 50 <= px <= mx1 + 50 and my0 - 50 <= py <= my1 + 50:
                points.append((px, py))

        def add_text_point(x, y):
            tx, ty = _matrix_apply(_matrix_multiply(ctm, text_matrix), float(x), float(y))
            if mx0 - 50 <= tx <= mx1 + 50 and my0 - 50 <= ty <= my1 + 50:
                points.append((tx, ty))

        for operands, operator in stream.operations:
            op = operator.decode("latin1") if isinstance(operator, bytes) else str(operator)
            try:
                if op == "q":
                    stack.append(ctm)
                elif op == "Q" and stack:
                    ctm = stack.pop()
                elif op == "cm" and len(operands) >= 6:
                    vals = tuple(float(v) for v in operands[:6])
                    ctm = _matrix_multiply(ctm, vals)
                elif op == "re" and len(operands) >= 4:
                    x, y, w, h = (float(v) for v in operands[:4])
                    for px, py in ((x, y), (x + w, y), (x, y + h), (x + w, y + h)):
                        add_point(px, py)
                elif op in ("m", "l") and len(operands) >= 2:
                    add_point(operands[0], operands[1])
                elif op in ("c", "v", "y"):
                    vals = [float(v) for v in operands]
                    for i in range(0, len(vals) - 1, 2):
                        add_point(vals[i], vals[i + 1])
                elif op == "Tm" and len(operands) >= 6:
                    text_matrix = tuple(float(v) for v in operands[:6])
                    add_text_point(0, 0)
                elif op in ("Td", "TD") and len(operands) >= 2:
                    text_matrix = _matrix_multiply(text_matrix, (1.0, 0.0, 0.0, 1.0, float(operands[0]), float(operands[1])))
                    add_text_point(0, 0)
                elif op == "Do":
                    for px, py in ((0, 0), (1, 0), (0, 1), (1, 1)):
                        add_point(px, py)
            except Exception:
                continue
        if len(points) < 4:
            return None
        xs = [p[0] for p in points]
        ys = [p[1] for p in points]
        x0, x1 = max(mx0, min(xs)), min(mx1, max(xs))
        y0, y1 = max(my0, min(ys)), min(my1, max(ys))
        if x1 <= x0 or y1 <= y0:
            return None
        pad_x = max((x1 - x0) * 0.035, 8)
        pad_y = max((y1 - y0) * 0.035, 8)
        x0, y0 = max(mx0, x0 - pad_x), max(my0, y0 - pad_y)
        x1, y1 = min(mx1, x1 + pad_x), min(my1, y1 + pad_y)
        if (x1 - x0) * (y1 - y0) < (mx1 - mx0) * (my1 - my0) * 0.08:
            return None
        return x0, y0, x1, y1
    except Exception:
        return None


def merge_source_pdf_into_sheet(sheet_pdf: bytes, source_path: Path, output: Path, draw_box: tuple[float, float, float, float], page_size: tuple[float, float]):
    sheet_page = PdfReader(io.BytesIO(sheet_pdf)).pages[0]
    source_page = PdfReader(str(source_path)).pages[0]
    try:
        source_page.transfer_rotation_to_content()
    except Exception:
        pass
    draw_x, draw_y, draw_w, draw_h = draw_box
    media = source_page.cropbox if source_page.cropbox else source_page.mediabox
    src_x0, src_y0 = float(media.left), float(media.bottom)
    src_x1, src_y1 = float(media.right), float(media.top)
    content_bbox = estimate_pdf_content_bbox(source_page)
    if content_bbox:
        src_x0, src_y0, src_x1, src_y1 = content_bbox
    src_w = src_x1 - src_x0
    src_h = src_y1 - src_y0
    scale = min(draw_w / src_w, draw_h / src_h) * 1.035
    scale = min(scale, draw_w / src_w, draw_h / src_h)
    tx = draw_x + (draw_w - src_w * scale) / 2
    ty = draw_y + (draw_h - src_h * scale) / 2
    final_page = PageObject.create_blank_page(width=page_size[0], height=page_size[1])
    final_page.merge_transformed_page(source_page, Transformation().translate(-src_x0, -src_y0).scale(scale).translate(tx, ty))
    final_page.merge_page(sheet_page)
    writer = PdfWriter()
    writer.add_page(final_page)
    with output.open("wb") as f:
        writer.write(f)


def draw_compass_rose(c: canvas.Canvas, x: float, y: float, size: float, color=colors.black) -> None:
    """Draw a compact north/south/east/west marker for topo plans."""
    c.saveState()
    c.setStrokeColor(color)
    c.setFillColor(color)
    c.setLineWidth(max(0.6, size * 0.035))
    half = size * 0.36
    c.line(x, y - half, x, y + half)
    c.line(x - half, y, x + half, y)
    c.line(x - half * 0.68, y - half * 0.68, x + half * 0.68, y + half * 0.68)
    c.line(x - half * 0.68, y + half * 0.68, x + half * 0.68, y - half * 0.68)
    c.setLineWidth(max(0.8, size * 0.045))
    c.line(x, y, x, y + size * 0.44)
    p = c.beginPath()
    p.moveTo(x, y + size * 0.55)
    p.lineTo(x - size * 0.08, y + size * 0.34)
    p.lineTo(x, y + size * 0.40)
    p.lineTo(x + size * 0.08, y + size * 0.34)
    p.close()
    c.drawPath(p, stroke=1, fill=1)
    c.setFont("Helvetica-Bold", max(6, size * 0.18))
    c.drawCentredString(x, y + size * 0.62, "N")
    c.setFont("Helvetica", max(5, size * 0.145))
    c.drawCentredString(x, y - size * 0.58, "S")
    c.drawCentredString(x + size * 0.54, y - size * 0.04, "E")
    c.drawCentredString(x - size * 0.54, y - size * 0.04, "O")
    c.restoreState()


def generate_pdf(output: Path, info: dict, legends: list[dict], page_size: tuple[float, float] = landscape(A4), source_path: Path | None = None):
    output.parent.mkdir(parents=True, exist_ok=True)
    width, height = page_size
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=page_size)
    format_scale = max(1.0, min(min(width / landscape(A4)[0], height / landscape(A4)[1]), 3.15))
    vertical_scale = format_scale
    if height > width and height > A3[1]:
        vertical_scale = max(format_scale, min(height / landscape(A4)[1], 4.4))
    margin = 18 * min(format_scale, 2.0)
    cart_w = 230 * format_scale
    compact_layout = (height / vertical_scale) < 650
    footer_h = (42 if compact_layout else 54) * vertical_scale
    logo_box_h = (62 if compact_layout else 88) * vertical_scale
    section_h = (30 if compact_layout else 44) * vertical_scale
    project_h = (35 if compact_layout else 50) * vertical_scale
    phone_h = (27 if compact_layout else 34) * vertical_scale
    double_h = (31 if compact_layout else 40) * vertical_scale
    date_h = (34 if compact_layout else 50) * vertical_scale
    format_h = (30 if compact_layout else 42) * vertical_scale
    bottom_limit = margin + footer_h + 6
    template_key = str(info.get("template_id") or "platform_standard")
    platform_template = PLATFORM_TEMPLATES.get(template_key)
    selected_template = get_template(template_key)
    template_color = selected_template["theme_color"] if selected_template else ""
    platform_color = platform_template["color"] if platform_template else ""
    theme_hex = info.get("theme_color") or template_color or platform_color or THEMES.get(info.get("theme", "Bleu BTP"), "#08213A")
    navy = colors.HexColor(theme_hex)
    blue = colors.HexColor(info.get("accent_color") or "#0E4C8A")
    if template_key == "platform_premium":
        logo_box_h *= 1.18
        project_h *= 1.08
        footer_h *= 1.08
    elif template_key == "platform_modern":
        logo_box_h *= 0.82
        section_h *= 1.08
    elif template_key == "platform_topo":
        section_h *= 1.04
        double_h *= 0.96
    elif template_key == "platform_engineering":
        logo_box_h *= 1.04
        section_h *= 0.98
    bottom_limit = margin + footer_h + 6
    x0, y0 = margin, margin
    w, h = width - 2 * margin, height - 2 * margin
    # Keep the cartouche wide enough for its internal professional grid.
    # A previous reduction made the right-side fields exceed the cartouche box
    # on A4 portrait, which visually cut the plan number and lower blocks.
    cart_w = min(cart_w, max(w * 0.27, 220 * format_scale))

    c.setStrokeColor(navy)
    c.setLineWidth(1.8 * min(format_scale, 2.2))
    c.rect(x0, y0, w, h)
    c.setLineWidth(0.8 * min(format_scale, 2.0))
    c.rect(x0 + 3 * format_scale, y0 + 3 * format_scale, w - 6 * format_scale, h - 6 * format_scale)
    side_x = x0 + w - cart_w
    c.line(side_x, y0, side_x, y0 + h)
    c.line(x0, y0 + footer_h, side_x, y0 + footer_h)

    draw_x, draw_y = x0 + 10 * format_scale, y0 + footer_h + 10 * format_scale
    draw_w, draw_h = w - cart_w - 20 * format_scale, h - footer_h - 20 * format_scale
    has_source_pdf = source_path is not None and source_path.exists() and source_path.suffix.lower() == ".pdf"
    has_source_image = source_path is not None and source_path.exists() and source_path.suffix.lower() in (".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".webp")
    if has_source_image:
        try:
            with Image.open(source_path) as img:
                iw, ih = img.size
            img_scale = min(draw_w / iw, draw_h / ih)
            img_w, img_h = iw * img_scale, ih * img_scale
            img_x = draw_x + (draw_w - img_w) / 2
            img_y = draw_y + (draw_h - img_h) / 2
            c.setFillColor(colors.white)
            c.rect(draw_x, draw_y, draw_w, draw_h, fill=1, stroke=0)
            c.drawImage(ImageReader(str(source_path)), img_x, img_y, width=img_w, height=img_h, preserveAspectRatio=True, mask="auto")
        except Exception:
            c.setFillColor(colors.white)
            c.rect(draw_x, draw_y, draw_w, draw_h, fill=1, stroke=0)
    elif not has_source_pdf:
        c.setFillColor(colors.white)
        c.rect(draw_x, draw_y, draw_w, draw_h, fill=1, stroke=0)
    c.setStrokeColor(colors.HexColor("#D5DDE8"))
    c.setLineWidth(0.8)
    c.rect(draw_x, draw_y, draw_w, draw_h)
    compass_size = max(34, min(62 * min(format_scale, 2.0), draw_w * 0.12, draw_h * 0.18))
    draw_compass_rose(c, draw_x + compass_size * 0.88, draw_y + draw_h - compass_size * 0.92, compass_size, colors.HexColor("#111827"))

    def prepare_logo_clean(path: Path) -> Path | None:
        try:
            with Image.open(path) as img:
                img = img.convert("RGBA")
                alpha = img.getchannel("A")
                alpha_bbox = alpha.getbbox()
                pixels = img.load()
                w0, h0 = img.size
                min_x, min_y, max_x, max_y = w0, h0, 0, 0
                found = False
                for yy in range(h0):
                    for xx in range(w0):
                        r, g, b, a = pixels[xx, yy]
                        # Remove transparent and near-white margins around company logos.
                        if a > 12 and not (r > 238 and g > 238 and b > 238):
                            min_x, min_y = min(min_x, xx), min(min_y, yy)
                            max_x, max_y = max(max_x, xx), max(max_y, yy)
                            found = True
                if found:
                    pad_x = max(2, int((max_x - min_x + 1) * 0.04))
                    pad_y = max(2, int((max_y - min_y + 1) * 0.04))
                    min_x = max(0, min_x - pad_x)
                    min_y = max(0, min_y - pad_y)
                    max_x = min(w0 - 1, max_x + pad_x)
                    max_y = min(h0 - 1, max_y + pad_y)
                    img = img.crop((min_x, min_y, max_x + 1, max_y + 1))
                elif alpha_bbox:
                    img = img.crop(alpha_bbox)
                prepared = UPLOADS / f"prepared_logo_{int(time.time() * 1000)}.png"
                img.save(prepared)
                return prepared
        except Exception:
            return None

    def draw_logo_contain(path: Path, box_x: float, box_y: float, box_w: float, box_h: float):
        prepared = prepare_logo_clean(path)
        final_path = prepared or path
        try:
            with Image.open(final_path) as img:
                img_w, img_h = img.size
            # Contain mode: never crop company name/text. Margins are removed first,
            # then the full logo is fitted inside the available rectangle.
            scale = min(box_w / img_w, box_h / img_h)
            final_w = img_w * scale
            final_h = img_h * scale
            final_x = box_x + (box_w - final_w) / 2
            final_y = box_y + (box_h - final_h) / 2
            c.drawImage(str(final_path), final_x, final_y, width=final_w, height=final_h, preserveAspectRatio=True, mask="auto")
            return True
        except Exception:
            return False

    if not has_source_pdf and not has_source_image:
        # Technical drawing area: clean demo preview only when no real PDF is imported.
        c.setStrokeColor(colors.HexColor("#D8DEE8"))
        c.setLineWidth(0.7)
        for i in range(6):
            xx = draw_x + 55 + i * (draw_w - 110) / 5
            c.line(xx, draw_y + 25, xx, draw_y + draw_h - 25)
            c.circle(xx, draw_y + draw_h - 22, 9)
            c.setFont("Helvetica-Bold", 7)
            c.setFillColor(navy)
            c.drawCentredString(xx, draw_y + draw_h - 25, chr(65 + i))
        for i in range(5):
            yy = draw_y + 55 + i * (draw_h - 110) / 4
            c.line(draw_x + 25, yy, draw_x + draw_w - 25, yy)
            c.circle(draw_x + 24, yy, 9)
            c.setFont("Helvetica-Bold", 7)
            c.drawCentredString(draw_x + 24, yy - 3, str(i + 1))

        c.setStrokeColor(colors.HexColor("#1F2937"))
        c.setLineWidth(2.6)
        bx, by = draw_x + 90, draw_y + 88
        bw, bh = draw_w - 175, draw_h - 150
        c.rect(bx, by, bw, bh)
        c.line(bx + bw * .28, by, bx + bw * .28, by + bh)
        c.line(bx + bw * .58, by, bx + bw * .58, by + bh)
        c.line(bx, by + bh * .48, bx + bw, by + bh * .48)
        c.line(bx + bw * .58, by + bh * .48, bx + bw * .58, by + bh)
        c.setFillColor(navy)
        for px, py, label in [
            (bx, by, "P1"), (bx + bw * .28, by, "P2"), (bx + bw * .58, by, "P3"), (bx + bw, by, "P4"),
            (bx, by + bh * .48, "P5"), (bx + bw * .58, by + bh * .48, "P6"), (bx + bw, by + bh * .48, "P7"),
            (bx, by + bh, "P8"), (bx + bw * .28, by + bh, "P9"), (bx + bw * .58, by + bh, "P10"), (bx + bw, by + bh, "P11"),
        ]:
            c.rect(px - 5, py - 5, 10, 10, fill=1, stroke=0)
            c.setFont("Helvetica", 6)
            c.drawString(px + 7, py + 7, label)

        c.setStrokeColor(colors.HexColor("#6B7280"))
        c.setLineWidth(0.6)
        c.line(bx, by - 22, bx + bw, by - 22)
        c.line(bx, by - 17, bx, by - 27)
        c.line(bx + bw, by - 17, bx + bw, by - 27)
        c.setFont("Helvetica-Bold", 8)
        c.setFillColor(navy)
        c.drawCentredString(bx + bw / 2, by - 36, "13,80")

    c.setFont("Helvetica-Bold", 16)
    c.setFillColor(navy)
    c.drawCentredString(draw_x + draw_w / 2, draw_y + 18, str(info["plan_type"]).upper())
    c.setFont("Helvetica", 9)
    c.drawCentredString(draw_x + draw_w / 2, draw_y + 4, f"ECHELLE : {info['scale']}")
    if template_key == "platform_premium":
        c.setFillColor(colors.HexColor("#07192E"))
        c.rect(draw_x, draw_y, draw_w, 34 * vertical_scale, fill=1, stroke=0)
        c.setFillColor(navy)
        c.rect(draw_x, draw_y + 34 * vertical_scale, draw_w, 3 * vertical_scale, fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 8 * format_scale)
        c.drawCentredString(draw_x + draw_w / 2, draw_y + 20 * vertical_scale, str(info["plan_type"]).upper())
        c.setFont("Helvetica", 6 * format_scale)
        c.drawCentredString(draw_x + draw_w / 2, draw_y + 9 * vertical_scale, f"ECHELLE : {info['scale']}")
    elif template_key == "platform_modern":
        c.setStrokeColor(colors.HexColor("#94A3B8"))
        c.line(draw_x + draw_w * 0.22, draw_y + 26 * vertical_scale, draw_x + draw_w * 0.78, draw_y + 26 * vertical_scale)
    elif template_key == "platform_topo":
        c.setFillColor(colors.HexColor("#0F766E"))
        c.circle(draw_x + 18 * format_scale, draw_y + 18 * vertical_scale, 5 * format_scale, fill=1, stroke=0)
        c.circle(draw_x + draw_w - 18 * format_scale, draw_y + 18 * vertical_scale, 5 * format_scale, fill=1, stroke=0)

    # Side panel
    cart_x = side_x
    def ux(value: float) -> float:
        return cart_x + value * format_scale

    def uw(value: float) -> float:
        return value * format_scale

    def uh(value: float) -> float:
        return value * vertical_scale

    def fs(value: float) -> float:
        return value * format_scale

    c.setFillColor(colors.white)
    c.rect(cart_x, y0, cart_w, h, fill=1, stroke=0)
    if platform_template:
        if template_key == "platform_premium":
            c.setFillColor(colors.HexColor("#FBF7EC"))
            c.rect(cart_x, y0, cart_w, h, fill=1, stroke=0)
            c.setFillColor(colors.HexColor("#07192E"))
            c.rect(cart_x, y0 + h - uh(124), cart_w, uh(124), fill=1, stroke=0)
            c.setFillColor(navy)
            c.rect(cart_x, y0 + h - uh(128), cart_w, uh(6), fill=1, stroke=0)
            c.rect(cart_x + uw(12), y0 + footer_h + uh(8), cart_w - uw(24), uh(4), fill=1, stroke=0)
        elif template_key == "platform_modern":
            c.setFillColor(colors.HexColor("#FAFBFC"))
            c.rect(cart_x, y0, cart_w, h, fill=1, stroke=0)
            c.setFillColor(colors.HexColor("#E2E8F0"))
            c.rect(cart_x, y0, uw(30), h, fill=1, stroke=0)
            c.setStrokeColor(colors.HexColor("#CBD5E1"))
            c.setLineWidth(uw(0.6))
            for i in range(1, 6):
                c.line(cart_x, y0 + i * h / 6, cart_x + cart_w, y0 + i * h / 6)
        elif template_key == "platform_topo":
            c.setFillColor(colors.HexColor("#F0FDFA"))
            c.rect(cart_x, y0, cart_w, h, fill=1, stroke=0)
            c.setFillColor(colors.HexColor("#CCFBF1"))
            c.rect(cart_x, y0 + footer_h, uw(7), h - footer_h, fill=1, stroke=0)
            c.setFillColor(colors.HexColor("#0F766E"))
            for i in range(6):
                c.circle(cart_x + uw(15), y0 + footer_h + uh(38 + i * 42), uw(3), fill=1, stroke=0)
        elif template_key == "platform_engineering":
            c.setFillColor(colors.HexColor("#F8FAFC"))
            c.rect(cart_x, y0, cart_w, h, fill=1, stroke=0)
            c.setStrokeColor(colors.HexColor("#DBEAFE"))
            c.setLineWidth(uw(0.4))
            step = uw(18)
            xx = cart_x
            while xx < cart_x + cart_w:
                c.line(xx, y0, xx, y0 + h)
                xx += step
            yy = y0
            while yy < y0 + h:
                c.line(cart_x, yy, cart_x + cart_w, yy)
                yy += step
            c.setFillColor(colors.HexColor("#DBEAFE"))
            c.rect(cart_x + cart_w - uw(9), y0 + footer_h, uw(5), h - footer_h, fill=1, stroke=0)
    # A user-imported cartouche is treated as a style reference.
    # It must not be pasted as a background, otherwise fields overlap and the
    # final PDF stops behaving like a clean dynamic technical template.
    template_path = Path(selected_template["source_file"]) if selected_template else None
    logo_y = y0 + h - logo_box_h - uh(16)
    logo_path = Path(info.get("logo_path", "")) if info.get("logo_path") else None
    c.setStrokeColor(colors.HexColor("#D6DEE8"))
    logo_box_x = ux(12)
    logo_box_y = logo_y
    logo_box_w = cart_w - uw(24)
    if template_key == "platform_premium":
        c.setStrokeColor(navy)
        c.setFillColor(colors.Color(1, 1, 1, alpha=0.08))
        c.roundRect(logo_box_x, logo_box_y, logo_box_w, logo_box_h, uw(8), fill=1, stroke=1)
    elif template_key == "platform_modern":
        c.setStrokeColor(colors.HexColor("#94A3B8"))
        c.roundRect(logo_box_x + uw(24), logo_box_y, logo_box_w - uw(24), logo_box_h, uw(2), fill=0, stroke=1)
    elif template_key == "platform_topo":
        c.setStrokeColor(colors.HexColor("#0F766E"))
        c.roundRect(logo_box_x, logo_box_y, logo_box_w, logo_box_h, uw(14), fill=0, stroke=1)
    else:
        c.roundRect(logo_box_x, logo_box_y, logo_box_w, logo_box_h, uw(6), fill=0, stroke=1)
    logo_drawn = bool(logo_path and logo_path.exists() and draw_logo_contain(logo_path, logo_box_x + uw(7), logo_box_y + uw(7), logo_box_w - uw(14), logo_box_h - uw(14)))
    if not logo_drawn:
        c.setFillColor(navy)
        center_x = logo_box_x + logo_box_w / 2
        bar_h = uw(36 if compact_layout else 44)
        c.rect(center_x - uw(46), logo_box_y + uh(14), uw(14), bar_h, fill=1, stroke=0)
        c.rect(center_x - uw(26), logo_box_y + uh(7), uw(14), bar_h + uh(10), fill=1, stroke=0)
        c.rect(center_x - uw(6), logo_box_y + uh(12), uw(14), bar_h + uh(3), fill=1, stroke=0)
        c.setFont("Helvetica-Bold", fs(17 if compact_layout else 19))
        c.drawString(center_x + uw(16), logo_box_y + logo_box_h - uh(29), str(info.get("company_short") or "SKE")[:10])
        c.setFont("Helvetica", fs(8 if compact_layout else 10))
        c.drawString(center_x + uw(18), logo_box_y + logo_box_h - uh(43), "SOLUTIONS")
    c.setFillColor(colors.white if template_key == "platform_premium" else colors.HexColor("#111827"))
    c.setFont("Helvetica", fs(5.8 if compact_layout else 6.8))
    company_line = f"{str(info.get('company') or 'SKE System')[:24]}  |  {str(info.get('company_email') or 'contact@btpsmarttools.com')[:24]}"
    c.drawCentredString(cart_x + cart_w / 2, logo_box_y - uh(9), company_line)

    def section(y: float, label: str, value: str, height_box: float = 42):
        c.setStrokeColor(colors.HexColor("#B8C4D2"))
        c.line(cart_x, y, cart_x + cart_w, y)
        if template_key == "platform_modern":
            c.setFillColor(colors.white)
            c.roundRect(ux(42), y - uh(18), cart_w - uw(54), uh(30), uw(4), fill=1, stroke=0)
            c.setStrokeColor(colors.HexColor("#CBD5E1"))
            c.roundRect(ux(42), y - uh(18), cart_w - uw(54), uh(30), uw(4), fill=0, stroke=1)
            c.setFillColor(navy)
            c.setFont("Helvetica-Bold", fs(5.5 if compact_layout else 6.2))
            c.drawString(ux(46), y - uh(8), label)
            value_y = y - uh(21 if compact_layout else 27)
        elif template_key == "platform_premium":
            c.setFillColor(colors.HexColor("#07192E"))
            c.roundRect(ux(9), y - uh(19), uw(88), uh(14), uw(3), fill=1, stroke=0)
            c.setFillColor(navy)
            c.rect(ux(9), y - uh(21), uw(88), uh(2), fill=1, stroke=0)
            c.setFillColor(colors.white)
            c.setFont("Helvetica-Bold", fs(5.8 if compact_layout else 6.5))
            c.drawString(ux(13), y - uh(14), label)
            value_y = y - uh(24 if compact_layout else 34)
        elif template_key == "platform_topo":
            c.setFillColor(colors.HexColor("#0F766E"))
            c.rect(ux(9), y - uh(17), uw(8), uh(12), fill=1, stroke=0)
            c.setFillColor(colors.HexColor("#DDFCF5"))
            c.rect(ux(20), y - uh(17), uw(82), uh(12), fill=1, stroke=0)
            c.setFillColor(colors.HexColor("#064E3B"))
            c.setFont("Helvetica-Bold", fs(5.8 if compact_layout else 6.5))
            c.drawString(ux(24), y - uh(14), label)
            value_y = y - uh(24 if compact_layout else 34)
        else:
            c.setFillColor(navy)
            c.rect(ux(9), y - uh(17), uw(82), uh(12), fill=1, stroke=0)
            c.setFillColor(colors.white)
            c.setFont("Helvetica-Bold", fs(5.8 if compact_layout else 6.5))
            c.drawString(ux(13), y - uh(14), label)
            value_y = y - uh(24 if compact_layout else 34)
        c.setFillColor(colors.HexColor("#111827"))
        c.setFont("Helvetica-Bold", fs(8 if compact_layout else 9.5))
        c.drawCentredString(cart_x + cart_w / 2, value_y, str(value)[:36])
        return y - height_box

    y = logo_box_y - uh(18 if compact_layout else 20)
    y = section(y, "PROJET", info["project"], project_h)
    y = section(y, "LOCALISATION", "GABON", section_h)
    y = section(y, "TYPE DE PLAN", info["plan_type"], section_h)
    y = section(y, "MAITRE D'OUVRAGE", info["company"], section_h)
    y = section(y, "TELEPHONE", info.get("company_phone", ""), phone_h)

    c.line(cart_x, y, cart_x + cart_w, y)
    c.setFillColor(navy)
    c.rect(ux(9), y - uw(17), uw(55), uw(12), fill=1, stroke=0)
    c.rect(ux(117), y - uw(17), uw(55), uw(12), fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", fs(5.8 if compact_layout else 6.5))
    c.drawString(ux(13), y - uw(14), "DATE")
    c.drawString(ux(121), y - uw(14), "N PLAN")
    c.setFillColor(colors.HexColor("#111827"))
    c.setFont("Helvetica-Bold", fs(8.5 if compact_layout else 10))
    c.drawString(ux(14), y - uw(26 if compact_layout else 36), datetime.now().strftime("%d/%m/%Y"))
    c.setFont("Helvetica-Bold", fs(12 if compact_layout else 15))
    c.drawString(ux(122), y - uw(27 if compact_layout else 37), str(info["number"])[:12])
    y -= date_h

    c.line(cart_x, y, cart_x + cart_w, y)
    c.setFillColor(navy)
    c.rect(ux(9), y - uw(17), uw(55), uw(12), fill=1, stroke=0)
    c.rect(ux(117), y - uw(17), uw(55), uw(12), fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", fs(5.8 if compact_layout else 6.5))
    c.drawString(ux(13), y - uw(14), "FORMAT")
    c.drawString(ux(121), y - uw(14), "REVISION")
    c.setFillColor(colors.HexColor("#111827"))
    c.setFont("Helvetica-Bold", fs(7.5 if compact_layout else 9))
    c.drawString(ux(14), y - uw(24 if compact_layout else 36), str(info.get("format_plan", "A4"))[:14])
    c.drawString(ux(122), y - uw(24 if compact_layout else 36), str(info.get("revision", "REV 00"))[:14])
    y -= format_h

    c.line(cart_x, y, cart_x + cart_w, y)
    c.setFillColor(navy)
    c.rect(ux(9), y - uw(17), uw(70), uw(12), fill=1, stroke=0)
    c.rect(ux(117), y - uw(17), uw(70), uw(12), fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", fs(5.8 if compact_layout else 6.5))
    c.drawString(ux(13), y - uw(14), "CHEF PROJET")
    c.drawString(ux(121), y - uw(14), "OPERATEUR")
    c.setFillColor(colors.HexColor("#111827"))
    c.setFont("Helvetica-Bold", fs(6.8 if compact_layout else 8))
    c.drawString(ux(14), y - uw(24 if compact_layout else 36), str(info.get("project_manager", ""))[:17])
    c.drawString(ux(122), y - uw(24 if compact_layout else 36), str(info.get("operator", ""))[:17])
    y -= double_h

    c.line(cart_x, y, cart_x + cart_w, y)
    c.setFillColor(navy)
    c.rect(ux(9), y - uw(17), uw(70), uw(12), fill=1, stroke=0)
    c.rect(ux(117), y - uw(17), uw(70), uw(12), fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", fs(5.8 if compact_layout else 6.5))
    c.drawString(ux(13), y - uw(14), "DESSINATEUR")
    c.drawString(ux(121), y - uw(14), "VALIDATEUR")
    c.setFillColor(colors.HexColor("#111827"))
    c.setFont("Helvetica-Bold", fs(6.8 if compact_layout else 8))
    c.drawString(ux(14), y - uw(24 if compact_layout else 36), str(info.get("author", ""))[:17])
    c.drawString(ux(122), y - uw(24 if compact_layout else 36), str(info.get("validator", ""))[:17])
    y -= double_h

    # Legend and notes
    if info.get("show_legend", "on") == "on" and y > bottom_limit + (48 if compact_layout else 80):
        legend_items = 3 if compact_layout else 4
        note_items = 2 if compact_layout else 4
        legend_step = uw(10 if compact_layout else 13)
        legend_h = uw(50 if compact_layout else 92)
        c.line(cart_x, y, cart_x + cart_w, y)
        c.setFillColor(navy)
        c.rect(ux(9), y - uw(17), uw(60), uw(12), fill=1, stroke=0)
        c.rect(ux(112), y - uw(17), uw(82), uw(12), fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", fs(5.8 if compact_layout else 6.5))
        c.drawString(ux(13), y - uw(14), "LEGENDE")
        c.drawString(ux(116), y - uw(14), "NOTES GENERALES")
        c.setFont("Helvetica", fs(5.5 if compact_layout else 6.5))
        ly = y - uw(28 if compact_layout else 33)
        for item in legends[:legend_items]:
            color = item.get("color", "#2487ff")
            try:
                c.setFillColor(colors.HexColor(color))
            except Exception:
                c.setFillColor(blue)
            c.rect(ux(15), ly - uw(4), uw(6), uw(6), fill=1, stroke=0)
            c.setFillColor(colors.black)
            c.drawString(ux(27), ly - uw(3), str(item.get("text", ""))[:18])
            ly -= legend_step
        c.setFillColor(colors.black)
        for i, note in enumerate(["Dimensions en metres.", "Beton conforme BTP.", "Verifier echelle.", "Document genere IA."][:note_items]):
            c.drawString(ux(116), y - uw(28 if compact_layout else 33) - i * legend_step, f"- {note}")
        y -= legend_h

    # Elements table
    if info.get("show_elements", "on") == "on" and y > bottom_limit + (45 if compact_layout else 78):
        element_columns = info.get("element_columns") or ["REPERE", "DESIGNATION", "DIM.", "QTE"]
        element_rows = info.get("element_rows") or [
            {"c1": "P1-P6", "c2": "Poteau", "c3": "30x30", "c4": "06"},
            {"c1": "L1", "c2": "Longrine", "c3": "30x40", "c4": "08"},
            {"c1": "L2", "c2": "Longrine", "c3": "25x30", "c4": "04"},
        ]
        row_h = uw(11 if compact_layout else 15)
        max_possible_rows = int(max(1, (y - bottom_limit - 34) // row_h))
        reserve_revisions = info.get("show_revisions", "on") == "on" and (format_scale > 2.0 or y < bottom_limit + uw(160 if not compact_layout else 115))
        row_cap = 2 if compact_layout or reserve_revisions else 6
        max_rows = max(1, min(row_cap, len(element_rows), max_possible_rows))
        table_h = uw(18) + max_rows * row_h
        c.line(cart_x, y, cart_x + cart_w, y)
        c.setFillColor(navy)
        c.rect(ux(9), y - uw(17), uw(95), uw(12), fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", fs(5.8 if compact_layout else 6.5))
        c.drawString(ux(13), y - uw(14), "TABLEAU DES ELEMENTS")
        table_y = y - uw(25 if compact_layout else 30)
        cols = [ux(10), ux(57), ux(122), ux(177), cart_x + cart_w - uw(10)]
        c.setStrokeColor(colors.HexColor("#B8C4D2"))
        for i in range(max_rows + 2):
            yy = table_y - i * row_h
            c.line(ux(10), yy, cart_x + cart_w - uw(10), yy)
        for xx in cols:
            c.line(xx, table_y, xx, table_y - (max_rows + 1) * row_h)
        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", fs(4.8 if compact_layout else 5.4))
        for j, val in enumerate(element_columns[:4]):
            c.drawString(cols[j] + uw(3), table_y - uw(8 if compact_layout else 10), str(val)[:14])
        c.setFont("Helvetica", fs(5 if compact_layout else 5.8))
        for i, row in enumerate(element_rows[:max_rows]):
            yy = table_y - uw(19 if compact_layout else 25) - i * row_h
            vals = [row.get("c1", ""), row.get("c2", ""), row.get("c3", ""), row.get("c4", "")]
            for j, val in enumerate(vals):
                c.drawString(cols[j] + uw(3), yy, str(val)[:16])
        y -= table_h + uw(16 if compact_layout or reserve_revisions else 30)

    # Revisions
    tight_revisions = y < bottom_limit + uw(58)
    if info.get("show_revisions", "on") == "on":
        if y < bottom_limit + uw(42):
            y = bottom_limit + uw(42)
        revision_rows = info.get("revision_rows") or [
            {
                "indice": info.get("revision", "REV 00"),
                "date": datetime.now().strftime("%d/%m/%Y"),
                "description": "Premiere emission",
            }
        ]
        rev_rows = max(1, min(len(revision_rows) + 1, 2 if (compact_layout or tight_revisions) else 4))
        rev_row_h = uw(12 if (compact_layout or tight_revisions) else 16)
        c.line(cart_x, y, cart_x + cart_w, y)
        c.setFillColor(navy)
        c.rect(ux(9), y - uw(17), uw(62), uw(12), fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", fs(5.8 if (compact_layout or tight_revisions) else 6.5))
        c.drawString(ux(13), y - uw(14), "REVISIONS")
        rev_y = y - uw(24 if (compact_layout or tight_revisions) else 30)
        c.setStrokeColor(colors.HexColor("#B8C4D2"))
        for i in range(rev_rows + 1):
            yy = rev_y - i * rev_row_h
            c.line(ux(10), yy, cart_x + cart_w - uw(10), yy)
        for xx in [ux(10), ux(60), ux(115), cart_x + cart_w - uw(10)]:
            c.line(xx, rev_y, xx, rev_y - rev_rows * rev_row_h)
        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", fs(5 if (compact_layout or tight_revisions) else 5.8))
        c.drawString(ux(18), rev_y - uw(9 if (compact_layout or tight_revisions) else 11), "INDICE")
        c.drawString(ux(70), rev_y - uw(9 if (compact_layout or tight_revisions) else 11), "DATE")
        c.drawString(ux(126), rev_y - uw(9 if (compact_layout or tight_revisions) else 11), "DESCRIPTION")
        c.setFont("Helvetica", fs(4.8 if (compact_layout or tight_revisions) else 5.5))
        for idx, row in enumerate(revision_rows[: max(1, rev_rows - 1)]):
            row_y = rev_y - (idx + 1) * rev_row_h - uw(8 if (compact_layout or tight_revisions) else 10)
            c.drawString(ux(15), row_y, str(row.get("indice", ""))[:10])
            c.drawString(ux(64), row_y, str(row.get("date", ""))[:12])
            c.drawString(ux(119), row_y, str(row.get("description", ""))[:28])

    # Footer quality band
    c.setFillColor(navy)
    c.rect(side_x, y0, cart_w, footer_h, fill=1, stroke=0)
    labels = [("RAPIDITE", "Generation auto"), ("PRECISION", "Cartouches BTP"), ("CONFORMITE", "Standards pro"), ("IA", "Aide integree")]
    part = cart_w / 4
    for i, (a, b) in enumerate(labels):
        xx = side_x + i * part
        if i:
            c.setStrokeColor(colors.HexColor("#5A7898"))
            c.line(xx, y0 + uw(8), xx, y0 + footer_h - uw(8))
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", fs(5.5 if compact_layout else 6.3))
        c.drawCentredString(xx + part / 2, y0 + uw(23 if compact_layout else 28), a)
        c.setFont("Helvetica", fs(4.9 if compact_layout else 5.4))
        c.drawCentredString(xx + part / 2, y0 + uw(12 if compact_layout else 16), b)

    c.setFillColor(navy)
    c.setFont("Helvetica-Bold", fs(7))
    c.drawString(x0 + uw(18), y0 + uw(24), "GENERE PAR BTP SMART TOOLS")
    c.setFont("Helvetica", fs(6))
    c.drawString(x0 + uw(18), y0 + uw(12), "La technologie au service de vos plans")
    c.drawCentredString(x0 + (side_x - x0) / 2, y0 + 12, "2026 BTP SMART TOOLS - by SKE System")
    c.save()
    sheet_pdf = buffer.getvalue()
    if has_source_pdf:
        try:
            merge_source_pdf_into_sheet(sheet_pdf, source_path, output, (draw_x + uw(6), draw_y + uw(28), draw_w - uw(12), draw_h - uw(38)), page_size)
        except Exception:
            output.write_bytes(sheet_pdf)
    else:
        output.write_bytes(sheet_pdf)


class App(BaseHTTPRequestHandler):
    server_version = "SKE/0.1"

    def get_user(self) -> sqlite3.Row | None:
        raw = self.headers.get("Cookie", "")
        jar = cookies.SimpleCookie(raw)
        token = jar.get("ske_session")
        if not token or token.value not in SESSIONS:
            return None
        user_id = SESSIONS[token.value]
        with db() as con:
            return con.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()

    def send_html(self, title: str, body: str, status: int = 200):
        data = render_page(title, body, self.get_user())
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def redirect(self, path: str):
        self.send_response(302)
        self.send_header("Location", path)
        self.end_headers()

    def require_login(self) -> sqlite3.Row | None:
        user = self.get_user()
        if not user:
            self.redirect("/login")
            return None
        return user

    def do_GET(self):
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path
        if path == "/":
            return self.home()
        if path == "/login":
            return self.login_page()
        if path == "/register":
            return self.register_page()
        if path == "/logout":
            return self.logout()
        if path == "/dashboard":
            return self.dashboard()
        if path == "/workspace":
            return self.redirect("/generator")
        if path == "/generator":
            return self.generator()
        if path == "/cartouches":
            return self.generator()
        if path == "/assistant":
            return self.assistant()
        if path == "/templates":
            return self.templates_page()
        if path.startswith("/template-file/"):
            return self.template_file(path.split("/", 2)[2])
        if path.startswith("/template-analysis/"):
            return self.template_analysis(path.split("/", 2)[2])
        if path == "/profile-analyzer":
            return self.redirect("/generator")
        if path.startswith("/profile-analysis/"):
            return self.redirect("/generator")
        if path.startswith("/profile-source/"):
            return self.redirect("/generator")
        if path == "/batch":
            return self.batch_page()
        if path == "/services":
            return self.redirect("/generator")
        if path == "/payment":
            return self.payment_page()
        if path.startswith("/payment/check/"):
            return self.payment_check(path.split("/", 3)[3])
        if path.startswith("/payment/qr/"):
            return self.payment_qr(path.split("/", 3)[3])
        if path == "/payment/success":
            return self.payment_result("success")
        if path == "/payment/failed":
            return self.payment_result("failed")
        if path == "/pvit/secret-receiver":
            return self.pvit_secret_receiver()
        if path == "/pvit/secret-receiver-open":
            return self.pvit_secret_receiver()
        if path == "/pvit/secret-status":
            return self.pvit_secret_status()
        if path == "/admin":
            return self.admin()
        if path == "/admin/validate-payment":
            return self.validate_payment()
        if path.startswith("/download/"):
            return self.download(path.split("/", 2)[2])
        self.send_html("Page introuvable", "<div class='card'><h2>Page introuvable</h2></div>", 404)

    def do_POST(self):
        path = urllib.parse.urlparse(self.path).path
        if path == "/login":
            return self.login_action()
        if path == "/register":
            return self.register_action()
        if path == "/generate":
            return self.generate_action()
        if path == "/assistant":
            return self.assistant_action()
        if path == "/workspace":
            return self.redirect("/generator")
        if path == "/templates":
            return self.templates_action()
        if path == "/profile-analyzer":
            return self.redirect("/generator")
        if path == "/batch":
            return self.batch_action()
        if path == "/payment":
            return self.payment_action()
        if path == "/payment/callback":
            return self.payment_callback()
        if path == "/pvit/secret-receiver":
            return self.pvit_secret_receiver()
        if path == "/pvit/secret-receiver-open":
            return self.pvit_secret_receiver()
        if path == "/admin/validate-payment":
            return self.validate_payment()
        self.send_html("Action introuvable", "<div class='card'><h2>Action introuvable</h2></div>", 404)

    def read_form(self) -> dict:
        length = int(self.headers.get("Content-Length", 0))
        body = self.rfile.read(length).decode("utf-8")
        return {k: v[0] if v else "" for k, v in urllib.parse.parse_qs(body).items()}

    def coming_soon(self, name: str):
        body = f"""
        <div class="card">
          <span class="pill">Indisponible</span>
          <h2>{html.escape(name)}</h2>
          <p>Ce module n'est pas disponible sur le site public pour le moment. BTP Smart Tools est concentre uniquement sur les cartouches professionnelles : import de modeles, generation PDF et traitement batch.</p>
          <a class="btn green" href="/generator">Aller au generateur de cartouches</a>
          <a class="btn" href="/templates">Importer une cartouche modele</a>
        </div>"""
        self.send_html(name, body)

    def home(self):
        body = cartouche_focus_panel() + """
        <section class="hero">
          <div>
            <span class="pill">Service actif : cartouches professionnelles</span>
            <h1>Cartouches PDF</h1>
            <h2>Le service principal de BTP Smart Tools.</h2>
            <p>La plateforme est concentree sur les cartouches PDF : cadres, logos, legendes, tableaux, formats, echelle et mise en page professionnelle.</p>
            <div class="row"><a class="btn green" href="/register">Creer un compte</a><a class="btn" href="/login">Acceder au compte</a><a class="btn dark" href="/generator">Generer une cartouche</a></div>
            <div class="feature-row">
              <div class="feature"><b>Actif maintenant</b><span>Generation de cartouches et PDF propres.</span></div>
              <div class="feature"><b>Compte client</b><span>Credits, abonnements, historique et modeles importes.</span></div>
              <div class="feature"><b>Controle avant PDF</b><span>Format, echelle, logo, legende et modele visibles avant generation.</span></div>
            </div>
          </div>
          <div class="sheet">
            <div class="draw">
              <h2>Plan final pret a envoyer</h2>
              <p>Cadre, cartouche, legende et informations projet.</p>
              <svg viewBox="0 0 440 280" width="100%" height="280">
                <rect x="25" y="25" width="360" height="220" fill="none" stroke="#111827" stroke-width="2"/>
                <path d="M70 220 L70 92 L145 92 L145 60 L310 60 L310 220 Z" fill="none" stroke="#111827" stroke-width="4"/>
                <path d="M98 220 L98 118 L166 118 L166 88 L285 88 L285 220" fill="none" stroke="#2f8cff" stroke-width="3"/>
                <line x1="45" y1="150" x2="370" y2="150" stroke="#94a3b8" stroke-dasharray="8 8"/>
                <circle cx="98" cy="118" r="5" fill="#dc2626"/>
                <circle cx="285" cy="88" r="5" fill="#dc2626"/>
                <text x="130" y="172" fill="#10b981" font-size="15">zone de projet</text>
              </svg>
            </div>
            <div class="cartouche"><div class="ske">BTP</div><div>PROJET<br><b>Jardin Botanique</b></div><div>ENTREPRISE<br><b>SKE System</b></div><div>TYPE PLAN<br><b>Plan beton</b></div><div>ECHELLE<br><b>1/100</b></div><div>DATE<br><b>11/05/2026</b></div><div>N PLAN<br><b>BTP-001</b></div></div>
          </div>
        </section>
        """ + cartouche_steps_html() + """
        <div class="grid3" style="margin-top:20px">
          <div class="card"><span class="pill">Ponctuel</span><h2>1 200 FCFA</h2><p>1 credit = 1 PDF avec cartouche.</p></div>
          <div class="card"><span class="pill">Decouverte</span><h2>5 000 FCFA</h2><p>10 credits : analyse d'une cartouche + generations PDF.</p></div>
          <div class="card"><span class="pill">Mensuel</span><h2>12 000 FCFA</h2><p>30 credits inclus pour cartouches, analyses IA et PDF.</p></div>
          <div class="card"><span class="pill">Entreprise</span><h2>250 000 FCFA</h2><p>1500 credits/an pour bureaux d'etudes, entreprises BTP et grands volumes.</p></div>
        </div>
        <div class="grid3" style="margin-top:20px">
          <div class="card roadmap-card"><span class="pill">Disponible</span><h3>Cartouches automatiques</h3><p>PDF, logo, formats, legendes et tableaux.</p></div>
          <div class="card roadmap-card"><span class="pill">Disponible</span><h3>Modeles personnalises</h3><p>Import d'une cartouche d'entreprise, analyse une seule fois, reutilisation ensuite.</p></div>
          <div class="card roadmap-card"><span class="pill">Disponible</span><h3>Batch PDF</h3><p>Traitement de plusieurs PDF avec la meme cartouche et numeros automatiques.</p></div>
        </div>"""
        self.send_html("Accueil", body)

    def login_page(self, error: str = ""):
        alert = f"<p class='alert'>{html.escape(error)}</p>" if error else ""
        body = f"""
        <div class="grid">
          <form class="card" method="post" action="/login">
            <h2>Connexion</h2>{alert}
            <label>Email</label><input name="email" type="email" placeholder="votre@email.com" autocomplete="email">
            <label>Mot de passe</label><input name="password" type="password" placeholder="Votre mot de passe" autocomplete="current-password">
            <button class="green">Se connecter</button>
            <p><a class="btn dark" href="/register">Creer un compte client</a></p>
          </form>
          <div class="card"><h2>Acces compte</h2><p>Connecte-toi avec ton compte client pour acheter des credits, importer des cartouches et retrouver tes PDF generes.</p><p class="muted">L'acces administrateur est reserve au fondateur et n'est pas affiche publiquement.</p></div>
        </div>"""
        self.send_html("Connexion", body)

    def register_page(self, error: str = ""):
        alert = f"<p class='alert'>{html.escape(error)}</p>" if error else ""
        body = f"""
        <div class="grid">
          <form class="card" method="post" action="/register">
            <h2>Creer un compte</h2>{alert}
            <p class="muted">Le compte permet d'acheter des credits, importer des cartouches personnalisees et retrouver les PDF generes.</p>
            <label>Nom / Entreprise</label><input name="name" placeholder="Entreprise ou nom complet" required>
            <label>Email</label><input name="email" type="email" placeholder="client@email.com" required>
            <label>Mot de passe</label><input name="password" type="password" minlength="6" required>
            <button class="green">Creer mon compte</button>
            <p class="muted">Apres creation, choisis un pack ou un abonnement pour recevoir tes credits.</p>
          </form>
          <div class="card">
            <h2>Parcours client</h2>
            <p>1. Creation du compte.</p>
            <p>2. Achat d'un pack ou abonnement.</p>
            <p>3. Import du PDF ou de la cartouche modele.</p>
            <p>4. Detection automatique si le PDF contient l'echelle.</p>
            <p>5. Generation du PDF final avec cartouche.</p>
          </div>
        </div>"""
        self.send_html("Creer un compte", body)

    def register_action(self):
        form = self.read_form()
        name = (form.get("name") or "").strip()
        email = (form.get("email") or "").strip().lower()
        password = form.get("password") or ""
        if not name or not email or len(password) < 6:
            return self.register_page("Complete le nom, l'email et un mot de passe de 6 caracteres minimum.")
        with db() as con:
            exists = con.execute("SELECT id FROM users WHERE email=?", (email,)).fetchone()
            if exists:
                return self.register_page("Ce compte existe deja. Connecte-toi ou utilise un autre email.")
            cur = con.execute(
                "INSERT INTO users(email,password_hash,name,role,credits,subscription,subscription_until,created_at) VALUES(?,?,?,?,?,?,?,?)",
                (email, hash_password(password), name, "user", 0, "none", "", now()),
            )
            user_id = cur.lastrowid
        token = secrets.token_urlsafe(32)
        SESSIONS[token] = user_id
        self.send_response(302)
        self.send_header("Location", "/payment")
        self.send_header("Set-Cookie", f"ske_session={token}; Path=/; HttpOnly; SameSite=Lax")
        self.end_headers()

    def login_action(self):
        form = self.read_form()
        with db() as con:
            user = con.execute("SELECT * FROM users WHERE email=?", (form.get("email", ""),)).fetchone()
        if not user or not verify_password(form.get("password", ""), user["password_hash"]):
            return self.login_page("Email ou mot de passe incorrect.")
        token = secrets.token_urlsafe(32)
        SESSIONS[token] = user["id"]
        self.send_response(302)
        self.send_header("Location", "/admin" if user["role"] == "admin" else "/dashboard")
        self.send_header("Set-Cookie", f"ske_session={token}; Path=/; HttpOnly; SameSite=Lax")
        self.end_headers()

    def logout(self):
        raw = self.headers.get("Cookie", "")
        jar = cookies.SimpleCookie(raw)
        token = jar.get("ske_session")
        if token:
            SESSIONS.pop(token.value, None)
        self.send_response(302)
        self.send_header("Location", "/")
        self.send_header("Set-Cookie", "ske_session=; Path=/; Max-Age=0")
        self.end_headers()

    def dashboard(self):
        user = self.require_login()
        if not user:
            return
        with db() as con:
            gens = con.execute("SELECT * FROM generations WHERE user_id=? ORDER BY id DESC LIMIT 10", (user["id"],)).fetchall()
            templates = con.execute("SELECT * FROM cartouche_templates WHERE user_id=? ORDER BY id DESC LIMIT 8", (user["id"],)).fetchall()
            batches = con.execute("SELECT * FROM batches WHERE user_id=? ORDER BY id DESC LIMIT 5", (user["id"],)).fetchall()
        rows = "".join(f"<tr><td>{g['created_at']}</td><td>{html.escape(g['plan_number'] or '')}</td><td>{html.escape(g['project'])}</td><td>{html.escape(g['format_plan'])}</td><td><a href='{download_url(g['output_file'])}'>Telecharger</a></td></tr>" for g in gens) or "<tr><td colspan='5'>Aucune generation.</td></tr>"
        template_rows = "".join(
            f"<tr><td>{html.escape(t['name'])}</td><td>{html.escape(t['source_type'])}</td><td>{html.escape(t['ai_status'] if 'ai_status' in t.keys() else 'preparation_locale')}</td><td>{html.escape(t['paid_status'] if 'paid_status' in t.keys() else 'test_gratuit')}</td><td><a href='/template-file/{t['id']}' target='_blank'>Fichier</a> | <a href='/template-analysis/{t['id']}'>Analyse</a></td></tr>"
            for t in templates
        ) or "<tr><td colspan='5'>Aucun modele importe.</td></tr>"
        batch_rows = "".join(
            f"<tr><td>{b['created_at']}</td><td>{html.escape(b['project'])}</td><td>{b['total_files']}</td><td>{html.escape(b['status'])}</td><td><a href='{download_url(b['output_zip'])}'>ZIP</a></td></tr>"
            for b in batches
        ) or "<tr><td colspan='5'>Aucun traitement batch.</td></tr>"
        body = f"""
        {cartouche_focus_panel()}
        <h2>Tableau de bord cartouches</h2>
        <div class="grid3">
          <div class="stat"><b>{user['credits']}</b>Credits</div>
          <div class="stat"><b>{html.escape(user['subscription'])}</b>Abonnement</div>
          <div class="stat"><b>{len(gens)}</b>Dernieres generations</div>
        </div>
        <div style="margin-top:16px">{cartouche_steps_html()}</div>
        <div class="grid3" style="margin-top:16px">
          <a class="btn green" href="/generator">Generer une cartouche</a>
          <a class="btn purple" href="/templates">Importer un modele</a>
          <a class="btn dark" href="/batch">Traiter plusieurs PDF</a>
        </div>
        <div class="grid" style="margin-top:16px">
          <div class="card">
            <h3>Paiement TEST PVit</h3>
            <p class="muted">Prepare les essais Airtel Money / Moov Money en mode TEST.</p>
            <a class="btn green" href="/payment">Ouvrir le paiement TEST</a>
          </div>
          <div class="card">
            <h3>Etat paiement</h3>
            <p>Mode actuel : <b>{html.escape(PVIT_MODE)}</b></p>
            <p>Marchand : <b>{html.escape(PVIT_MERCHANT_SLUG)}</b></p>
            <p class="muted">La production sera activee apres validation administrative PVit.</p>
          </div>
        </div>
        <div class="card" style="margin-top:16px">
          <h3>Bibliotheque intelligente de modeles</h3>
          <p class="muted">Chaque cartouche importee est liee au compte utilisateur, stockee, analysee une seule fois et reutilisable plus tard dans le generateur sans refaire l'analyse.</p>
          <table><tr><th>Modele</th><th>Type</th><th>Analyse</th><th>Acces</th><th>Actions</th></tr>{template_rows}</table>
          <p><a class="btn purple" href="/templates">Importer ou gerer mes modeles</a></p>
        </div>
        <div class="card" style="margin-top:16px"><h3>Bibliotheque des cartouches realisees</h3><p class="muted">Tous les PDF/cartouches generes restent dans le compte utilisateur. Le client peut les retrouver, les telecharger ou les renvoyer plus tard, meme s'il les a deja envoyes par mail ou sauvegardes sur son ordinateur.</p><table><tr><th>Date</th><th>N plan</th><th>Projet</th><th>Format</th><th>PDF</th></tr>{rows}</table></div>
        <div class="card" style="margin-top:16px"><h3>Traitements Batch PDF / ZIP</h3><p class="muted">Les lots de plusieurs PDF sont aussi conserves dans l'historique du compte avec le fichier ZIP final.</p><table><tr><th>Date</th><th>Projet</th><th>PDF</th><th>Statut</th><th>Fichier</th></tr>{batch_rows}</table></div>"""
        self.send_html("Tableau de bord", body)

    def workspace(self, message: str = ""):
        user = self.require_login()
        if not user:
            return
        parsed = urllib.parse.urlparse(self.path)
        params = urllib.parse.parse_qs(parsed.query)
        q = (params.get("q") or [""])[0].strip()
        with db() as con:
            space = get_or_create_company_space(con, user)
            projects = con.execute("SELECT * FROM workspace_projects WHERE user_id=? ORDER BY id DESC", (user["id"],)).fetchall()
            if q:
                like = f"%{q}%"
                docs = con.execute(
                    """
                    SELECT d.*, p.name project_name
                    FROM workspace_documents d
                    LEFT JOIN workspace_projects p ON p.id=d.project_id
                    WHERE d.user_id=? AND (d.title LIKE ? OR d.category LIKE ? OR d.original_name LIKE ? OR p.name LIKE ?)
                    ORDER BY d.id DESC
                    LIMIT 50
                    """,
                    (user["id"], like, like, like, like),
                ).fetchall()
            else:
                docs = con.execute(
                    """
                    SELECT d.*, p.name project_name
                    FROM workspace_documents d
                    LEFT JOIN workspace_projects p ON p.id=d.project_id
                    WHERE d.user_id=?
                    ORDER BY d.id DESC
                    LIMIT 25
                    """,
                    (user["id"],),
                ).fetchall()
            templates = con.execute("SELECT * FROM cartouche_templates WHERE user_id=? ORDER BY id DESC LIMIT 10", (user["id"],)).fetchall()
        project_options = "".join(f"<option value='{p['id']}'>{html.escape(p['name'])}</option>" for p in projects)
        project_rows = "".join(
            f"<tr><td>{html.escape(p['name'])}</td><td>{html.escape(p['client'] or '')}</td><td>{html.escape(p['location'] or '')}</td><td>{html.escape(p['status'])}</td><td>{html.escape(p['created_at'])}</td><td><form method='post' action='/workspace' onsubmit=\"return confirm('Supprimer ce projet ? Les documents seront remis en Non classe.');\"><input type='hidden' name='action' value='delete_project'><input type='hidden' name='project_id' value='{p['id']}'><button class='red'>Supprimer</button></form></td></tr>"
            for p in projects
        ) or "<tr><td colspan='6'>Aucun projet pour le moment.</td></tr>"
        doc_rows = "".join(
            f"<tr><td>{html.escape(d['title'])}</td><td>{html.escape(d['category'])}</td><td>{html.escape(d['project_name'] or 'Non classe')}</td><td>{html.escape(d['version_label'])}</td><td>{round(float(d['size_kb'] or 0), 1)} Ko</td><td>{html.escape(d['created_at'])}</td><td><a href='{download_url(d['stored_file'])}'>Telecharger</a></td><td><form method='post' action='/workspace' onsubmit=\"return confirm('Supprimer ce document ?');\"><input type='hidden' name='action' value='delete_document'><input type='hidden' name='document_id' value='{d['id']}'><button class='red'>Supprimer</button></form></td></tr>"
            for d in docs
        ) or "<tr><td colspan='8'>Aucun document stocke.</td></tr>"
        template_rows = "".join(
            f"<tr><td>{html.escape(t['name'])}</td><td>{html.escape(t['ai_status'])}</td><td>{html.escape(t['paid_status'] if 'paid_status' in t.keys() else 'test_gratuit')}</td><td><a href='/template-analysis/{t['id']}'>Analyse</a></td></tr>"
            for t in templates
        ) or "<tr><td colspan='4'>Aucun modele personnalise.</td></tr>"
        body = f"""
        <h2>Espace professionnel prive</h2>
        {message}
        <div class="grid3">
          <div class="stat"><b>{html.escape(space['name'])}</b>Espace entreprise</div>
          <div class="stat"><b>{len(projects)}</b>Projets</div>
          <div class="stat"><b>{len(docs)}</b>Documents recents</div>
        </div>
        <div class="card" style="margin-top:16px">
          <h3>Securite et confidentialite</h3>
          <p class="muted">Premiere base : chaque fichier est lie au compte connecte et reste invisible pour les autres utilisateurs. En production avancee, on passera au stockage prive permanent avec liens temporaires, roles par entreprise et sauvegardes.</p>
          <div class="grid3">
            <div class="stat"><b>2FA</b>Reserve admin</div>
            <div class="stat"><b>Roles</b>Admin, chef projet, lecteur</div>
            <div class="stat"><b>Prive</b>Documents separes par compte</div>
          </div>
        </div>
        <div class="grid" style="margin-top:16px">
          <form class="card" method="post" action="/workspace">
            <h3>Creer un projet / chantier</h3>
            <input type="hidden" name="action" value="create_project">
            <label>Nom du projet</label><input name="name" value="Jardin Botanique">
            <label>Client / Maitre d'ouvrage</label><input name="client" placeholder="Entreprise, ministere, client">
            <label>Localisation</label><input name="location" placeholder="Libreville, Owendo, Akanda...">
            <button class="green">Creer le projet</button>
          </form>
          <form class="card" method="post" action="/workspace" enctype="multipart/form-data">
            <h3>Ajouter un document prive</h3>
            <input type="hidden" name="action" value="upload_document">
            <label>Projet</label><select name="project_id"><option value="">Non classe</option>{project_options}</select>
            <label>Titre du document</label><input name="title" placeholder="Plan topographique, DAO, Rapport...">
            <label>Categorie</label><select name="category"><option>Automatique</option><option>Plan PDF</option><option>Cartouche</option><option>Rapport chantier</option><option>DAO / Administratif</option><option>DWG / AutoCAD</option><option>Image</option><option>Autre</option></select>
            <label>Version</label><input name="version_label" value="V1">
            <label>Fichier</label><input type="file" name="document_file" accept=".pdf,.dwg,.dxf,.png,.jpg,.jpeg,.doc,.docx,.xls,.xlsx,.txt,.zip" required>
            <button class="purple">Stocker dans mon espace</button>
          </form>
        </div>
        <div class="card" style="margin-top:16px"><h3>Projets / chantiers</h3><table><tr><th>Projet</th><th>Client</th><th>Localisation</th><th>Statut</th><th>Date</th><th>Action</th></tr>{project_rows}</table></div>
        <div class="card" style="margin-top:16px">
          <h3>Recherche intelligente simple</h3>
          <form method="get" action="/workspace" class="row"><input name="q" value="{html.escape(q)}" placeholder="Rechercher par projet, fichier, type, chantier..." style="max-width:520px"><button>Rechercher</button><a class="btn dark" href="/workspace">Tout afficher</a></form>
        </div>
        <div class="card" style="margin-top:16px"><h3>Documents stockes</h3><table><tr><th>Titre</th><th>Categorie</th><th>Projet</th><th>Version</th><th>Taille</th><th>Date</th><th>Fichier</th><th>Action</th></tr>{doc_rows}</table></div>
        <div class="card" style="margin-top:16px"><h3>Modeles de cartouches lies au compte</h3><table><tr><th>Modele</th><th>Analyse</th><th>Acces</th><th>Action</th></tr>{template_rows}</table></div>
        """
        self.send_html("Espace entreprise", body)

    def workspace_action(self):
        user = self.require_login()
        if not user:
            return
        content_type = self.headers.get("Content-Type", "")
        if content_type.startswith("multipart/form-data"):
            form = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": content_type})
            item = form["document_file"] if "document_file" in form else None
            if item is None or not getattr(item, "filename", ""):
                return self.workspace("<p class='alert'>Choisis un fichier a stocker.</p>")
            title = (form.getfirst("title") or Path(item.filename).stem).strip() or "Document"
            category_choice = (form.getfirst("category") or "Automatique").strip()
            version_label = (form.getfirst("version_label") or "V1").strip()
            try:
                project_id = int(form.getfirst("project_id") or 0) or None
            except Exception:
                project_id = None
            with db() as con:
                if project_id:
                    exists = con.execute("SELECT id FROM workspace_projects WHERE id=? AND user_id=?", (project_id, user["id"])).fetchone()
                    if not exists:
                        project_id = None
                base_dir = WORKSPACE_FILES / f"user_{user['id']}" / f"project_{project_id or 'non_classe'}"
                base_dir.mkdir(parents=True, exist_ok=True)
                suffix = Path(item.filename).suffix.lower()
                stored_name = f"{int(time.time())}_{secrets.token_hex(4)}_{safe_file(Path(item.filename).stem)}{suffix}"
                stored_path = base_dir / stored_name
                with stored_path.open("wb") as f:
                    shutil.copyfileobj(item.file, f)
                size_kb = round(stored_path.stat().st_size / 1024, 1) if stored_path.exists() else 0
                category = classify_document(item.filename, "Autre") if category_choice == "Automatique" else category_choice
                con.execute(
                    "INSERT INTO workspace_documents(user_id,project_id,title,category,original_name,stored_file,mime_type,size_kb,version_label,confidentiality,status,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
                    (user["id"], project_id, title, category, item.filename, str(stored_path), content_type, size_kb, version_label, "prive", "stocke", now()),
                )
                imported_count = 1
                if suffix == ".zip":
                    extract_dir = base_dir / f"extrait_{stored_path.stem}"
                    extract_dir.mkdir(parents=True, exist_ok=True)
                    try:
                        with zipfile.ZipFile(stored_path, "r") as zf:
                            for member in zf.namelist():
                                if member.endswith("/") or Path(member).name.startswith("."):
                                    continue
                                member_name = Path(member).name
                                member_suffix = Path(member_name).suffix.lower()
                                if member_suffix not in (".pdf", ".dwg", ".dxf", ".png", ".jpg", ".jpeg", ".doc", ".docx", ".xls", ".xlsx", ".txt"):
                                    continue
                                out_name = f"{int(time.time())}_{secrets.token_hex(3)}_{safe_file(Path(member_name).stem)}{member_suffix}"
                                out_path = extract_dir / out_name
                                with out_path.open("wb") as f:
                                    f.write(zf.read(member))
                                imported_count += 1
                                con.execute(
                                    "INSERT INTO workspace_documents(user_id,project_id,title,category,original_name,stored_file,mime_type,size_kb,version_label,confidentiality,status,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
                                    (user["id"], project_id, Path(member_name).stem, classify_document(member_name, "Autre"), member_name, str(out_path), "archive/extracted", round(out_path.stat().st_size / 1024, 1), version_label, "prive", "stocke", now()),
                                )
                    except Exception:
                        pass
            return self.workspace(f"<p class='alert'>{imported_count} document(s) stocke(s) dans votre espace prive.</p>")
        form = self.read_form()
        if form.get("action") == "create_project":
            name = (form.get("name") or "").strip() or "Nouveau projet"
            client = (form.get("client") or "").strip()
            location = (form.get("location") or "").strip()
            with db() as con:
                space = get_or_create_company_space(con, user)
                con.execute(
                    "INSERT INTO workspace_projects(user_id,company_id,name,client,location,status,created_at) VALUES(?,?,?,?,?,?,?)",
                    (user["id"], space["id"], name, client, location, "actif", now()),
                )
            return self.workspace("<p class='alert'>Projet cree dans l'espace entreprise.</p>")
        if form.get("action") == "delete_document":
            try:
                document_id = int(form.get("document_id") or 0)
            except Exception:
                document_id = 0
            with db() as con:
                doc = con.execute("SELECT * FROM workspace_documents WHERE id=? AND user_id=?", (document_id, user["id"])).fetchone()
                if not doc:
                    return self.workspace("<p class='alert'>Document introuvable ou non autorise.</p>")
                path = Path(doc["stored_file"])
                con.execute("DELETE FROM workspace_documents WHERE id=? AND user_id=?", (document_id, user["id"]))
            try:
                if path.exists() and WORKSPACE_FILES in path.resolve().parents:
                    path.unlink()
            except Exception:
                pass
            return self.workspace("<p class='alert'>Document supprime.</p>")
        if form.get("action") == "delete_project":
            try:
                project_id = int(form.get("project_id") or 0)
            except Exception:
                project_id = 0
            with db() as con:
                project = con.execute("SELECT * FROM workspace_projects WHERE id=? AND user_id=?", (project_id, user["id"])).fetchone()
                if not project:
                    return self.workspace("<p class='alert'>Projet introuvable ou non autorise.</p>")
                con.execute("UPDATE workspace_documents SET project_id=NULL WHERE project_id=? AND user_id=?", (project_id, user["id"]))
                con.execute("DELETE FROM workspace_projects WHERE id=? AND user_id=?", (project_id, user["id"]))
            return self.workspace("<p class='alert'>Projet supprime. Les documents associes sont conserves en Non classe.</p>")
        return self.workspace("<p class='alert'>Action non reconnue.</p>")

    def templates_page(self, message: str = ""):
        user = self.require_login()
        if not user:
            return
        with db() as con:
            templates = con.execute(
                "SELECT * FROM cartouche_templates WHERE user_id=? OR ?='admin' ORDER BY id DESC",
                (user["id"], user["role"]),
            ).fetchall()
        rows = "".join(
            f"<tr><td>{html.escape(t['name'])}</td><td>{html.escape(t['source_type'])}</td><td>{html.escape(t['created_at'])}</td><td>{html.escape(t['ai_status'] if 'ai_status' in t.keys() else 'preparation_locale')}</td><td>{html.escape((t['analysis_summary'] if 'analysis_summary' in t.keys() else '') or 'Analyse locale en attente')}</td><td><a href='/template-file/{t['id']}' target='_blank'>Fichier</a> | <a href='/template-analysis/{t['id']}'>Analyse</a></td></tr>"
            for t in templates
        ) or "<tr><td colspan='6'>Aucun modele personnalise pour le moment.</td></tr>"
        platform_cards = "".join(
            f"<div class='stat'><b style='color:{tpl['color']}'>{html.escape(tpl['category'])}</b><strong>{html.escape(tpl['name'])}</strong><p class='muted'>{html.escape(tpl['description'])}</p></div>"
            for tpl in PLATFORM_TEMPLATES.values()
        )
        body = f"""
        {cartouche_focus_panel()}
        <div class="card">
          <h2>Bibliotheque de modeles BTP Smart Tools</h2>
          <p class="muted">Ces modeles servent uniquement aux cartouches. Le client choisit un modele pret a l'emploi ou importe une cartouche d'entreprise qui sera analysee une seule fois, stockee dans son compte, puis reutilisee plusieurs fois.</p>
          <div class="grid3">{platform_cards}</div>
        </div>
        <div class="grid">
          <form class="card" method="post" action="/templates" enctype="multipart/form-data">
            <h2>Modeles de cartouche personnalises</h2>
            {message}
            <p class="muted">Importe une cartouche personnalisee : image, capture ou PDF. L'analyse coute {CREDIT_COST_TEMPLATE_ANALYSIS} credits une seule fois. Ensuite, le modele reste dans le compte et peut etre reutilise pour plusieurs PDF sans refaire l'analyse.</p>
            <label>Nom du modele</label>
            <input name="name" value="Modele entreprise">
            <label>Choisir une cartouche personnalisee</label>
            <input type="file" name="template_file" accept=".png,.jpg,.jpeg,.pdf" required>
            <label>Couleur principale detectee ou souhaitee</label>
            <input type="color" name="theme_color" value="#08213A">
            <button class="purple">Importer et analyser - {CREDIT_COST_TEMPLATE_ANALYSIS} credits</button>
          </form>
          <div class="card">
            <h2>Principe simple</h2>
            <p>1. Le client importe sa cartouche une seule fois.</p>
            <p>2. Le systeme analyse le style, la couleur, le format, l'orientation et les champs.</p>
            <p>3. Le modele est sauvegarde dans sa bibliotheque personnelle.</p>
            <p>4. Il le reutilise dans le generateur sans repayer l'analyse a chaque PDF.</p>
            <p>5. Chaque PDF final consomme seulement le credit de generation.</p>
            <p>6. Les PDF finaux, lots ZIP et historiques restent rattaches au compte utilisateur.</p>
            <p class="alert">Important : l'objectif final n'est pas de coller l'image en fond, mais de reconstruire une cartouche dynamique propre, modifiable et reutilisable.</p>
            {cartouche_quality_html()}
          </div>
        </div>
        <div class="card" style="margin-top:16px">
          <h2>Mes modeles importes</h2>
          <table><tr><th>Nom</th><th>Type</th><th>Date</th><th>Analyse</th><th>Resume</th><th>Actions</th></tr>{rows}</table>
        </div>
        <div class="card" style="margin-top:16px">
          <h2>Preparation API IA</h2>
          <p>L'analyse de cartouche est incluse dans le systeme de credits pour couvrir les couts API et garder une marge.</p>
          <p class="muted">API IA : {html.escape('active' if OPENAI_API_KEY else 'non configuree')} - variable : OPENAI_API_KEY.</p>
          <p class="muted">Les images sont analysees visuellement. Les PDF sont prepares avec metadonnees et texte extrait, puis pourront recevoir une conversion image avancee ensuite.</p>
          <p class="muted">Pour les offres payantes, le modele final pourra etre marque comme paye/valide, telechargeable et reutilisable selon l'abonnement.</p>
        </div>"""
        self.send_html("Modeles", body)

    def templates_action(self):
        user = self.require_login()
        if not user:
            return
        try:
            form = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": self.headers.get("Content-Type")})
            name = (form.getfirst("name") or "Modele entreprise").strip()
            theme_color = (form.getfirst("theme_color") or "#08213A").strip()
            item = form["template_file"] if "template_file" in form else None
            if item is None or not getattr(item, "filename", ""):
                return self.templates_page("<p class='alert'>Choisis d'abord une image ou un PDF de modele.</p>")
            if user["role"] != "admin" and user["credits"] < CREDIT_COST_TEMPLATE_ANALYSIS:
                return self.templates_page(f"<p class='alert'>Credit insuffisant : l'analyse d'une cartouche personnalisee demande {CREDIT_COST_TEMPLATE_ANALYSIS} credits. Ajoute un pack ou un abonnement.</p>")
            suffix = Path(item.filename).suffix.lower()
            if suffix not in (".png", ".jpg", ".jpeg", ".pdf"):
                return self.templates_page("<p class='alert'>Format non accepte. Utilise PNG, JPG, JPEG ou PDF.</p>")
            saved_name = f"template_{int(time.time())}_{secrets.token_hex(4)}_{safe_file(Path(item.filename).stem)}{suffix}"
            saved_path = UPLOADS / saved_name
            saved_path.parent.mkdir(parents=True, exist_ok=True)
            with saved_path.open("wb") as f:
                shutil.copyfileobj(item.file, f)
            if saved_path.stat().st_size == 0:
                saved_path.unlink(missing_ok=True)
                return self.templates_page("<p class='alert'>Le fichier importe est vide. Choisis une image ou un PDF valide.</p>")
            if suffix in (".png", ".jpg", ".jpeg"):
                try:
                    with Image.open(saved_path) as img:
                        img.verify()
                except Exception:
                    saved_path.unlink(missing_ok=True)
                    return self.templates_page("<p class='alert'>L'image n'est pas lisible. Essaie avec PNG ou JPG propre.</p>")
                if theme_color.upper() == "#08213A":
                    theme_color = detect_image_style_color(saved_path, theme_color)
                source_type = "image_reference"
                message = "<p class='alert'>Modele image enregistre comme reference intelligente. Le generateur reconstruit une cartouche propre avec champs dynamiques, sans utiliser l'image comme fond.</p>"
            else:
                try:
                    PdfReader(str(saved_path))
                except Exception:
                    saved_path.unlink(missing_ok=True)
                    return self.templates_page("<p class='alert'>Le PDF modele n'est pas lisible. Essaie avec un PDF propre ou une capture PNG/JPG.</p>")
                source_type = "pdf_reference"
                message = "<p class='alert'>PDF modele enregistre comme reference. La reconstruction dynamique avance par etapes : le systeme garde une cartouche propre au lieu de coller le PDF comme fond.</p>"
            analysis = build_template_analysis(saved_path, source_type, theme_color)
            analysis_summary = summarize_template_analysis(analysis)
            ai_status = "analyse_ia_terminee" if str(analysis.get("engine", "")).startswith("openai") else "preparation_locale"
            if analysis.get("ai_error"):
                ai_status = "analyse_ia_erreur_mode_local"
            paid_status = "admin_gratuit" if user["role"] == "admin" else "test_gratuit"
            with db() as con:
                con.execute(
                    "INSERT INTO cartouche_templates(user_id,name,source_type,source_file,theme_color,ai_status,analysis_json,analysis_summary,paid_status,final_template_file,last_used_at,status,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    (user["id"], name, source_type, str(saved_path), theme_color, ai_status, json.dumps(analysis, ensure_ascii=False), analysis_summary, paid_status, "", "", "active", now()),
                )
                if user["role"] != "admin":
                    con.execute("UPDATE users SET credits=credits-? WHERE id=?", (CREDIT_COST_TEMPLATE_ANALYSIS, user["id"]))
            self.templates_page(message + f"<p class='alert'>{html.escape(analysis_summary)}</p>")
        except Exception as exc:
            self.templates_page(f"<p class='alert'>Import impossible : {html.escape(str(exc))}</p>")

    def profile_analyzer(self, message: str = "", result_html: str = ""):
        user = self.require_login()
        if not user:
            return
        type_options = "".join(f"<option>{html.escape(value)}</option>" for value in PROFILE_ANALYSIS_TYPES)
        with db() as con:
            rows_data = con.execute(
                "SELECT * FROM profile_analyses WHERE user_id=? ORDER BY id DESC LIMIT 12",
                (user["id"],),
            ).fetchall()
        rows = "".join(
            f"<tr><td>{html.escape(row['project'] or 'Non renseigne')}</td><td>{html.escape(row['document_type'])}</td><td>{html.escape(row['source_name'])}</td><td>{html.escape(row['status'])}</td><td>{html.escape(row['created_at'])}</td><td><a href='/profile-analysis/{row['id']}'>Voir</a></td></tr>"
            for row in rows_data
        ) or "<tr><td colspan='6'>Aucune analyse de profil pour le moment.</td></tr>"
        body = f"""
        <div class="hero" style="grid-template-columns:1fr; padding:34px">
          <div>
            <span class="pill">Nouveau module actif</span>
            <h1 style="font-size:54px">Analyse profils routiers</h1>
            <h2>Charge un profil ou un plan, puis le systeme explique le projet comme un technicien BTP.</h2>
            <p>Le module peut recevoir PDF, images, CSV/TXT, Excel, DWG/DXF. Il conserve le fichier, prepare l'analyse et affiche les axes, couches, altitudes, pentes, terrassements, risques et recommandations.</p>
          </div>
        </div>
        <div class="grid" style="margin-top:16px">
          <form class="card" method="post" action="/profile-analyzer" enctype="multipart/form-data">
            <h2>Analyser un profil / plan</h2>
            {message}
            <label>Projet / chantier</label>
            <input name="project" value="Jardin Botanique" placeholder="Ex : Jardin Botanique, Canal Gustave">
            <label>Type de document</label>
            <select name="document_type">{type_options}</select>
            <label>Fichier a analyser</label>
            <input type="file" name="profile_file" accept=".pdf,.png,.jpg,.jpeg,.txt,.csv,.xlsx,.xls,.dwg,.dxf,.zip" required>
            <button class="green">Analyser le profil</button>
            <p class="muted">Le systeme detecte automatiquement le format. PDF/images s'affichent directement, Excel/TXT/CSV sont lus, ZIP est liste, et DWG/DXF sont prepares pour conversion CAO future.</p>
          </form>
          <div class="card">
            <h2>Ce que l'analyse doit expliquer</h2>
            <div class="grid3">
              <div class="stat"><b>Axe</b>PK, emprise, alignement</div>
              <div class="stat"><b>Route</b>Chaussee, trottoir, bordure</div>
              <div class="stat"><b>Couches</b>PST, forme, fondation, base</div>
              <div class="stat"><b>Niveaux</b>Altitudes, pentes, raccordements</div>
              <div class="stat"><b>Terrassement</b>Remblai, deblai, purge</div>
              <div class="stat"><b>Conseils</b>Risques et points a verifier</div>
            </div>
          </div>
        </div>
        {result_html}
        <div class="card" style="margin-top:16px">
          <h2>Historique des analyses</h2>
          <table><tr><th>Projet</th><th>Type</th><th>Fichier</th><th>Statut</th><th>Date</th><th>Action</th></tr>{rows}</table>
        </div>
        """
        self.send_html("Analyse profils routiers", body)

    def profile_analyzer_action(self):
        user = self.require_login()
        if not user:
            return
        try:
            form = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": self.headers.get("Content-Type")})
            project = (form.getfirst("project") or "Projet BTP").strip()
            document_type = (form.getfirst("document_type") or "Profil routier").strip()
            item = form["profile_file"] if "profile_file" in form else None
            if item is None or not getattr(item, "filename", ""):
                return self.profile_analyzer("<p class='alert'>Choisis d'abord un fichier a analyser.</p>")
            suffix = Path(item.filename).suffix.lower()
            accepted = (".pdf", ".png", ".jpg", ".jpeg", ".txt", ".csv", ".xlsx", ".xls", ".dwg", ".dxf", ".zip")
            if suffix not in accepted:
                return self.profile_analyzer("<p class='alert'>Format non accepte pour le moment. Utilise PDF, image, TXT/CSV, Excel ou DWG/DXF.</p>")
            user_dir = PROFILE_UPLOADS / f"user_{user['id']}"
            user_dir.mkdir(parents=True, exist_ok=True)
            saved_name = f"profile_{int(time.time())}_{secrets.token_hex(4)}_{safe_file(Path(item.filename).stem)}{suffix}"
            saved_path = user_dir / saved_name
            with saved_path.open("wb") as f:
                shutil.copyfileobj(item.file, f)
            if saved_path.stat().st_size == 0:
                saved_path.unlink(missing_ok=True)
                return self.profile_analyzer("<p class='alert'>Le fichier est vide. Choisis un document valide.</p>")
            analysis = build_profile_analysis(saved_path, document_type, project)
            status = "analyse_ia_terminee" if str(analysis.get("engine", "")).startswith("openai") else "analyse_locale"
            if analysis.get("ai_error"):
                status = "analyse_locale_ia_limitee"
            with db() as con:
                cur = con.execute(
                    "INSERT INTO profile_analyses(user_id,project,document_type,source_file,source_name,analysis_json,analysis_html,status,created_at) VALUES(?,?,?,?,?,?,?,?,?)",
                    (user["id"], project, document_type, str(saved_path), item.filename, json.dumps(analysis, ensure_ascii=False), "", status, now()),
                )
                analysis_id = cur.lastrowid
                result_html = render_profile_workspace(analysis, f"/profile-source/{analysis_id}", item.filename, saved_path)
                con.execute("UPDATE profile_analyses SET analysis_html=? WHERE id=?", (result_html, analysis_id))
            return self.profile_analyzer("<p class='alert'>Analyse terminee et sauvegardee dans ton historique.</p>", result_html)
        except Exception as exc:
            return self.profile_analyzer(f"<p class='alert'>Analyse impossible : {html.escape(str(exc))}</p>")

    def profile_analysis_detail(self, analysis_id: str):
        user = self.require_login()
        if not user:
            return
        try:
            aid = int(analysis_id)
        except Exception:
            return self.send_html("Analyse introuvable", "<div class='card'><h2>Analyse introuvable</h2></div>", 404)
        with db() as con:
            row = con.execute("SELECT * FROM profile_analyses WHERE id=? AND user_id=?", (aid, user["id"])).fetchone()
        if not row:
            return self.send_html("Analyse introuvable", "<div class='card'><h2>Analyse introuvable ou non autorisee.</h2></div>", 404)
        analysis = json.loads(row["analysis_json"] or "{}")
        visual_html = render_profile_workspace(analysis, f"/profile-source/{row['id']}", row["source_name"], Path(row["source_file"]))
        body = f"""
        <div class="card">
          <h2>Analyse sauvegardee</h2>
          <p><b>Projet :</b> {html.escape(row['project'] or '')}</p>
          <p><b>Type :</b> {html.escape(row['document_type'])}</p>
          <p><b>Fichier :</b> {html.escape(row['source_name'])}</p>
          <p><b>Date :</b> {html.escape(row['created_at'])}</p>
          <p><a class="btn" href="/profile-analyzer">Retour a l'analyseur</a></p>
        </div>
        {visual_html}
        """
        self.send_html("Analyse sauvegardee", body)

    def profile_source(self, analysis_id: str):
        user = self.require_login()
        if not user:
            return
        try:
            aid = int(analysis_id)
        except Exception:
            return self.send_html("Fichier introuvable", "<div class='card'><h2>Fichier introuvable.</h2></div>", 404)
        with db() as con:
            row = con.execute("SELECT * FROM profile_analyses WHERE id=? AND user_id=?", (aid, user["id"])).fetchone()
        if not row:
            return self.send_html("Fichier introuvable", "<div class='card'><h2>Fichier introuvable ou non autorise.</h2></div>", 404)
        path = Path(row["source_file"])
        if not path.exists():
            return self.send_html("Fichier introuvable", "<div class='card'><h2>Le fichier source n'est plus disponible.</h2></div>", 404)
        suffix = path.suffix.lower()
        content_types = {
            ".pdf": "application/pdf",
            ".png": "image/png",
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".txt": "text/plain; charset=utf-8",
            ".csv": "text/plain; charset=utf-8",
        }
        data = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_types.get(suffix, "application/octet-stream"))
        self.send_header("Content-Disposition", f"inline; filename*=UTF-8''{urllib.parse.quote(row['source_name'])}")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def template_file(self, template_id: str):
        user = self.require_login()
        if not user:
            return
        try:
            tid = int(urllib.parse.unquote(template_id))
        except Exception:
            return self.send_html("Modele introuvable", "<div class='card'><h2>Modele introuvable.</h2></div>", 404)
        with db() as con:
            template = con.execute(
                "SELECT * FROM cartouche_templates WHERE id=? AND (user_id=? OR ?='admin')",
                (tid, user["id"], user["role"]),
            ).fetchone()
        if not template:
            return self.send_html("Modele introuvable", "<div class='card'><h2>Modele introuvable.</h2></div>", 404)
        path = Path(template["source_file"])
        if not path.exists():
            return self.send_html("Modele introuvable", "<div class='card'><h2>Fichier modele introuvable.</h2><p>Le modele a peut-etre ete deplace. Reimporte-le depuis la page Modeles.</p></div>", 404)
        data = path.read_bytes()
        content_type = "application/pdf" if path.suffix.lower() == ".pdf" else "image/jpeg"
        if path.suffix.lower() == ".png":
            content_type = "image/png"
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def template_analysis(self, template_id: str):
        user = self.require_login()
        if not user:
            return
        try:
            tid = int(urllib.parse.unquote(template_id))
        except Exception:
            return self.send_html("Analyse introuvable", "<div class='card'><h2>Analyse introuvable.</h2></div>", 404)
        with db() as con:
            template = con.execute(
                "SELECT * FROM cartouche_templates WHERE id=? AND (user_id=? OR ?='admin')",
                (tid, user["id"], user["role"]),
            ).fetchone()
        if not template:
            return self.send_html("Analyse introuvable", "<div class='card'><h2>Modele introuvable.</h2></div>", 404)
        raw = template["analysis_json"] if "analysis_json" in template.keys() else ""
        try:
            analysis = json.loads(raw) if raw else {}
        except Exception:
            analysis = {}
        fields = "".join(f"<li>{html.escape(str(field))}</li>" for field in analysis.get("editable_fields", [])) or "<li>Champs a definir</li>"
        detected = "".join(f"<li>{html.escape(str(item))}</li>" for item in analysis.get("detected_elements", [])) or "<li>Elements a confirmer</li>"
        json_pretty = html.escape(json.dumps(analysis, indent=2, ensure_ascii=False))
        body = f"""
        <div class="grid">
          <div class="card">
            <h2>Analyse du modele personnalise</h2>
            <p><b>Modele :</b> {html.escape(template['name'])}</p>
            <p><b>Statut :</b> {html.escape(template['ai_status'] if 'ai_status' in template.keys() else 'preparation_locale')}</p>
            <p><b>Resume :</b> {html.escape((template['analysis_summary'] if 'analysis_summary' in template.keys() else '') or 'Analyse locale preparee.')}</p>
            <p><b>Couleur principale :</b> <span class="pill">{html.escape(template['theme_color'])}</span></p>
            <p><a class="btn" href="/template-file/{template['id']}" target="_blank">Voir le fichier importe</a></p>
          </div>
          <div class="card">
            <h2>Analyse IA / structure</h2>
            <p>Si la cle OpenAI est configuree dans Render, cette analyse provient du moteur IA. Sinon le systeme affiche une preparation locale de secours.</p>
            <p class="muted">Objectif : detecter cadres, blocs textes, logo, tableaux, revisions, signatures, couleurs et positions exactes pour reconstruire une cartouche dynamique.</p>
          </div>
        </div>
        <div class="grid" style="margin-top:16px">
          <div class="card"><h2>Elements detectes / a confirmer</h2><ul>{detected}</ul></div>
          <div class="card"><h2>Champs dynamiques prevus</h2><ul>{fields}</ul></div>
        </div>
        <div class="card" style="margin-top:16px">
          <h2>Structure technique JSON</h2>
          <pre style="white-space:pre-wrap;background:#020617;color:#dbeafe;padding:16px;border-radius:12px;overflow:auto">{json_pretty}</pre>
        </div>"""
        self.send_html("Analyse modele", body)

    def generator(self, values: dict | None = None):
        user = self.require_login()
        if not user:
            return
        if values is None:
            values = {}
            query = urllib.parse.parse_qs(urllib.parse.urlparse(self.path).query)
            for key in ("project", "company", "plan_type", "scale", "author", "number", "revision", "revision_index", "revision_date", "revision_description"):
                if query.get(key):
                    values[key] = query[key][0]
        with db() as con:
            templates = con.execute(
                "SELECT * FROM cartouche_templates WHERE user_id=? OR ?='admin' ORDER BY id DESC",
                (user["id"], user["role"]),
            ).fetchall()
        platform_options = "".join(
            f"<option value='{key}'>{html.escape(tpl['name'])} - {html.escape(tpl['category'])}</option>"
            for key, tpl in PLATFORM_TEMPLATES.items()
        )
        user_options = "".join(
            f"<option value='{t['id']}'>{html.escape(t['name'])} - {html.escape(t['source_type'])}</option>" for t in templates
        ) or "<option disabled>Aucun modele importe</option>"
        template_options = f"<optgroup label='Modeles BTP Smart Tools'>{platform_options}</optgroup><optgroup label='Mes modeles importes'>{user_options}</optgroup>"
        today_fr = datetime.now().strftime("%d/%m/%Y")
        body = f"""
        {cartouche_focus_panel()}
        {cartouche_steps_html()}
        <form class="grid" method="post" action="/generate" enctype="multipart/form-data">
          {datalist_html()}
          <div class="card">
            <h2>Generateur de cartouche</h2>
            <label>Fichiers PDF/Image pour cartouche finale</label><input type="file" name="file" multiple accept=".pdf,.png,.jpg,.jpeg,.webp,.bmp,.tif,.tiff,.dwg,.dxf">
            <p class="muted">Important : l'aperçu et la cartouche finale utilisent directement les PDF/images. Pour un DWG/DXF, exporte d'abord le plan en PDF depuis AutoCAD afin de conserver le vrai dessin.</p>
            <label>Logo societe</label><input type="file" name="logo" accept=".png,.jpg,.jpeg">
            <div class="grid">
              <div><label>Projet</label><input name="project" value="{html.escape(values.get('project','Jardin Botanique'))}"></div>
              <div><label>Entreprise</label><input name="company" value="{html.escape(values.get('company','SKE System'))}"></div>
            </div>
            <div class="grid">
              <div><label>Telephone societe</label><input name="company_phone" value="{html.escape(values.get('company_phone','+241 74 15 37 16'))}"></div>
              <div><label>Mail societe</label><input name="company_email" value="{html.escape(values.get('company_email','contact@ske-btp.com'))}"></div>
            </div>
            <div class="grid">
              <div><label>Type de plan</label><input list="plan_type_options" name="plan_type" value="{html.escape(values.get('plan_type','Plan beton'))}" placeholder="Choisir ou ecrire"></div>
              <div><label>Echelle</label><input list="scale_options" name="scale" value="{html.escape(values.get('scale','1/100'))}" placeholder="Detectee automatiquement si presente dans le PDF"></div>
            </div>
            <div class="grid">
              <div><label>Auteur</label><input name="author" value="{html.escape(values.get('author','Michael'))}"></div>
              <div><label>Format du plan</label><select name="format_plan"><option selected>Automatique</option><option>A4 Portrait</option><option>A4 Paysage</option><option>A3 Portrait</option><option>A3 Paysage</option><option>A2 Portrait</option><option>A2 Paysage</option><option>A1 Portrait</option><option>A1 Paysage</option><option>A0 Portrait</option><option>A0 Paysage</option></select></div>
            </div>
            <label>Modele de cartouche</label>
            <select name="template_id">{template_options}</select>
            <p class="muted">Choisis un modele BTP Smart Tools pret a l'emploi ou un modele personnalise importe dans la page Modeles.</p>
            <div class="grid">
              <div><label>Chef de projet</label><input name="project_manager" value="{html.escape(values.get('project_manager',''))}"></div>
              <div><label>Operateur</label><input name="operator" value="{html.escape(values.get('operator',''))}"></div>
            </div>
            <div class="grid">
              <div><label>Validateur</label><input name="validator" value="{html.escape(values.get('validator',''))}"></div>
              <div><label>Revision principale</label><input list="revision_index_options" name="revision" value="{html.escape(values.get('revision','REV 00'))}"></div>
            </div>
            <div class="grid">
              <div><label>Theme couleur</label><select name="theme"><option>Bleu BTP</option><option>Noir Pro</option><option>Rouge Entreprise</option><option>Vert Chantier</option><option>Gris Technique</option></select></div>
              <div><label>Couleur personnalisee</label><input name="theme_color" type="color" value="#08213A"></div>
            </div>
            <p class="alert">Numero du plan automatique : le systeme genere BTP-001, BTP-002, BTP-003... a chaque nouveau PDF. Si le PDF contient une echelle lisible, elle est detectee automatiquement.</p>
            {cartouche_quality_html()}
            <h3>Zones a afficher</h3>
            <div class="grid">
              <label><input type="checkbox" name="show_legend" checked style="width:auto"> Afficher la legende</label>
              <label><input type="checkbox" name="show_elements" checked style="width:auto"> Afficher le tableau des elements</label>
            </div>
            <label><input type="checkbox" name="show_revisions" checked style="width:auto"> Afficher les revisions</label>
            <h3>Tableau des revisions dynamique</h3>
            <div class="grid">
              <div><label>Indice</label><input list="revision_index_options" id="revIndex" name="revision_index" value="{html.escape(values.get('revision_index','REV 00'))}" placeholder="Choisir ou ecrire"></div>
              <div><label>Date</label><input id="revDate" name="revision_date" value="{html.escape(values.get('revision_date', today_fr))}"></div>
            </div>
            <label>Description revision</label>
            <input list="revision_description_options" id="revDescription" name="revision_description" value="{html.escape(values.get('revision_description','Premiere emission'))}" placeholder="Choisir ou ecrire">
            <button type="button" class="dark" onclick="addRevisionRow()">+ Ajouter la revision</button>
            <div id="revisionRows"></div>
            <h3>Legendes dynamiques</h3>
            <div id="legends"></div>
            <button type="button" class="dark" onclick="addLegend()">+ Ajouter une legende</button>
            <h3>Tableau des elements dynamique</h3>
            <div class="grid">
              <div><label>Colonne 1</label><input id="col1" value="DESIGNATION"></div>
              <div><label>Colonne 2</label><input id="col2" value="UNITE"></div>
              <div><label>Colonne 3</label><input id="col3" value="QUANTITE"></div>
              <div><label>Colonne 4</label><input id="col4" value="OBSERVATION"></div>
            </div>
            <div id="elementsRows"></div>
            <button type="button" class="dark" onclick="addElementRow()">+ Ajouter une ligne</button>
            <input type="hidden" name="legends_json" id="legends_json">
            <input type="hidden" name="element_columns_json" id="element_columns_json">
            <input type="hidden" name="element_rows_json" id="element_rows_json">
            <input type="hidden" name="revision_rows_json" id="revision_rows_json">
            <hr><button class="green" onclick="saveLegends()">Generer PDF final</button>
          </div>
          {preview_panel("Apercu avant generation")}
        </form>
        <script>
        function moveLine(btn,dir){{let r=btn.parentElement,p=r.parentElement;if(dir<0&&r.previousElementSibling)p.insertBefore(r,r.previousElementSibling);if(dir>0&&r.nextElementSibling)p.insertBefore(r.nextElementSibling,r);}}
        function addLegend(t='',c='#2487ff',s='Ligne'){{let d=document.createElement('div');d.className='legend-line';d.innerHTML=`<select><option>Ligne</option><option>Rectangle</option><option>Cercle</option><option>Croix</option><option>Hachure</option><option>Texte</option></select><input type="color" value="${{c}}"><input placeholder="Intitule de la legende" value="${{t}}"><button type="button" class="dark" onclick="moveLine(this,-1)">↑</button><button type="button" class="dark" onclick="moveLine(this,1)">↓</button><button type="button" class="red" onclick="this.parentElement.remove()">X</button>`;d.children[0].value=s;d.style.gridTemplateColumns='110px 80px 1fr 44px 44px 44px';document.getElementById('legends').appendChild(d);}}
        function addElementRow(a='',b='',c='',d=''){{let r=document.createElement('div');r.className='legend-line';r.innerHTML=`<input placeholder="Designation : Surface de remblai" value="${{a}}"><input placeholder="Unite : m2, m3, ml" value="${{b}}"><input placeholder="Quantite" value="${{c}}"><input placeholder="Observation" value="${{d}}"><button type="button" class="red" onclick="this.parentElement.remove()">X</button>`;r.style.gridTemplateColumns='1.4fr .7fr .8fr 1.2fr 44px';document.getElementById('elementsRows').appendChild(r);}}
        function addRevisionRow(i='',d='',desc=''){{let r=document.createElement('div');r.className='legend-line revision-line';r.innerHTML=`<input placeholder="Indice" list="revision_index_options" value="${{i || revIndex.value || 'REV 00'}}"><input placeholder="Date" value="${{d || revDate.value}}"><input placeholder="Description" list="revision_description_options" value="${{desc || revDescription.value}}"><button type="button" class="red" onclick="this.parentElement.remove()">X</button>`;r.style.gridTemplateColumns='0.8fr 1fr 2fr 44px';document.getElementById('revisionRows').appendChild(r);}}
        function saveLegends(){{let arr=[...document.querySelectorAll('#legends .legend-line')].map(x=>({{symbol:x.children[0].value,color:x.children[1].value,text:x.children[2].value}}));document.getElementById('legends_json').value=JSON.stringify(arr);let cols=[col1.value,col2.value,col3.value,col4.value];document.getElementById('element_columns_json').value=JSON.stringify(cols);let rows=[...document.querySelectorAll('#elementsRows .legend-line')].map(x=>({{c1:x.children[0].value,c2:x.children[1].value,c3:x.children[2].value,c4:x.children[3].value}}));document.getElementById('element_rows_json').value=JSON.stringify(rows);let revs=[...document.querySelectorAll('#revisionRows .revision-line')].map(x=>({{indice:x.children[0].value,date:x.children[1].value,description:x.children[2].value}}));if(!revs.length){{revs=[{{indice:revIndex.value || 'REV 00',date:revDate.value,description:revDescription.value || 'Premiere emission'}}];}}document.getElementById('revision_rows_json').value=JSON.stringify(revs);}}
        addLegend('Canal du Staff','#2487ff','Ligne');addLegend('Purge','#0f766e','Rectangle');addLegend('Point de reference','#dc2626','Croix');
        addElementRow('Surface de remblai','m2','','');addElementRow('Volume de purge','m3','','');addElementRow('Beton','m3','','');
        addRevisionRow('REV 00','{today_fr}','Premiere emission');
        </script>
        {PREVIEW_SCRIPT}"""
        self.send_html("Generateur", body)

    def generate_action(self):
        user = self.require_login()
        if not user:
            return
        form = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": self.headers.get("Content-Type")})
        info = {k: (form.getfirst(k) or "").strip() for k in (
            "project", "company", "company_phone", "company_email", "plan_type", "scale",
            "author", "format_plan", "project_manager", "operator", "validator",
            "revision", "theme", "theme_color", "template_id"
        )}
        legends = json.loads(form.getfirst("legends_json") or "[]")
        info["element_columns"] = json.loads(form.getfirst("element_columns_json") or '["REPERE","DESIGNATION","SURFACE","QTE"]')
        info["element_rows"] = json.loads(form.getfirst("element_rows_json") or "[]")
        info["revision_rows"] = json.loads(form.getfirst("revision_rows_json") or "[]")
        info["show_legend"] = "on" if form.getfirst("show_legend") else "off"
        info["show_elements"] = "on" if form.getfirst("show_elements") else "off"
        info["show_revisions"] = "on" if form.getfirst("show_revisions") else "off"
        info["element_columns"] = json.loads(form.getfirst("element_columns_json") or '["REPERE","DESIGNATION","DIM.","QTE"]')
        info["element_rows"] = json.loads(form.getfirst("element_rows_json") or "[]")
        info["revision_rows"] = json.loads(form.getfirst("revision_rows_json") or "[]")
        info["show_legend"] = "on" if form.getfirst("show_legend") else "off"
        info["show_elements"] = "on" if form.getfirst("show_elements") else "off"
        info["show_revisions"] = "on" if form.getfirst("show_revisions") else "off"
        file_items = form["file"] if "file" in form else []
        if not isinstance(file_items, list):
            file_items = [file_items]
        uploaded_sources: list[tuple[str, Path | None]] = []
        for item in file_items:
            if item is not None and getattr(item, "filename", ""):
                source_name = f"{int(time.time())}_{secrets.token_hex(3)}_{Path(item.filename).name}"
                source_path = UPLOADS / source_name
                with source_path.open("wb") as f:
                    shutil.copyfileobj(item.file, f)
                uploaded_sources.append((source_name, source_path))
        if not uploaded_sources:
            uploaded_sources.append(("", None))
        supported_exts = {".pdf", ".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".webp"}
        unsupported_sources = [source_path.name for _, source_path in uploaded_sources if source_path is not None and source_path.suffix.lower() not in supported_exts]
        if unsupported_sources:
            items = "".join(f"<li>{html.escape(name)}</li>" for name in unsupported_sources)
            return self.send_html(
                "Format non compatible",
                f"""
                <div class='card'>
                  <h2>Le fichier charge ne peut pas etre affiche directement</h2>
                  <p>Le generateur de cartouches applique la cartouche sur le vrai document uniquement pour les formats PDF et images.</p>
                  <p>Les fichiers suivants ne peuvent pas encore etre rendus dans le navigateur ou dans le PDF final :</p>
                  <ul>{items}</ul>
                  <p class='alert'>Pour un DWG/DXF : ouvre le dessin dans AutoCAD/Covadis, exporte-le en PDF, puis recharge ce PDF dans BTP Smart Tools.</p>
                  <a class='btn' href='/generator'>Retour au generateur</a>
                </div>
                """,
                user,
            )
        required_credits = len(uploaded_sources) * CREDIT_COST_PDF
        if user["role"] != "admin" and user["credits"] < required_credits:
            return self.send_html("Credit requis", f"<div class='card'><h2>Credit requis</h2><p>Cette generation demande {required_credits} credit(s). Ajoute un pack ou un abonnement pour continuer.</p><a class='btn green' href='/payment'>Ajouter des credits</a><a class='btn' href='/dashboard'>Retour</a></div>")
        logo_item = form["logo"] if "logo" in form else None
        if logo_item is not None and getattr(logo_item, "filename", ""):
            logo_name = f"logo_{int(time.time())}_{Path(logo_item.filename).name}"
            logo_path = UPLOADS / logo_name
            with logo_path.open("wb") as f:
                shutil.copyfileobj(logo_item.file, f)
            info["logo_path"] = str(logo_path)
        generated: list[Path] = []
        with db() as con:
            for source_name, source_path in uploaded_sources:
                plan_number = next_plan_number(con)
                format_name, page_size = resolve_format(info.get("format_plan", "Automatique"), source_path)
                item_info = dict(info)
                detected = guess_plan_metadata(source_path)
                if detected.get("plan_type") and item_info.get("plan_type") in ("", "Plan technique", "Plan beton"):
                    item_info["plan_type"] = detected["plan_type"]
                if detected.get("scale") and item_info.get("scale") in ("", "1/100"):
                    item_info["scale"] = detected["scale"]
                item_info["number"] = plan_number
                item_info["format_plan"] = format_name
                base_name = safe_file(source_path.stem if source_path else item_info["project"])
                out_name = f"{plan_number}_{base_name}.pdf"
                out_path = OUTPUTS / out_name
                generate_pdf(out_path, item_info, legends, page_size, source_path)
                generated.append(out_path)
                con.execute(
                    "INSERT INTO generations(user_id,project,company,plan_type,scale,format_plan,plan_number,source_file,output_file,status,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
                    (user["id"], item_info["project"], item_info["company"], item_info["plan_type"], item_info["scale"], item_info["format_plan"], item_info["number"], source_name, str(out_path), "generated_multi" if len(uploaded_sources) > 1 else "generated", now()),
                )
            if user["role"] != "admin":
                con.execute("UPDATE users SET credits=credits-? WHERE id=?", (required_credits, user["id"]))
        if len(generated) == 1:
            self.redirect(download_url(generated[0]))
            return
        zip_name = f"BTP_LOT_{int(time.time())}_{safe_file(info['project'])}.zip"
        zip_path = OUTPUTS / zip_name
        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for path in generated:
                zf.write(path, path.name)
        self.redirect(download_url(zip_name))

    def assistant(self, result: str = ""):
        body = f"""
        {cartouche_focus_panel()}
        <div class="grid">
          <form class="card" method="post" action="/assistant">
            <h2>Assistant cartouches PDF</h2>
            <p class="muted">Pose une question sur la generation de cartouches, les credits, les modeles importes, le batch PDF, le paiement ou l'historique du compte.</p>
            <label>Votre question</label>
            <textarea name="question" placeholder="Exemple : combien coute une cartouche ? ou sont mes PDF ? comment reutiliser mon modele ?"></textarea>
            <button class="purple">Obtenir une reponse</button>
            <p class="alert">L'assistant guide le client. Les generations se font dans les pages Cartouche, Modeles et Batch PDF.</p>
          </form>
          <div>
            <div class="card">
              <h2>Aide utilisateur</h2>
              {result or "<p>Bonjour, je suis l'assistant cartouche de BTP Smart Tools. Je peux expliquer comment importer un PDF, creer une cartouche, reutiliser un modele, gerer les credits et retrouver les fichiers generes.</p>"}
              <hr>
              <p class="muted">L'assistant explique le fonctionnement. Les PDF, modeles et ZIP restent rattaches au compte utilisateur apres generation.</p>
              <p class="muted">Support : sessouedem15@gmail.com | +241 74 15 37 16 | +241 65 28 05 25</p>
            </div>
            <div class="card" style="margin-top:16px">
              <h3>Questions rapides</h3>
              <p><b>1 credit :</b> 1 PDF final genere.</p>
              <p><b>Modele importe :</b> analyse une seule fois, puis reutilisation dans le compte.</p>
              <p><b>Historique :</b> les PDF et ZIP restent disponibles dans le tableau de bord.</p>
            </div>
          </div>
        </div>"""
        self.send_html("Assistant cartouches", body)

    def assistant_action(self):
        question = self.read_form().get("question", "")
        result = support_answer(question)
        self.assistant(result)
    def batch_page(self):
        user = self.require_login()
        if not user:
            return
        body = cartouche_focus_panel() + cartouche_steps_html() + """
        <form class="grid" method="post" action="/batch" enctype="multipart/form-data">
          {DATALISTS}
          <div class="card">
            <h2>Batch cartouches PDF</h2>
            <p class="muted">Charge plusieurs PDF ou un ZIP contenant des PDF. Le systeme applique automatiquement la meme cartouche a tous les plans, avec numeros automatiques.</p>
            <label>PDF multiples ou fichier ZIP</label>
            <input type="file" name="files" multiple accept=".pdf,.zip">
            <label>Logo entreprise</label>
            <input type="file" name="logo" accept=".png,.jpg,.jpeg">
            <div class="grid">
              <div><label>Nom du projet</label><input name="project" value="Jardin Botanique"></div>
              <div><label>Entreprise</label><input name="company" value="SKE System"></div>
            </div>
            <div class="grid">
              <div><label>Type de plan</label><input list="plan_type_options" name="plan_type" value="Plan technique" placeholder="Choisir ou ecrire"></div>
              <div><label>Echelle</label><input list="scale_options" name="scale" value="1/100" placeholder="Choisir ou ecrire"></div>
            </div>
            <div class="grid">
              <div><label>Format</label><select name="format_plan"><option selected>Automatique</option><option>A4 Portrait</option><option>A4 Paysage</option><option>A3 Portrait</option><option>A3 Paysage</option><option>A2 Portrait</option><option>A2 Paysage</option><option>A1 Portrait</option><option>A1 Paysage</option><option>A0 Portrait</option><option>A0 Paysage</option></select></div>
              <div><label>Couleur principale</label><input name="theme_color" type="color" value="#08213A"></div>
            </div>
            <label>Modele de cartouche</label>
            <select name="template_id"><option value="platform_standard">BTP Standard</option><option value="platform_premium">BTP Premium Bleu</option><option value="platform_modern">BTP Moderne Gris</option><option value="platform_topo">BTP Topographie</option><option value="platform_engineering">BTP Ingenierie</option></select>
            <div class="grid">
              <div><label>Operateur</label><input name="operator"></div>
              <div><label>Chef de projet</label><input name="project_manager"></div>
            </div>
            <div class="grid">
              <div><label>Dessinateur / auteur</label><input name="author" value="Michael"></div>
              <div><label>Validateur</label><input name="validator"></div>
            </div>
            <div class="grid">
              <div><label>Telephone societe</label><input name="company_phone" value="+241 74 15 37 16"></div>
              <div><label>Mail societe</label><input name="company_email" value="contact@ske-btp.com"></div>
            </div>
            <label>Legendes communes</label>
            <div id="legends"></div>
            <button type="button" class="dark" onclick="addLegend()">+ Ajouter une legende</button>
            <h3>Zones communes</h3>
            <label><input type="checkbox" name="show_legend" checked style="width:auto"> Afficher la legende</label>
            <label><input type="checkbox" name="show_elements" checked style="width:auto"> Afficher le tableau des elements</label>
            <label><input type="checkbox" name="show_revisions" checked style="width:auto"> Afficher les revisions</label>
            <h3>Tableau des elements commun</h3>
            <div class="grid">
              <div><label>Colonne 1</label><input id="col1" value="REPERE"></div>
              <div><label>Colonne 2</label><input id="col2" value="DESIGNATION"></div>
              <div><label>Colonne 3</label><input id="col3" value="SURFACE"></div>
              <div><label>Colonne 4</label><input id="col4" value="QTE"></div>
            </div>
            <div id="elementsRows"></div>
            <button type="button" class="dark" onclick="addElementRow()">+ Ajouter une ligne</button>
            <input type="hidden" name="legends_json" id="legends_json">
            <input type="hidden" name="element_columns_json" id="element_columns_json">
            <input type="hidden" name="element_rows_json" id="element_rows_json">
            <hr>
            <button class="purple" onclick="saveLegends()">Lancer le traitement batch PDF</button>
          </div>
          <div class="card">
            <h2>Ce que le batch fait</h2>
            <p>1. Detecte tous les PDF.</p>
            <p>2. Applique les memes informations a chaque plan.</p>
            <p>3. Detecte le format PDF si Automatique est choisi.</p>
            <p>4. Genere BTP-001, BTP-002, BTP-003...</p>
            <p>5. Cree tous les PDF finaux.</p>
            <p>6. Prepare un ZIP final a telecharger.</p>
            <p class="alert">Admin : traitement gratuit pour les tests.</p>
          </div>
        </form>
        <script>
        function addLegend(t='',c='#2487ff',s='Ligne'){let d=document.createElement('div');d.className='legend-line';d.innerHTML=`<select><option>Ligne</option><option>Rectangle</option><option>Cercle</option><option>Croix</option><option>Hachure</option></select><input type="color" value="${c}"><input placeholder="Description" value="${t}">`;d.children[0].value=s;document.getElementById('legends').appendChild(d);}
        function addElementRow(a='',b='',c='',d=''){let r=document.createElement('div');r.className='legend-line';r.innerHTML=`<input placeholder="Col.1" value="${a}"><input placeholder="Col.2" value="${b}"><input placeholder="Col.3" value="${c}"><input placeholder="Col.4" value="${d}"><button type="button" class="red" onclick="this.parentElement.remove()">X</button>`;r.style.gridTemplateColumns='1fr 1fr 1fr 1fr 44px';document.getElementById('elementsRows').appendChild(r);}
        function saveLegends(){let arr=[...document.querySelectorAll('#legends .legend-line')].map(x=>({symbol:x.children[0].value,color:x.children[1].value,text:x.children[2].value}));document.getElementById('legends_json').value=JSON.stringify(arr);document.getElementById('element_columns_json').value=JSON.stringify([col1.value,col2.value,col3.value,col4.value]);let rows=[...document.querySelectorAll('#elementsRows .legend-line')].map(x=>({c1:x.children[0].value,c2:x.children[1].value,c3:x.children[2].value,c4:x.children[3].value}));document.getElementById('element_rows_json').value=JSON.stringify(rows);}
        addLegend('Plan importe','#2487ff','Ligne'); addLegend('Zone projet','#0f766e','Rectangle');
        addElementRow('A1','Surface beton','25 m2',''); addElementRow('V1','Volume','25 x 0.20','5 m3');
        </script>"""
        body = body.replace("{DATALISTS}", datalist_html())
        body = body.replace("<div class=\"card\">\n            <h2>Ce que le batch fait</h2>", preview_panel("Apercu Batch PDF") + "<div class=\"card\" style=\"margin-top:16px\">\n            <h2>Ce que le batch fait</h2>")
        body += PREVIEW_SCRIPT
        self.send_html("Batch PDF", body)

    def batch_action(self):
        user = self.require_login()
        if not user:
            return
        form = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": self.headers.get("Content-Type")})
        batch_id = f"batch_{int(time.time())}_{secrets.token_hex(4)}"
        batch_dir = BATCHES / batch_id
        pdfs = collect_batch_pdfs(form, batch_dir)
        if not pdfs:
            return self.send_html("Aucun PDF", "<div class='card'><h2>Aucun PDF detecte</h2><p>Charge plusieurs PDF ou un ZIP contenant des PDF.</p><a class='btn' href='/batch'>Retour</a></div>")
        required_credits = len(pdfs) * CREDIT_COST_PDF
        if user["role"] != "admin" and user["credits"] < required_credits:
            return self.send_html("Credit requis", f"<div class='card'><h2>Credit requis</h2><p>Ce traitement batch demande {required_credits} credit(s), soit 1 credit par PDF.</p><a class='btn green' href='/payment'>Ajouter des credits</a><a class='btn' href='/batch'>Retour</a></div>")
        info = {k: (form.getfirst(k) or "").strip() for k in (
            "project", "company", "company_phone", "company_email", "plan_type", "scale",
            "author", "format_plan", "project_manager", "operator", "validator", "theme_color", "template_id"
        )}
        info["revision"] = "REV 00"
        legends = json.loads(form.getfirst("legends_json") or "[]")
        logo_item = form["logo"] if "logo" in form else None
        if logo_item is not None and getattr(logo_item, "filename", ""):
            logo_path = batch_dir / f"logo_{Path(logo_item.filename).name}"
            with logo_path.open("wb") as f:
                shutil.copyfileobj(logo_item.file, f)
            info["logo_path"] = str(logo_path)
        final_dir = batch_dir / "final"
        final_dir.mkdir(parents=True, exist_ok=True)
        generated: list[Path] = []
        with db() as con:
            for source in pdfs:
                plan_number = next_plan_number(con)
                format_name, page_size = resolve_format(info.get("format_plan", "Automatique"), source)
                item_info = dict(info)
                detected = guess_plan_metadata(source)
                if detected.get("plan_type") and item_info.get("plan_type") in ("", "Plan technique", "Plan beton"):
                    item_info["plan_type"] = detected["plan_type"]
                if detected.get("scale") and item_info.get("scale") in ("", "1/100"):
                    item_info["scale"] = detected["scale"]
                item_info["number"] = plan_number
                item_info["format_plan"] = format_name
                out_path = final_dir / f"{plan_number}_{safe_file(source.stem)}.pdf"
                generate_pdf(out_path, item_info, legends, page_size, source)
                generated.append(out_path)
                con.execute(
                    "INSERT INTO generations(user_id,project,company,plan_type,scale,format_plan,plan_number,source_file,output_file,status,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
                    (user["id"], item_info["project"], item_info["company"], item_info["plan_type"], item_info["scale"], item_info["format_plan"], item_info["number"], str(source), str(out_path), "generated_batch", now()),
                )
            zip_path = OUTPUTS / f"{batch_id}_PDF_FINALISES.zip"
            with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                for path in generated:
                    zf.write(path, path.name)
            con.execute(
                "INSERT INTO batches(user_id,project,total_files,output_zip,status,created_at) VALUES(?,?,?,?,?,?)",
                (user["id"], info.get("project", ""), len(generated), str(zip_path), "completed", now()),
            )
            if user["role"] != "admin":
                con.execute("UPDATE users SET credits=credits-? WHERE id=?", (required_credits, user["id"]))
        body = f"""
        <div class="card">
          <h2>Traitement Batch PDF termine</h2>
          <p>{len(generated)} PDF ont ete finalises automatiquement.</p>
          <a class="btn green" href="{download_url(zip_path)}">Telecharger tous les plans finalises</a>
          <a class="btn" href="/dashboard">Voir historique</a>
        </div>"""
        self.send_html("Batch termine", body)

    def payment_action(self):
        user = self.require_login()
        if not user:
            return
        form = self.read_form()
        offer_key = form.get("offer_key", "pdf")
        offer = PAYMENT_OFFERS.get(offer_key, PAYMENT_OFFERS["pdf"])
        method = form.get("method", "Airtel Money")
        phone = form.get("phone", "")
        with db() as con:
            cur = con.execute(
                "INSERT INTO payments(user_id,offer,amount,method,phone,provider,provider_mode,status,created_at,offer_key,credits) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
                (user["id"], offer["label"], offer["display"], method, phone, "mypvit", PVIT_MODE, "created", now(), offer_key, int(offer.get("credits", 0))),
            )
            payment_id = cur.lastrowid
            result = create_pvit_payment(user, offer_key, phone, payment_id)
            con.execute(
                "UPDATE payments SET status=?, transaction_ref=?, payment_url=?, provider_response=? WHERE id=?",
                (result["status"], result["reference"], result["payment_url"], json.dumps(result["raw"], ensure_ascii=False), payment_id),
            )
        if result["payment_url"]:
            self.redirect(result["payment_url"])
        else:
            body = f"""
            <div class="card">
              <h2>Paiement TEST prepare</h2>
              <p><b>Offre :</b> {html.escape(offer['label'])} - {html.escape(offer['display'])}</p>
              <p><b>Credits inclus :</b> {int(offer.get('credits', 0))}</p>
              <p><b>Methode :</b> {html.escape(method)}</p>
              <p><b>Reference :</b> {html.escape(result['reference'])}</p>
              <p class="alert">{html.escape(result['message'])}</p>
              <p class="muted">Le paiement est enregistre dans l'admin. Il restera en attente tant que la cle secrete et le compte d'operation TEST ne sont pas ajoutes.</p>
              <a class="btn" href="/payment">Revenir au paiement</a>
              <a class="btn dark" href="/dashboard">Tableau de bord</a>
            </div>"""
            self.send_html("Paiement TEST", body)

    def payment_page(self):
        user = self.require_login()
        if not user:
            return
        configured, missing = pvit_config_status()
        config_message = "Configuration PVit TEST complete." if configured else "Configuration incomplete : " + ", ".join(missing)
        offer_options = "".join(
            f"<option value='{key}'>{html.escape(offer['label'])} - {html.escape(offer['display'])} - {int(offer.get('credits', 0))} credits</option>"
            for key, offer in PAYMENT_OFFERS.items()
        )
        body = f"""
        <div class="grid">
          <form class="card" method="post" action="/payment">
            <h2>Paiement TEST Airtel / Moov</h2>
            <p class="muted">Integration preparee avec MyPVit en mode TEST.</p>
            <label>Offre</label>
            <select name="offer_key">{offer_options}</select>
            <label>Methode de paiement</label>
            <select name="method"><option>Airtel Money</option><option>Moov Money</option><option>MobiCash</option></select>
            <label>Numero telephone client</label>
            <input name="phone" placeholder="+241..." required>
            <button class="green">Creer le paiement TEST</button>
          </form>
          <div class="card">
            <h2>Etat API PVit</h2>
            <p><b>Mode :</b> {html.escape(PVIT_MODE)}</p>
            <p><b>URL :</b> {html.escape(PVIT_BASE_URL)}</p>
            <p><b>SLUG marchand :</b> {html.escape(PVIT_MERCHANT_SLUG)}</p>
            <p><b>LINK :</b> {html.escape(PVIT_LINK_PATH)}</p>
            <p><b>STATUS :</b> {html.escape(PVIT_STATUS_PATH)}</p>
            <p><b>QR CODE :</b> {html.escape(PVIT_QR_PATH)}</p>
            <p><b>OPERATORS :</b> {html.escape(PVIT_OPERATORS_PATH)}</p>
            <p class="alert">{html.escape(config_message)}</p>
            <p class="muted">Il manque encore la cle secrete X-Secret et le compte d'operation TEST pour envoyer une vraie requete PVit complete.</p>
          </div>
        </div>
        <div class="grid3" style="margin-top:16px">
          <div class="card"><span class="pill">1 credit</span><h3>1 PDF</h3><p>Generation simple avec cartouche.</p></div>
          <div class="card"><span class="pill">5 credits</span><h3>Analyse cartouche</h3><p>Import et analyse IA d'une cartouche personnalisee reutilisable.</p></div>
          <div class="card"><span class="pill">Abonnement</span><h3>Credits inclus</h3><p>Les packs couvrent OpenAI, stockage, PVit, support et marge.</p></div>
        </div>
        <div class="card" style="margin-top:16px">
          <h2>Workflows prepares</h2>
          <p>1. Creation du lien de paiement via LINK.</p>
          <p>2. Redirection client vers paiement Airtel/Moov.</p>
          <p>3. Reception callback PVit.</p>
          <p>4. Verification statut via CHECK STATUS.</p>
          <p>5. Preparation QR code via GENERATE QR CODE.</p>
          <p>6. Passage en production plus tard apres NIF et validation officielle.</p>
        </div>"""
        self.send_html("Paiement TEST", body)

    def payment_callback(self):
        length = int(self.headers.get("Content-Length", 0))
        raw = self.rfile.read(length).decode("utf-8", errors="replace")
        try:
            data = json.loads(raw) if raw else {}
        except Exception:
            data = urllib.parse.parse_qs(raw)
        reference = str(data.get("reference") or data.get("transaction_ref") or data.get("ref") or "")
        status = str(data.get("status") or data.get("state") or "callback_received")
        with db() as con:
            if reference:
                con.execute("UPDATE payments SET status=?, provider_response=? WHERE transaction_ref=?", (status, json.dumps(data, ensure_ascii=False), reference))
                if status.lower() in ("success", "successful", "paid", "validated", "valide", "validé", "completed", "approved"):
                    row = con.execute("SELECT id FROM payments WHERE transaction_ref=?", (reference,)).fetchone()
                    if row:
                        grant_payment_credits(con, row["id"])
        response = b"OK"
        self.send_response(200)
        self.send_header("Content-Type", "text/plain; charset=utf-8")
        self.send_header("Content-Length", str(len(response)))
        self.end_headers()
        self.wfile.write(response)

    def payment_check(self, payment_id: str):
        user = self.require_login()
        if not user:
            return
        try:
            pid = int(urllib.parse.unquote(payment_id))
        except Exception:
            return self.send_html("Paiement introuvable", "<div class='card'><h2>Paiement introuvable.</h2></div>", 404)
        with db() as con:
            payment = con.execute("SELECT * FROM payments WHERE id=? AND (user_id=? OR ?='admin')", (pid, user["id"], user["role"])).fetchone()
        if not payment:
            return self.send_html("Paiement introuvable", "<div class='card'><h2>Paiement introuvable.</h2></div>", 404)
        configured, missing = pvit_config_status()
        if not configured:
            body = f"""
            <div class="card">
              <h2>Verification statut TEST</h2>
              <p><b>Reference :</b> {html.escape(payment['transaction_ref'] or '')}</p>
              <p class="alert">Impossible d'interroger PVit pour le moment : {html.escape(', '.join(missing))}</p>
              <a class="btn" href="/dashboard">Retour</a>
            </div>"""
            return self.send_html("Statut paiement", body)
        payload = {"merchant_slug": PVIT_MERCHANT_SLUG, "reference": payment["transaction_ref"], "mode": PVIT_MODE}
        try:
            raw = pvit_request(PVIT_STATUS_PATH, payload)
            status = str(raw.get("status") or raw.get("state") or payment["status"])
            with db() as con:
                con.execute("UPDATE payments SET status=?, provider_response=? WHERE id=?", (status, json.dumps(raw, ensure_ascii=False), pid))
                if status.lower() in ("success", "successful", "paid", "validated", "valide", "validé", "completed", "approved"):
                    grant_payment_credits(con, pid)
            body = f"<div class='card'><h2>Statut PVit</h2><p>Reference : {html.escape(payment['transaction_ref'] or '')}</p><p class='alert'>Statut : {html.escape(status)}</p><pre style='white-space:pre-wrap'>{html.escape(json.dumps(raw, indent=2, ensure_ascii=False))}</pre></div>"
        except Exception as exc:
            body = f"<div class='card'><h2>Erreur statut PVit</h2><p class='alert'>{html.escape(str(exc))}</p></div>"
        self.send_html("Statut paiement", body)

    def payment_qr(self, payment_id: str):
        user = self.require_login()
        if not user:
            return
        try:
            pid = int(urllib.parse.unquote(payment_id))
        except Exception:
            return self.send_html("Paiement introuvable", "<div class='card'><h2>Paiement introuvable.</h2></div>", 404)
        with db() as con:
            payment = con.execute("SELECT * FROM payments WHERE id=? AND (user_id=? OR ?='admin')", (pid, user["id"], user["role"])).fetchone()
        if not payment:
            return self.send_html("Paiement introuvable", "<div class='card'><h2>Paiement introuvable.</h2></div>", 404)
        configured, missing = pvit_config_status()
        if not configured:
            return self.send_html("QR Code TEST", f"<div class='card'><h2>QR Code TEST</h2><p class='alert'>Configuration incomplete : {html.escape(', '.join(missing))}</p></div>")
        payload = {"merchant_slug": PVIT_MERCHANT_SLUG, "reference": payment["transaction_ref"], "amount": payment["amount"], "mode": PVIT_MODE}
        try:
            raw = pvit_request(PVIT_QR_PATH, payload)
            body = f"<div class='card'><h2>QR Code PVit</h2><p>Reference : {html.escape(payment['transaction_ref'] or '')}</p><pre style='white-space:pre-wrap'>{html.escape(json.dumps(raw, indent=2, ensure_ascii=False))}</pre></div>"
        except Exception as exc:
            body = f"<div class='card'><h2>Erreur QR Code PVit</h2><p class='alert'>{html.escape(str(exc))}</p></div>"
        self.send_html("QR Code paiement", body)

    def payment_result(self, result: str):
        title = "Paiement reussi" if result == "success" else "Paiement echoue"
        message = "Le paiement a ete valide ou est en cours de confirmation." if result == "success" else "Le paiement n'a pas ete valide. Vous pouvez refaire un essai."
        self.send_html(title, f"<div class='card'><h2>{title}</h2><p>{message}</p><a class='btn' href='/dashboard'>Retour au tableau de bord</a></div>")

    def pvit_secret_receiver(self):
        parsed = urllib.parse.urlparse(self.path)
        params = urllib.parse.parse_qs(parsed.query)
        token = ((params.get("token") or [""])[0] or "").strip()
        token_ok = token in PVIT_ACCEPTED_RECEIVER_TOKENS
        if PVIT_MODE.upper() == "TEST" and (not token or parsed.path.endswith("-open")):
            token_ok = True
        method = self.command.upper()
        length = int(self.headers.get("Content-Length", 0))
        raw = self.rfile.read(length).decode("utf-8", errors="replace") if length else ""
        append_pvit_receiver_log(
            f"{method} {self.path} token_ok={token_ok} "
            f"token_received={token} token_expected={PVIT_SECRET_RECEIVER_TOKEN} "
            f"length={length} ip={self.client_address[0]} body={raw[:1000]}"
        )
        save_setting("mypvit_secret_receiver_last_method", method)
        save_setting("mypvit_secret_receiver_last_path", self.path)
        save_setting("mypvit_secret_receiver_last_headers", str(dict(self.headers)))
        if not token_ok:
            save_setting("mypvit_secret_receiver_last_error", "Token invalide ou absent")
            save_setting("mypvit_secret_receiver_last_at", now())
            save_setting("mypvit_secret_receiver_last_payload", raw)
            response = b"FORBIDDEN"
            self.send_response(403)
            self.send_header("Content-Type", "text/plain; charset=utf-8")
            self.send_header("Content-Length", str(len(response)))
            self.end_headers()
            self.wfile.write(response)
            return
        data = {k: v[0] if v else "" for k, v in params.items()}
        if raw:
            try:
                data.update(json.loads(raw))
            except Exception:
                data.update({k: v[0] if v else "" for k, v in urllib.parse.parse_qs(raw).items()})
        secret = extract_secret_from_payload(data, raw + "&" + parsed.query)
        save_setting("mypvit_secret_receiver_last_payload", raw)
        save_setting("mypvit_secret_receiver_last_at", now())
        save_setting("mypvit_secret_receiver_last_error", "" if secret else "Requete recue mais aucune cle detectee dans le payload")
        if secret:
            save_setting("mypvit_secret_received", secret)
            save_setting("mypvit_secret_received_masked", secret[:4] + "..." + secret[-4:] if len(secret) > 8 else "recu")
        response = json.dumps({"ok": True, "method": method, "secret_received": bool(secret)}).encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(response)))
        self.end_headers()
        self.wfile.write(response)

    def pvit_secret_status(self):
        received = get_setting("mypvit_secret_received_masked", "")
        last_at = get_setting("mypvit_secret_receiver_last_at", "Aucune reception")
        raw = get_setting("mypvit_secret_receiver_last_payload", "")
        method = get_setting("mypvit_secret_receiver_last_method", "")
        last_path = get_setting("mypvit_secret_receiver_last_path", "")
        last_error = get_setting("mypvit_secret_receiver_last_error", "")
        headers = get_setting("mypvit_secret_receiver_last_headers", "")
        log_path = DATA / "pvit_secret_receiver.log"
        log_tail = ""
        if log_path.exists():
            try:
                log_tail = "".join(log_path.read_text(encoding="utf-8", errors="replace").splitlines(True)[-25:])
            except Exception:
                log_tail = ""
        receiver_url = f"{PUBLIC_SITE_URL}/pvit/secret-receiver?token={PVIT_SECRET_RECEIVER_TOKEN}"
        body = f"""
        <div class="card">
          <h2>Reception Secret Key PVit</h2>
          <p>URL a configurer dans PVit quand le site sera public :</p>
          <p class="alert">{html.escape(receiver_url)}</p>
          <p>URL alternative TEST si PVit n'accepte pas les parametres dans l'URL :</p>
          <p class="alert">{html.escape(PUBLIC_SITE_URL + '/pvit/secret-receiver-open')}</p>
          <p><b>Derniere reception :</b> {html.escape(last_at)}</p>
          <p><b>Methode :</b> {html.escape(method or 'Aucune')}</p>
          <p><b>Chemin recu :</b> {html.escape(last_path or 'Aucun')}</p>
          <p><b>Secret recu :</b> {html.escape(received or 'Non recu')}</p>
          <p><b>Derniere erreur :</b> {html.escape(last_error or 'Aucune')}</p>
          <p><b>Token attendu :</b> <code>{html.escape(PVIT_SECRET_RECEIVER_TOKEN)}</code></p>
          <p><b>Tokens acceptes :</b> <code>{html.escape(', '.join(sorted(PVIT_ACCEPTED_RECEIVER_TOKENS)))}</code></p>
          <p class="muted">La valeur complete est stockee en base locale dans les settings. Elle n'est pas affichee en clair dans l'interface.</p>
        </div>
        <div class="card" style="margin-top:16px">
          <h2>Dernier payload recu</h2>
          <pre style="white-space:pre-wrap;background:#020617;color:#dbeafe;padding:16px;border-radius:12px;overflow:auto">{html.escape(raw or 'Aucun payload recu.')}</pre>
        </div>
        <div class="card" style="margin-top:16px">
          <h2>Derniers headers recus</h2>
          <pre style="white-space:pre-wrap;background:#020617;color:#dbeafe;padding:16px;border-radius:12px;overflow:auto">{html.escape(headers or 'Aucun header recu.')}</pre>
        </div>
        <div class="card" style="margin-top:16px">
          <h2>Journal receiver</h2>
          <pre style="white-space:pre-wrap;background:#020617;color:#dbeafe;padding:16px;border-radius:12px;overflow:auto">{html.escape(log_tail or 'Aucun log local.')}</pre>
        </div>"""
        self.send_html("Secret PVit", body)

    def admin(self):
        user = self.require_login()
        if not user:
            return
        if user["role"] != "admin":
            return self.send_html("Interdit", "<div class='card'><h2>Acces admin reserve.</h2></div>", 403)
        with db() as con:
            stats = {
                "users": con.execute("SELECT COUNT(*) c FROM users").fetchone()["c"],
                "gens": con.execute("SELECT COUNT(*) c FROM generations").fetchone()["c"],
                "payments": con.execute("SELECT COUNT(*) c FROM payments WHERE status='pending'").fetchone()["c"],
                "annual": con.execute("SELECT COUNT(*) c FROM users WHERE subscription='annual'").fetchone()["c"],
            }
            users = con.execute("SELECT * FROM users ORDER BY id DESC").fetchall()
            gens = con.execute("SELECT g.*,u.email FROM generations g JOIN users u ON u.id=g.user_id ORDER BY g.id DESC LIMIT 8").fetchall()
            pays = con.execute("SELECT p.*,u.email FROM payments p JOIN users u ON u.id=p.user_id ORDER BY p.id DESC LIMIT 8").fetchall()
        user_rows = "".join(f"<tr><td>{u['email']}</td><td>{u['role']}</td><td>{u['credits']}</td><td>{u['subscription']}</td></tr>" for u in users)
        gen_rows = "".join(f"<tr><td>{g['created_at']}</td><td>{g['email']}</td><td>{g['plan_number'] or ''}</td><td>{g['format_plan']}</td><td>{g['project']}</td></tr>" for g in gens) or "<tr><td colspan='5'>Aucune generation.</td></tr>"
        pay_rows = "".join(f"<tr><td>{p['created_at']}</td><td>{p['email']}</td><td>{p['offer']}</td><td>{p['amount']}</td><td>{p['method']}</td><td>{p['transaction_ref'] or ''}</td><td>{p['status']}</td><td>{p['credits'] if 'credits' in p.keys() else ''}</td><td><a href='/payment/check/{p['id']}'>Statut</a> | <a href='/payment/qr/{p['id']}'>QR</a> | <a href='/admin/validate-payment?id={p['id']}'>Valider</a></td></tr>" for p in pays) or "<tr><td colspan='9'>Aucun paiement.</td></tr>"
        body = f"""
        <h2>Espace administrateur personnel</h2>
        <div class="grid3"><div class="stat"><b>{stats['users']}</b>Utilisateurs</div><div class="stat"><b>{stats['gens']}</b>PDF generes</div><div class="stat"><b>{stats['payments']}</b>Paiements attente</div><div class="stat"><b>{stats['annual']}</b>Abonn. annuels</div></div>
        <div class="card" style="margin-top:16px"><h3>PVit TEST</h3><p>Reception Secret Key et verification technique.</p><a class="btn" href="/pvit/secret-status">Voir reception Secret Key</a><a class="btn dark" href="/payment" style="margin-left:8px">Tester paiement</a></div>
        <div class="grid" style="margin-top:16px"><div class="card"><h3>Utilisateurs</h3><table><tr><th>Email</th><th>Role</th><th>Credits</th><th>Abonnement</th></tr>{user_rows}</table></div><div class="card"><h3>Generations</h3><table><tr><th>Date</th><th>Email</th><th>N plan</th><th>Format</th><th>Projet</th></tr>{gen_rows}</table></div></div>
        <div class="card" style="margin-top:16px"><h3>Paiements PVit TEST</h3><table><tr><th>Date</th><th>Email</th><th>Offre</th><th>Montant</th><th>Methode</th><th>Reference</th><th>Statut</th><th>Credits</th><th>Actions</th></tr>{pay_rows}</table></div>"""
        self.send_html("Admin", body)

    def services(self):
        return self.redirect("/generator")

    def validate_payment(self):
        user = self.require_login()
        if not user:
            return
        if user["role"] != "admin":
            return self.send_html("Interdit", "<div class='card'><h2>Acces admin reserve.</h2></div>", 403)
        parsed = urllib.parse.urlparse(self.path)
        params = urllib.parse.parse_qs(parsed.query)
        try:
            payment_id = int((params.get("id") or ["0"])[0])
        except Exception:
            payment_id = 0
        with db() as con:
            if payment_id:
                con.execute("UPDATE payments SET status='validated_admin' WHERE id=?", (payment_id,))
                grant_payment_credits(con, payment_id)
        self.redirect("/admin")

    def download(self, name: str):
        safe = Path(urllib.parse.unquote(name)).name
        path = OUTPUTS / safe
        if not path.exists():
            candidates = [candidate for candidate in DATA.rglob(safe) if candidate.is_file()]
            if candidates:
                path = candidates[0]
            else:
                self.send_html("Introuvable", "<div class='card'><h2>Fichier introuvable.</h2><p>Le PDF a peut-etre ete deplace ou supprime. Relance la generation depuis le generateur.</p></div>", 404)
                return
        data = path.read_bytes()
        self.send_response(200)
        content_type = "application/zip" if path.suffix.lower() == ".zip" else "application/pdf"
        self.send_header("Content-Type", content_type)
        ascii_name = safe.encode("ascii", "ignore").decode("ascii") or "BTP_document.pdf"
        self.send_header("Content-Disposition", f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{urllib.parse.quote(safe)}")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)


def support_answer(question: str) -> str:
    q = (question or "").strip()
    low = q.lower()
    contact = "<p class='muted'>Support fondateur : sessouedem15@gmail.com | +241 74 15 37 16 | +241 65 28 05 25</p>"
    if not q:
        return "<p>Posez votre question, par exemple : comment creer une cartouche, comment ajouter un logo, ou retrouver mes PDF generes.</p>" + contact

    if any(word in low for word in ["temps", "duree", "durée", "minute", "rapide", "combien"]):
        answer = """
        <p><b>Duree de generation :</b> pour un seul PDF, la generation prend generalement quelques secondes a quelques minutes selon la taille du fichier.</p>
        <p>Pour plusieurs PDF, le temps depend du nombre de plans, du poids des fichiers et du traitement Batch PDF.</p>
        <p>Si un fichier est tres lourd, il faut attendre la fin du traitement avant de telecharger le resultat.</p>
        """
    elif any(word in low for word in ["import", "importer", "charger", "upload", "pdf", "dwg", "fichier"]):
        answer = """
        <p><b>Importer un fichier :</b> ouvrez le generateur, utilisez le champ de chargement PDF, puis selectionnez le plan a traiter.</p>
        <p>Le site utilise le vrai PDF importe pour preparer l'apercu et generer le document final avec cartouche.</p>
        <p>Le format conseille est le PDF. Pour un dessin DWG/DXF, exportez d'abord le plan en PDF depuis AutoCAD, puis importez ce PDF dans BTP Smart Tools.</p>
        """
    elif any(word in low for word in ["telecharger", "télécharger", "download", "recuperer", "récupérer", "sortie", "resultat", "résultat"]):
        answer = """
        <p><b>Telechargement :</b> apres la generation, le site ouvre automatiquement le fichier final ou propose le fichier a telecharger.</p>
        <p>Si vous traitez un seul PDF, vous obtenez un PDF final. Si vous traitez plusieurs PDF avec Batch PDF, vous obtenez un fichier ZIP.</p>
        <p>Les PDF et ZIP generes restent aussi dans le tableau de bord du compte utilisateur.</p>
        """
    elif any(word in low for word in ["cartouche", "generer", "générer", "pdf final", "creer", "créer"]):
        answer = """
        <p><b>Pour creer une cartouche :</b></p>
        <p>1. Ouvrez le generateur.</p>
        <p>2. Chargez le PDF du plan.</p>
        <p>3. Remplissez les informations du projet, de l'entreprise, du format et du modele.</p>
        <p>4. Verifiez l'apercu, puis lancez la generation finale.</p>
        <p>Chaque PDF final genere consomme 1 credit et reste disponible dans l'historique du compte.</p>
        """
    elif any(word in low for word in ["batch", "plusieurs", "zip", "lot", "multi"]):
        answer = """
        <p><b>Pour traiter plusieurs PDF :</b> utilisez Batch PDF.</p>
        <p>Vous pouvez charger plusieurs PDF ou un fichier ZIP. Le systeme applique les memes parametres et prepare un ZIP final avec tous les PDF termines.</p>
        <p>Le cout est de 1 credit par PDF traite. Le ZIP final reste dans l'historique du compte.</p>
        """
    elif any(word in low for word in ["logo", "societe", "société", "entreprise"]):
        answer = """
        <p><b>Pour ajouter le logo :</b> dans le generateur, utilisez le champ logo de l'entreprise.</p>
        <p>Le logo est ajuste automatiquement dans la cartouche. Utilisez une image nette, de preference PNG ou JPG.</p>
        <p>Vous pouvez aussi renseigner le nom de l'entreprise, le telephone, le mail, le chef de projet, l'operateur et les informations du plan.</p>
        """
    elif any(word in low for word in ["modele", "modèle", "personnalise", "personnalisé", "template"]):
        answer = """
        <p><b>Pour les modeles :</b> vous pouvez choisir un modele BTP Smart Tools ou importer une cartouche personnalisee comme reference.</p>
        <p>Une cartouche personnalisee coute 5 credits pour l'analyse une seule fois. Ensuite, elle reste dans le compte et peut etre reutilisee pour plusieurs PDF sans refaire l'analyse.</p>
        """
    elif any(word in low for word in ["format", "a4", "a3", "a2", "a1", "a0", "portrait", "paysage"]):
        answer = """
        <p><b>Formats disponibles :</b> A4, A3, A2, A1 et A0, en portrait ou paysage.</p>
        <p>Le systeme ajuste la mise en page selon le format choisi pour garder un rendu propre et lisible.</p>
        <p>Si vous ne savez pas quel format choisir, commencez par le format du plan original ou utilisez A4/A3 pour un test rapide.</p>
        """
    elif any(word in low for word in ["paiement", "airtel", "moov", "abonnement", "credit", "crédit"]):
        answer = """
        <p><b>Paiement et credits :</b> la plateforme fonctionne avec un solde de credits.</p>
        <p>1 PDF final genere = 1 credit. 1 analyse de cartouche personnalisee = 5 credits une seule fois. Le telechargement d'un PDF deja genere ne consomme pas de nouveau credit.</p>
        <p>Si le solde arrive a 0, le compte reste accessible, mais il faut acheter un pack ou renouveler l'abonnement pour generer de nouveaux PDF.</p>
        <p>Offres : 1 200 FCFA pour 1 credit, 5 000 FCFA pour 10 credits, 10 000 FCFA pour 20 credits, 12 000 FCFA/mois pour 30 credits, 108 000 FCFA/an pour 450 credits, 250 000 FCFA/an pour 1500 credits entreprise.</p>
        """
    elif any(word in low for word in ["historique", "compte", "retrouver", "sauvegarde", "stocke", "stockage", "mes pdf", "archive"]):
        answer = """
        <p><b>Historique du compte :</b> les PDF generes, les modeles importes et les ZIP Batch PDF restent rattaches au compte utilisateur.</p>
        <p>Le client peut revenir plus tard dans son tableau de bord pour telecharger a nouveau ses fichiers.</p>
        <p>Il faut eviter de supprimer les fichiers du stockage serveur, sinon les anciens liens de telechargement ne seront plus disponibles.</p>
        """
    elif any(word in low for word in ["admin", "administrateur", "connexion", "mot de passe"]):
        answer = """
        <p><b>Administration :</b> l'espace admin permet de suivre les utilisateurs, les generations, les abonnements, les paiements et les fichiers.</p>
        <p>Le fondateur peut tester les fonctions gratuitement depuis son compte administrateur.</p>
        """
    elif any(word in low for word in ["fonctionne", "fonctionnement", "comment ca marche", "comment ça marche", "site", "outil", "outils"]):
        answer = """
        <p><b>Fonctionnement du site :</b> BTP Smart Tools aide a preparer des plans PDF professionnels avec cadres, cartouches, logos, legendes, tableaux et informations projet.</p>
        <p>L'utilisateur charge son plan, remplit les informations, choisit le modele, verifie l'apercu puis genere le PDF final.</p>
        <p>Les traitements se font dans les pages Cartouche, Modeles cartouche et Batch PDF. L'assistant sert a expliquer et guider.</p>
        """
    else:
        answer = """
        <p>Je peux vous aider sur : creation de cartouche, import PDF, logo, formats, modeles, Batch PDF, paiement, credits, historique du compte ou espace admin.</p>
        <p>Pour une question precise, indiquez simplement ce que vous voulez faire dans le site.</p>
        """
    return f"<p class='alert'>Question recue : {html.escape(q)}</p>{answer}{contact}"


def parse_prompt(prompt: str) -> dict:
    values = {}
    patterns = {
        "project": r"(?:projet|chantier)\s+([^,.;]+)",
        "company": r"(?:entreprise|societe|sociÃ©tÃ©|maitre d'ouvrage|maÃ®tre d'ouvrage)\s+([^,.;]+)",
        "scale": r"(?:echelle|Ã©chelle)\s+([^,.;]+)",
        "author": r"(?:auteur|dessinateur|operateur|opÃ©rateur)\s+([^,.;]+)",
    }
    for key, pattern in patterns.items():
        m = re.search(pattern, prompt, flags=re.I)
        if m:
            values[key] = m.group(1).strip()
    low = prompt.lower()
    plan_keywords = [
        (("ferraillage", "armature"), "Plan de ferraillage"),
        (("coffrage",), "Plan de coffrage"),
        (("profil en travers", "travers"), "Profil en travers"),
        (("profil en long",), "Profil en long"),
        (("vrd", "reseau", "rÃ©seau"), "Plan VRD"),
        (("topographique", "topo", "implantation"), "Plan topographique"),
        (("fondation", "semelle"), "Plan de fondations"),
        (("situation",), "Plan de situation"),
        (("masse",), "Plan de masse"),
        (("route", "chaussee", "chaussÃ©e"), "Plan de route"),
        (("dallot", "dalot"), "Plan de dallot"),
        (("bÃ©ton", "beton", "radier", "voile"), "Plan beton"),
    ]
    for keywords, plan_type in plan_keywords:
        if any(keyword in low for keyword in keywords):
            values["plan_type"] = plan_type
            break
    if "scale" not in values:
        match = re.search(r"\b1\s*[/:\-\s]\s*(20|25|50|75|100|200|250|500|1000|2000|5000)\b", low)
        if match:
            values["scale"] = f"1/{match.group(1)}"
    values.setdefault("number", "BTP-001")
    return values


def safe_file(text: str) -> str:
    return "_".join("".join(ch if ch.isalnum() else "_" for ch in text).split("_")) or "plan"


def collect_batch_pdfs(form: cgi.FieldStorage, batch_dir: Path) -> list[Path]:
    batch_dir.mkdir(parents=True, exist_ok=True)
    pdfs: list[Path] = []
    items = form["files"] if "files" in form else []
    if not isinstance(items, list):
        items = [items]
    for item in items:
        if not getattr(item, "filename", ""):
            continue
        name = Path(item.filename).name
        saved = batch_dir / f"{int(time.time() * 1000)}_{name}"
        with saved.open("wb") as f:
            shutil.copyfileobj(item.file, f)
        if saved.suffix.lower() == ".zip":
            extract_dir = batch_dir / f"extract_{saved.stem}"
            extract_dir.mkdir(parents=True, exist_ok=True)
            with zipfile.ZipFile(saved, "r") as zf:
                for member in zf.namelist():
                    if member.lower().endswith(".pdf"):
                        target = extract_dir / Path(member).name
                        with target.open("wb") as f:
                            f.write(zf.read(member))
                        pdfs.append(target)
        elif saved.suffix.lower() == ".pdf":
            pdfs.append(saved)
    return pdfs


def run():
    init_db()
    server = ThreadingHTTPServer((HOST, PORT), App)
    print(f"SKE site reel lance : http://{HOST}:{PORT}")
    server.serve_forever()


if __name__ == "__main__":
    run()
